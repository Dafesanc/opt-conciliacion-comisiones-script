import os
import re
import math
import time
import warnings
import unicodedata
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any, Union
from functools import lru_cache
from collections import defaultdict

import numpy as np
import pandas as pd
from dateutil.relativedelta import relativedelta

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment, PatternFill

from scipy.optimize import linear_sum_assignment

try:
    from python_calamine import CalamineWorkbook
    _HAS_CALAMINE = True
except ImportError:
    _HAS_CALAMINE = False

warnings.filterwarnings("ignore")

# ============================================================
# A) CONFIGURACIÓN
# ============================================================

class Config:
    CARPETA = os.environ.get("RPE_LOOK", r"C:/Users/fbazurto/Desktop/Rodri - copia/Rodri/AUTOMATIZACION RPE/LOOK")
    CARPETA_RESULTADOS = os.environ.get("RPE_RESULTADOS", r"C:/Users/fbazurto/Desktop/Rodri - copia/Rodri/AUTOMATIZACION RPE/RESULTADOS")
    
    NOMBRE_VENTAS = "VENTAS.xlsx"
    NOMBRE_COMISIONES = "COMISIONES.xlsx"
    NOMBRE_SLA = "MAESTRO SLA EMISION.xlsx"
    SHEET_COMISIONES = "Base"
    SHEET_VENTAS = "BASE"
    SHEET_SLA = "BASE"  
    
    ACTUALIZAR_COMISIONES = True
    ACTUALIZAR_VENTAS = True
    MOSTRAR_PROGRESO = True
    PROGRESO_INTERVALO = 10000
    
    COL_VENTAS_CEDULA = "B"
    COL_VENTAS_PRIMA = "AC"
    COL_VENTAS_INGRESO = "M"
    COL_VENTAS_RAMO = "U"
    COL_START_BN = "BN"
    
    TOL_PRIMA = 0.03
    TOL_ABS = 10
    MAX_MESES = 11
    
    TOL_CANCEL_REL = 0.005
    TOL_CANCEL_ABS = 2
    
    HABILITAR_ASISTENCIA = True
    MAX_DIAS_ASISTENCIA = 10
    ASISTENCIA_MAX_FRAC = 0.35
    MAIN_TOP = 40
    OTHER_TOP = 200
    
    USAR_LSAP = True
    LSAP_MAX_FILAS = 500
    LSAP_MAX_CANDIDATOS = 2000
    DUMMY_UNMATCH_COST = 1e12
    IMPOSSIBLE_COST = 1e15
    
    HABILITAR_PRIORIDAD_ANTIGUEDAD = True
    PRIORIDAD_ANTIGUEDAD_COST = 1e4
    
    HABILITAR_AJUSTE_POSTERIOR = True
    AJUSTE_POS_MAX_REL = 0.10
    AJUSTE_POS_MAX_ABS = 150
    AJUSTE_NEG_MAX_DIAS = 45
    AJUSTE_POS_TOP = 60
    
    HABILITAR_CANCELADO_POSTERIOR = True
    CANCELADO_POST_MAX_DIAS = 365
    
    HABILITAR_EMISION_ANTES_INGRESO_1D = True
    EMISION_ANTES_INGRESO_MAX_DIAS = 1
    
    EXC_01_CODE = "EXC_01_AJUSTE_POSTERIOR_PARCIAL"
    EXC_02_CODE = "EXC_02_GRUPO_REPARTO_UNICO"
    EXC_03_CODE = "EXC_03_CANCELACION_POSTERIOR_TOTAL"
    EXC_04_CODE = "EXC_04_EMISION_ANTES_INGRESO_1D"
    
    EPS_ZERO = 1e-9

# ============================================================
# B) HELPERS OPTIMIZADOS
# ============================================================

def msg(*args) -> None:
    print("".join(str(a) for a in args), flush=True)

def normalize_name(x: Any) -> str:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ""
    s = str(x).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def is_blank(x: Any) -> bool:
    if x is None or x is pd.NaT:
        return True
    if isinstance(x, float) and math.isnan(x):
        return True
    return str(x).strip() == ""

@lru_cache(maxsize=50000)
def excel_to_timestamp(n: float) -> pd.Timestamp:
    try:
        days = int(math.floor(float(n)))
        return pd.Timestamp(date(1899, 12, 30) + timedelta(days=days))
    except:
        return pd.NaT

def parse_fecha_fast(x: Any) -> pd.Timestamp:
    if x is None or x is pd.NaT:
        return pd.NaT
    if isinstance(x, (datetime, pd.Timestamp, date)):
        return pd.Timestamp(x).normalize()
    if isinstance(x, (int, float, np.integer, np.floating)):
        v = float(x)
        if math.isnan(v):
            return pd.NaT
        if 20000 < v < 60000:
            return excel_to_timestamp(v)
        if v >= 1e7:
            try:
                return pd.Timestamp(datetime.strptime(f"{int(v):08d}", "%Y%m%d"))
            except:
                return pd.NaT
        return pd.NaT
    s = str(x).strip()
    if not s:
        return pd.NaT
    try:
        v = float(s)
        if 20000 < v < 60000:
            return excel_to_timestamp(v)
    except:
        pass
    if re.fullmatch(r"\d{8}", s):
        try:
            return pd.Timestamp(datetime.strptime(s, "%Y%m%d"))
        except:
            pass
    try:
        return pd.Timestamp(pd.to_datetime(s, errors="raise")).normalize()
    except:
        return pd.NaT

def parse_fecha_series(series: pd.Series) -> pd.Series:
    if series.empty:
        return pd.Series(dtype='datetime64[ns]')
    return series.apply(parse_fecha_fast)

def to_num_fast(x: Any) -> float:
    if x is None or x is pd.NaT:
        return np.nan
    if isinstance(x, bool):
        return np.nan
    if isinstance(x, (int, float, np.integer, np.floating)):
        v = float(x)
        return v if not math.isnan(v) else np.nan
    s = str(x).strip()
    if not s:
        return np.nan
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^0-9,\.\-]", "", s)
    s = re.sub(r"[.,]+$", "", s)
    if s in ("", "-"):
        return np.nan
    try:
        has_comma = "," in s
        has_dot = "." in s
        if has_comma and has_dot:
            if s.rfind(",") > s.rfind("."):
                t = s.replace(".", "").replace(",", ".", 1)
            else:
                t = s.replace(",", "")
            return float(t)
        elif has_comma and not has_dot:
            if s.count(",") > 1:
                return float(s.replace(",", ""))
            return float(s.replace(",", ".", 1))
        elif has_dot and not has_comma:
            if s.count(".") > 1:
                return float(s.replace(".", ""))
            return float(s)
        else:
            return float(s)
    except:
        return np.nan

def clean_cedula(x: Any) -> Optional[str]:
    if x is None or x is pd.NaT:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float, np.integer, np.floating)):
        s = f"{float(x):.0f}"
    else:
        s = str(x).strip()
        if 'e' in s.lower():
            try:
                s = f"{float(s):.0f}"
            except:
                return None
    s = re.sub(r"[^0-9]", "", s)
    if not s:
        return None
    if len(s) == 13 and s.endswith("001"):
        s = s[:10]
    if len(s) > 10 and s.endswith("001"):
        s = s[:len(s) - 3]
    if len(s) > 10:
        s = s[:10]
    s = s.rjust(10, "0")
    return s if len(s) == 10 else None

def rpe_code(rpe: str) -> Optional[int]:
    m = re.match(r"^\d+", str(rpe))
    return int(m.group(0)) if m else None

def rpe_base(rpe: str) -> Optional[str]:
    m = re.match(r"^\d+-\d+", str(rpe))
    return m.group(0) if m else None

@lru_cache(maxsize=50000)
def add_months_cached(ts: pd.Timestamp, n: int) -> pd.Timestamp:
    if pd.isna(ts):
        return pd.NaT
    return pd.Timestamp(ts) + pd.DateOffset(months=n)

def days_between(a: pd.Timestamp, b: pd.Timestamp) -> int:
    if pd.isna(a) or pd.isna(b):
        return 0
    return (pd.Timestamp(a).normalize() - pd.Timestamp(b).normalize()).days

def fmt_num(x) -> str:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return "NaN"
    return f"{round(float(x), 2):.2f}"

# ============================================================
# C) LECTURA EXCEL OPTIMIZADA
# ============================================================

_CALAMINE_CACHE = {}
_CALAMINE_CACHE_MAX = 5

def read_excel_fast(path: str, sheet: str, n_rows: int = None, max_cols: int = None) -> List[List]:
    if _HAS_CALAMINE:
        try:
            key = (path, sheet, os.path.getmtime(path))
            if key in _CALAMINE_CACHE:
                return _CALAMINE_CACHE[key]
            wb = CalamineWorkbook.from_path(path)
            ws = wb.get_sheet_by_name(sheet)
            data = ws.to_python(skip_empty_area=False)
            data = [list(r) for r in data]
            if len(_CALAMINE_CACHE) >= _CALAMINE_CACHE_MAX:
                _CALAMINE_CACHE.clear()
            _CALAMINE_CACHE[key] = data
            return data
        except:
            pass
    
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        rows = []
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=n_rows if n_rows else None,
                                               max_col=max_cols if max_cols else None,
                                               values_only=True)):
            rows.append(list(row))
        return rows
    finally:
        wb.close()

def read_sheet_to_df(path: str, sheet: str, header_row: int, cols: List[int]) -> pd.DataFrame:
    if _HAS_CALAMINE:
        try:
            data = read_excel_fast(path, sheet)
            if len(data) <= header_row:
                return pd.DataFrame()
            rows = []
            for r in data[header_row:]:
                if any(not is_blank(r[ci]) for ci in cols if ci < len(r)):
                    rows.append([r[ci] if ci < len(r) else None for ci in cols])
            return pd.DataFrame(rows)
        except:
            pass
    
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, min_col=1,
                                max_col=max(cols) + 1, values_only=True):
            if any(not is_blank(row[ci]) for ci in cols if ci < len(row)):
                rows.append([row[ci] if ci < len(row) else None for ci in cols])
        return pd.DataFrame(rows)
    finally:
        wb.close()

# ============================================================
# C.1) LECTURA DE SLA MAESTRO (VERSIÓN CORREGIDA)
# ============================================================

def cargar_sla_maestro() -> pd.DataFrame:
    """
    Carga la tabla maestra de SLA desde el archivo MAESTRO SLA EMISION.xlsx
    usando la BASE que tiene la estructura: ID, PRODUCTO, TIPO DE PRODUCTO, RAMO, SLA
    """
    path = os.path.join(Config.CARPETA, Config.NOMBRE_SLA)
    
    if not os.path.exists(path):
        msg(f"⚠️ Advertencia: No se encontró el archivo {Config.NOMBRE_SLA}")
        return pd.DataFrame()
    
    try:
        # Leer la hoja Hoja2 del archivo SLA (estructura detallada)
        data = read_excel_fast(path, Config.SHEET_SLA)
        
        if not data or len(data) < 2:
            msg(f"⚠️ El archivo {Config.NOMBRE_SLA} no tiene datos suficientes en BASE")
            return pd.DataFrame()
        
        # Obtener headers (primera fila)
        headers = ["" if v is None else str(v).strip() for v in data[0]]
        
        # Identificar columnas por nombre normalizado
        norm_headers = [normalize_name(h) for h in headers]
        
        # Buscar índices de columnas según la estructura de Hoja2
        idx_id = None
        idx_producto = None
        idx_tipo = None
        idx_ramo = None
        idx_sla = None
        
        for i, h in enumerate(norm_headers):
            if h in ['id']:
                idx_id = i
            elif h in ['producto']:
                idx_producto = i
            elif h in ['tipo de producto', 'tipo producto', 'tipo']:
                idx_tipo = i
            elif h in ['ramo']:
                idx_ramo = i
            elif h in ['sla de emision (dias)', 'sla de emisión (dias)', 'sla', 'sla emision', 'sla_emision']:
                idx_sla = i
        
        # Si no encuentra todas las columnas, usar posiciones fijas según la estructura mostrada
        if idx_producto is None:
            idx_producto = 1  # Columna B
        if idx_tipo is None:
            idx_tipo = 2      # Columna C
        if idx_ramo is None:
            idx_ramo = 3      # Columna D
        if idx_sla is None:
            idx_sla = 4       # Columna E
        
        # Extraer datos
        sla_rows = []
        for row in data[1:]:  # Saltar header
            if len(row) <= max(idx_producto, idx_tipo, idx_ramo, idx_sla):
                continue
            
            producto = str(row[idx_producto]).strip() if idx_producto < len(row) and row[idx_producto] is not None else ""
            tipo = str(row[idx_tipo]).strip() if idx_tipo < len(row) and row[idx_tipo] is not None else ""
            ramo = str(row[idx_ramo]).strip() if idx_ramo < len(row) and row[idx_ramo] is not None else ""
            sla = row[idx_sla] if idx_sla < len(row) else None
            
            # Saltar filas vacías
            if not producto and not ramo:
                continue
            
            # Convertir SLA a número
            try:
                sla_num = float(sla) if sla is not None else None
            except (ValueError, TypeError):
                sla_num = None
            
            if sla_num is not None:
                sla_rows.append({
                    'producto': normalize_name(producto),
                    'producto_original': producto,
                    'tipo': normalize_name(tipo) if tipo else '',
                    'tipo_original': tipo,
                    'ramo': normalize_name(ramo) if ramo else '',
                    'ramo_original': ramo,
                    'sla_dias': int(sla_num),
                })
        
        sla_df = pd.DataFrame(sla_rows)
        
        # También leer Hoja1 como respaldo (estructura más simple)
        try:
            data1 = read_excel_fast(path, "Hoja1")
            if data1 and len(data1) > 1:
                headers1 = ["" if v is None else str(v).strip() for v in data1[0]]
                norm1 = [normalize_name(h) for h in headers1]
                
                idx_prod1 = None
                idx_sla1 = None
                
                for i, h in enumerate(norm1):
                    if h in ['producto', 'productos']:
                        idx_prod1 = i
                    elif h in ['sla de emision (dias)', 'sla de emisión (dias)', 'sla', 'sla emision']:
                        idx_sla1 = i
                
                if idx_prod1 is None:
                    idx_prod1 = 0
                if idx_sla1 is None:
                    idx_sla1 = 2
                
                for row in data1[1:]:
                    if len(row) <= max(idx_prod1, idx_sla1):
                        continue
                    
                    producto = str(row[idx_prod1]).strip() if idx_prod1 < len(row) and row[idx_prod1] is not None else ""
                    sla = row[idx_sla1] if idx_sla1 < len(row) else None
                    
                    if not producto:
                        continue
                    
                    try:
                        sla_num = float(sla) if sla is not None else None
                    except (ValueError, TypeError):
                        sla_num = None
                    
                    if sla_num is not None:
                        # Verificar si ya existe en sla_df (evitar duplicados)
                        existing = sla_df[
                            (sla_df['producto'] == normalize_name(producto)) & 
                            (sla_df['ramo'] == '')
                        ]
                        if existing.empty:
                            sla_df = pd.concat([
                                sla_df,
                                pd.DataFrame([{
                                    'producto': normalize_name(producto),
                                    'producto_original': producto,
                                    'tipo': '',
                                    'tipo_original': '',
                                    'ramo': '',
                                    'ramo_original': '',
                                    'sla_dias': int(sla_num),
                                }])
                            ], ignore_index=True)
        except Exception as e:
            # Si falla la lectura de Hoja1, continuar
            pass
        
        msg(f"   ✅ Cargados {len(sla_df)} registros de SLA desde BASE")
        return sla_df
        
    except Exception as e:
        msg(f"⚠️ Error al cargar SLA maestro: {e}")
        return pd.DataFrame()

def get_sla_for_producto_ramo(sla_df: pd.DataFrame, producto: str, ramo: str) -> Optional[int]:
    """
    Obtiene los días de SLA para un producto y ramo específicos.
    Prioriza coincidencias exactas de producto + ramo.
    """
    if sla_df.empty:
        return None
    
    if not producto:
        return None
    
    producto_norm = normalize_name(producto)
    ramo_norm = normalize_name(ramo) if ramo else ''
    
    # 1. Buscar coincidencia exacta: producto + ramo
    if ramo_norm:
        exact_match = sla_df[
            (sla_df['producto'] == producto_norm) & 
            (sla_df['ramo'] == ramo_norm)
        ]
        if not exact_match.empty:
            return exact_match.iloc[0]['sla_dias']
    
    # 2. Buscar coincidencia: producto (sin ramo) y ramo en tipo o en ramo
    if ramo_norm:
        match = sla_df[
            (sla_df['producto'] == producto_norm) & 
            ((sla_df['ramo'] == ramo_norm) | (sla_df['tipo'] == ramo_norm))
        ]
        if not match.empty:
            return match.iloc[0]['sla_dias']
    
    # 3. Buscar solo por producto (cualquier ramo)
    product_match = sla_df[sla_df['producto'] == producto_norm]
    if not product_match.empty:
        # Si hay múltiples, devolver el primero
        return product_match.iloc[0]['sla_dias']
    
    # 4. Buscar coincidencia parcial: producto contiene parte del nombre
    for _, row in sla_df.iterrows():
        if producto_norm in row['producto'] or row['producto'] in producto_norm:
            # Verificar si también coincide el ramo
            if ramo_norm and (ramo_norm in row['ramo'] or ramo_norm in row['tipo']):
                return row['sla_dias']
            # Si no hay ramo específico, devolver el primero que coincida parcialmente
            return row['sla_dias']
    
    # 5. Buscar por ramo solamente (si el producto no está mapeado)
    if ramo_norm:
        ramo_match = sla_df[sla_df['ramo'] == ramo_norm]
        if not ramo_match.empty:
            return ramo_match.iloc[0]['sla_dias']
    
    # 6. Buscar por tipo de producto
    if producto_norm:
        tipo_match = sla_df[sla_df['tipo'] == producto_norm]
        if not tipo_match.empty:
            return tipo_match.iloc[0]['sla_dias']
    
    return None

# ============================================================
# D) COMISIONES OPTIMIZADO
# ============================================================

def build_com_key(com_df: pd.DataFrame) -> pd.DataFrame:
    if com_df.empty:
        return pd.DataFrame()
    
    mask = (com_df['cedula10'].notna() & (com_df['cedula10'] != '') &
            com_df['rpe'].notna() & (com_df['rpe'] != '') &
            com_df['emision'].notna())
    
    if not mask.any():
        return pd.DataFrame()
    
    base = com_df[mask].copy()
    
    # Agrupar optimizado
    rows = []
    for (ced, rpe), g in base.groupby(['cedula10', 'rpe'], sort=False):
        fe = g['emision'].dropna()
        fv = g['fin_vig'].dropna() if 'fin_vig' in g.columns else pd.Series()
        
        # Prima
        prima_vals = g['prima'].dropna().tolist()
        prima_conj = g['prima_conj'].dropna().tolist() if 'prima_conj' in g.columns else []
        
        if prima_conj:
            prima = max(prima_conj, key=abs) if prima_conj else float('nan')
        else:
            prima = sum(prima_vals) if prima_vals else float('nan')
        
        # Comisión
        subcom_vals = g['subcom'].dropna().tolist() if 'subcom' in g.columns else []
        subcom_conj = g['subcom_conj'].dropna().tolist() if 'subcom_conj' in g.columns else []
        
        if subcom_conj:
            comision = max(subcom_conj, key=abs) if subcom_conj else float('nan')
        else:
            comision = sum(subcom_vals) if subcom_vals else float('nan')
        
        rows.append({
            'cedula10': ced,
            'rpe': rpe,
            'emision': fe.min() if len(fe) > 0 else pd.NaT,
            'fin_vig': fv.max() if len(fv) > 0 else pd.NaT,
            'suma_prima': prima,
            'suma_comis': comision,
        })
    
    result = pd.DataFrame(rows)
    if result.empty:
        return pd.DataFrame()
    
    result['rpe_code'] = result['rpe'].apply(rpe_code)
    result['base'] = result['rpe'].apply(rpe_base)
    result['is_negative'] = result['suma_prima'] < 0
    result['is_canceled'] = False
    
    # Detectar cancelaciones (optimizado)
    neg = result[result['is_negative'] & result['base'].notna()]
    pos = result[(~result['is_negative']) & result['base'].notna() & result['suma_prima'].notna()]
    
    if not neg.empty and not pos.empty:
        neg_dict = neg.groupby(['cedula10', 'base']).agg({
            'emision': 'min',
            'suma_prima': lambda x: abs(x.min())
        }).reset_index()
        neg_dict.columns = ['cedula10', 'base', 'em_neg', 'neg_abs']
        
        pos_merge = pos.merge(neg_dict, on=['cedula10', 'base'], how='inner')
        pos_merge = pos_merge[pos_merge['em_neg'] >= pos_merge['emision']]
        
        if not pos_merge.empty:
            pos_merge['cancel_ok'] = (pos_merge['neg_abs'] - pos_merge['suma_prima']).abs() <= \
                                     np.maximum(Config.TOL_CANCEL_ABS, Config.TOL_CANCEL_REL * pos_merge['suma_prima'])
            canc_rpes = set(pos_merge[pos_merge['cancel_ok']]['rpe'])
            result['is_canceled'] = result['rpe'].isin(canc_rpes)
    
    return result

def cargar_comisiones() -> pd.DataFrame:
    path = os.path.join(Config.CARPETA, Config.NOMBRE_COMISIONES)
    top = read_excel_fast(path, Config.SHEET_COMISIONES, n_rows=5)
    headers = ["" if v is None else str(v) for v in (top[0] if top else [])]
    norm = [normalize_name(h) for h in headers]
    
    def find_col(names):
        for nm in names:
            nn = normalize_name(nm)
            for i, h in enumerate(norm):
                if h == nn:
                    return i
        return None
    
    col_cc = find_col(["cc", "Cédula Real", "Cédula"])
    col_rpe = find_col(["Clave R-P-E", "Clave RPE"])
    col_emision = find_col(["F EMISION", "Fecha Emision"])
    col_finv = find_col(["FECHA FIN VIG", "Fin Vig"])
    col_prima = find_col(["Prima"])
    col_subcom = find_col(["Sub Total Comisión"])
    col_prima_conj = find_col(["PRIMA CONJUNTO"])
    col_sub_conj = find_col(["SUBTOTAL CONJUNTO"])
    
    if col_cc is None or col_rpe is None or col_emision is None:
        raise RuntimeError("No se pudieron mapear columnas esenciales de COMISIONES")
    
    cols_needed = [col_cc, col_rpe, col_emision, col_finv, col_prima]
    if col_subcom is not None:
        cols_needed.append(col_subcom)
    if col_prima_conj is not None:
        cols_needed.append(col_prima_conj)
    if col_sub_conj is not None:
        cols_needed.append(col_sub_conj)
    
    df = read_sheet_to_df(path, Config.SHEET_COMISIONES, 1, cols_needed)
    if df.empty:
        return pd.DataFrame()
    
    col_names = ['cedula10', 'rpe', 'emision', 'fin_vig', 'prima']
    if col_subcom is not None:
        col_names.append('subcom')
    if col_prima_conj is not None:
        col_names.append('prima_conj')
    if col_sub_conj is not None:
        col_names.append('subcom_conj')
    
    df.columns = col_names
    df['cedula10'] = df['cedula10'].apply(clean_cedula)
    df['emision'] = parse_fecha_series(df['emision'])
    df['fin_vig'] = parse_fecha_series(df['fin_vig'])
    df['prima'] = df['prima'].apply(to_num_fast)
    if 'subcom' in df.columns:
        df['subcom'] = df['subcom'].apply(to_num_fast)
    if 'prima_conj' in df.columns:
        df['prima_conj'] = df['prima_conj'].apply(to_num_fast)
    if 'subcom_conj' in df.columns:
        df['subcom_conj'] = df['subcom_conj'].apply(to_num_fast)
    
    return df

# ============================================================
# E) VENTAS OPTIMIZADO
# ============================================================

RAMO_MAP = {
    "accidentes personales": 8, "calderas y maquinarias": 11,
    "equipo de contratistas": 5, "incendio": 1,
    "responsabilidad civil": 12, "transporte": 16,
    "vehiculos": 4, "vida termino": 30,
}

def ramo_code_fast(x: Any) -> Optional[int]:
    if x is None or is_blank(x):
        return None
    xn = normalize_name(x)
    if not xn:
        return None
    if xn in RAMO_MAP:
        return RAMO_MAP[xn]
    for key, code in RAMO_MAP.items():
        if key in xn:
            return code
    return None

def detectar_header_ventas() -> int:
    path = os.path.join(Config.CARPETA, Config.NOMBRE_VENTAS)
    for row_num in range(1, 16):
        top = read_excel_fast(path, Config.SHEET_VENTAS, n_rows=row_num + 1)
        if not top or len(top) < row_num:
            continue
        headers = ["" if v is None else str(v) for v in top[row_num-1]]
        norm = [normalize_name(h) for h in headers]
        esenciales = ["Cédula", "Fecha de Ingreso", "Prima Neta", "Producto", "Ramo"]
        encontradas = sum(1 for e in esenciales if any(normalize_name(e) in h for h in norm))
        if encontradas >= 3:
            return row_num
    return 2

def leer_ventas() -> pd.DataFrame:
    hrow = detectar_header_ventas()
    path = os.path.join(Config.CARPETA, Config.NOMBRE_VENTAS)
    top = read_excel_fast(path, Config.SHEET_VENTAS, n_rows=hrow)
    headers = ["" if v is None else str(v) for v in (top[hrow-1] if len(top) >= hrow else [])]
    norm = [normalize_name(h) for h in headers]
    
    def find_col(names):
        for nm in names:
            nn = normalize_name(nm)
            for i, h in enumerate(norm):
                if h == nn:
                    return i
        return None
    
    col_ced = find_col(["10 DIGITOS", "Cédula", "Cedula"]) or column_index_from_string(Config.COL_VENTAS_CEDULA) - 1
    col_pri = find_col(["Prima Neta"]) or column_index_from_string(Config.COL_VENTAS_PRIMA) - 1
    col_ing = find_col(["Fecha de Ingreso", "Fecha Ingreso"]) or column_index_from_string(Config.COL_VENTAS_INGRESO) - 1
    col_ram = find_col(["Ramo"]) or column_index_from_string(Config.COL_VENTAS_RAMO) - 1
    col_prod = find_col(["Producto"])
    
    cols_needed = [col_ced, col_pri, col_ing, col_ram]
    if col_prod is not None:
        cols_needed.append(col_prod)
    
    df = read_sheet_to_df(path, Config.SHEET_VENTAS, hrow, cols_needed)
    if df.empty:
        raise RuntimeError("No se pudieron leer datos de VENTAS")
    
    result = pd.DataFrame({
        'row_excel': df.index + hrow + 1,
        'cedula_raw': df[0],
        'prima_raw': df[1],
        'ingreso_raw': df[2],
        'ramo_raw': df[3],
    })
    
    if col_prod is not None and len(df.columns) > 4:
        result['producto_raw'] = df[4]
    else:
        result['producto_raw'] = None
    
    result['cedula10'] = result['cedula_raw'].apply(clean_cedula)
    result['prima_ventas'] = result['prima_raw'].apply(to_num_fast)
    result['ingreso'] = parse_fecha_series(result['ingreso_raw'])
    result['ramo_code'] = result['ramo_raw'].apply(ramo_code_fast)
    result['producto_norm'] = result['producto_raw'].apply(lambda x: normalize_name(x) if x is not None else '')
    
    # Group key optimizado
    def make_group_key(row):
        if row['cedula10'] and pd.notna(row['ingreso']) and pd.notna(row['ramo_code']) and not math.isnan(row['ramo_code']):
            return f"{row['cedula10']}|{row['ingreso'].strftime('%Y-%m-%d')}|{int(row['ramo_code'])}"
        return f"ROW|{row.name + hrow + 1}"
    
    result['group_key'] = result.apply(make_group_key, axis=1)
    result['group_n'] = result.groupby('group_key')['group_key'].transform('size')
    result['row_id'] = range(1, len(result) + 1)
    
    return result

# ============================================================
# F) MATCH ENGINE OPTIMIZADO
# ============================================================

class MatchEngine:
    def __init__(self, com_key: pd.DataFrame, sla_df: pd.DataFrame):
        self.com_key = com_key
        self.sla_df = sla_df
        self._cache = {}
        
        # Pre-indexar com_key por cédula para búsqueda rápida
        self._com_by_cedula = defaultdict(list)
        for _, row in com_key.iterrows():
            if row['cedula10']:
                self._com_by_cedula[row['cedula10']].append(row)
        
        # Pre-indexar por base para búsqueda rápida
        self._com_by_base = defaultdict(list)
        for _, row in com_key.iterrows():
            if row['base']:
                self._com_by_base[row['base']].append(row)
        
        # Convertir a DataFrames para operaciones vectorizadas
        self._com_df = com_key
    
    def ejecutar_match(self, ven: pd.DataFrame) -> pd.DataFrame:
        resultados = []
        ven_valid = ven[ven['cedula10'].notna() & (ven['cedula10'] != '')]
        total = len(ven_valid)
        
        # Procesar por cédula agrupado
        for idx, (cedula, grupo) in enumerate(ven_valid.groupby('cedula10', sort=False)):
            if Config.MOSTRAR_PROGRESO and idx > 0 and idx % Config.PROGRESO_INTERVALO == 0:
                msg(f"  Procesando: {idx}/{total}")
            
            com_list = self._com_by_cedula.get(cedula, [])
            if com_list:
                com_sub = pd.DataFrame(com_list)
            else:
                com_sub = pd.DataFrame()
            
            resultados.extend(self._match_grupo(grupo, com_sub, cedula))
        
        # Cédulas inválidas
        ven_invalid = ven[ven['cedula10'].isna() | (ven['cedula10'] == '')]
        for _, row in ven_invalid.iterrows():
            resultados.append(self._crear_vacio(row, "Cédula inválida"))
        
        return pd.DataFrame(resultados)
    
    def _match_grupo(self, grupo: pd.DataFrame, com_sub: pd.DataFrame, cedula: str) -> List[Dict]:
        if len(grupo) == 1:
            return [self._match_single(grupo.iloc[0], com_sub, cedula)]
        return self._match_multi(grupo, com_sub, cedula)
    
    def _match_single(self, row: pd.Series, com_sub: pd.DataFrame, cedula: str) -> Dict:
        resultado = self._crear_vacio(row, "")
        
        if pd.isna(row['prima_ventas']) or row['prima_ventas'] <= 0:
            resultado['razon_no_rpe'] = "Prima inválida"
            return resultado
        if pd.isna(row['ingreso']):
            resultado['razon_no_rpe'] = "Fecha de ingreso inválida"
            return resultado
        
        # Buscar match directo
        match, used = self._find_match(row, com_sub)
        if match:
            resultado.update(match)
            if match.get('exc_code'):
                resultado['revision_manual_txt'] = "SI"
            resultado['sla_dias'] = self._get_sla(row)
            return resultado
        
        # Intentar excepciones
        exc_match = self._try_exceptions(row, used)
        if exc_match:
            resultado.update(exc_match)
            resultado['revision_manual_txt'] = "SI"
            resultado['sla_dias'] = self._get_sla(row)
            return resultado
        
        resultado['razon_no_rpe'] = "No se encontró match"
        resultado['sla_dias'] = self._get_sla(row)
        return resultado
    
    def _get_sla(self, row: pd.Series) -> Optional[int]:
        """Obtiene los días de SLA para una fila de ventas usando producto y ramo."""
        producto = row.get('producto_norm', '')
        ramo = row.get('ramo_raw', '')
        return get_sla_for_producto_ramo(self.sla_df, producto, ramo)
    
    def _find_match(self, row: pd.Series, com_sub: pd.DataFrame) -> Tuple[Optional[Dict], set]:
        if com_sub.empty:
            return None, set()
        
        target = row['prima_ventas']
        ingreso = row['ingreso']
        ramo = row['ramo_code']
        lim = add_months_cached(ingreso, Config.MAX_MESES)
        
        # Filtrar candidatos (versión optimizada)
        mask = (~com_sub['is_negative']) & (~com_sub['is_canceled']) & \
               com_sub['suma_prima'].notna() & (com_sub['suma_prima'] > 0) & \
               com_sub['emision'].notna() & (com_sub['emision'] >= ingreso) & \
               (com_sub['emision'] <= lim)
        
        if ramo is not None and not pd.isna(ramo) and not math.isnan(ramo):
            mask = mask & (com_sub['rpe_code'] == ramo)
        
        candidatos = com_sub[mask].copy()
        if candidatos.empty:
            return None, set()
        
        candidatos['abs_diff'] = (candidatos['suma_prima'] - target).abs()
        candidatos['rel_diff'] = candidatos['abs_diff'] / target
        mask_tol = (candidatos['rel_diff'] <= Config.TOL_PRIMA) | (candidatos['abs_diff'] <= Config.TOL_ABS)
        candidatos = candidatos[mask_tol]
        
        if candidatos.empty:
            return None, set()
        
        best = candidatos.sort_values(['abs_diff', 'emision']).iloc[0]
        used = {best['rpe']}
        
        return {
            'RPE1': best['rpe'],
            'RPE2': None,
            'RPE3': None,
            'emision': best['emision'],
            'fin_vig': best['fin_vig'],
            'prima1': best['suma_prima'],
            'prima2': np.nan,
            'prima3': np.nan,
            'comision': best['suma_comis'] if pd.notna(best['suma_comis']) else 0,
            'multi_n': 1,
            'ramo_ok': True,
            'razon_no_rpe': '',
            'exc_code': '',
            'exc_detalle': '',
            'revision_manual': False,
            'revision_manual_txt': 'NO',
            'anomalia': ramo is None or pd.isna(ramo) or math.isnan(ramo),
            'anomalia_detalle': "RAMO no mapeado" if (ramo is None or pd.isna(ramo) or math.isnan(ramo)) else "",
        }, used
    
    def _try_exceptions(self, row: pd.Series, used_rpes: set) -> Optional[Dict]:
        target = row['prima_ventas']
        ingreso = row['ingreso']
        ramo = row['ramo_code']
        
        if pd.isna(ramo) or math.isnan(ramo):
            ramo = None
        
        # EXC_01: Ajuste posterior parcial
        if Config.HABILITAR_AJUSTE_POSTERIOR:
            exc = self._try_ajuste_posterior(target, ingreso, ramo, used_rpes)
            if exc:
                return exc
        
        # EXC_03: Cancelación posterior total
        if Config.HABILITAR_CANCELADO_POSTERIOR:
            exc = self._try_cancelado_posterior(target, ingreso, ramo, used_rpes)
            if exc:
                return exc
        
        # EXC_04: Emisión antes de ingreso
        if Config.HABILITAR_EMISION_ANTES_INGRESO_1D:
            exc = self._try_emision_antes_ingreso(target, ingreso, ramo, used_rpes)
            if exc:
                return exc
        
        return None
    
    def _try_ajuste_posterior(self, target: float, ingreso: pd.Timestamp, ramo: Optional[int], used: set) -> Optional[Dict]:
        if pd.isna(target) or target <= 0 or pd.isna(ingreso) or ramo is None:
            return None
        
        lim = add_months_cached(ingreso, Config.MAX_MESES)
        
        # Buscar positivos (optimizado)
        pos_mask = (~self._com_df['is_negative']) & (~self._com_df['is_canceled']) & \
                   self._com_df['base'].notna() & self._com_df['suma_prima'].notna() & \
                   (self._com_df['suma_prima'] > 0) & self._com_df['emision'].notna() & \
                   (self._com_df['emision'] >= ingreso) & (self._com_df['emision'] <= lim) & \
                   (self._com_df['rpe_code'] == ramo) & (~self._com_df['rpe'].isin(used))
        
        pos = self._com_df[pos_mask].copy()
        if pos.empty:
            return None
        
        pos['absd_target'] = (pos['suma_prima'] - target).abs()
        pos['rel_target'] = pos['absd_target'] / target
        
        near_cap = max(Config.TOL_ABS + 0.01, min(Config.AJUSTE_POS_MAX_ABS, Config.AJUSTE_POS_MAX_REL * target))
        pos = pos[(pos['suma_prima'] > target) &
                  ((pos['rel_target'] > Config.TOL_PRIMA) | (pos['absd_target'] > Config.TOL_ABS)) &
                  (pos['absd_target'] <= near_cap)]
        
        if pos.empty:
            return None
        
        pos = pos.sort_values(['absd_target', 'emision']).head(Config.AJUSTE_POS_TOP)
        
        # Buscar negativos (optimizado)
        neg_mask = self._com_df['is_negative'] & self._com_df['base'].notna() & \
                   self._com_df['suma_prima'].notna() & self._com_df['emision'].notna()
        neg = self._com_df[neg_mask]
        if neg.empty:
            return None
        
        best = None
        best_abs_adj = float('inf')
        best_gap = float('inf')
        
        for _, p in pos.iterrows():
            n_mask = (neg['base'] == p['base']) & \
                     (neg['emision'] >= p['emision']) & \
                     (neg['emision'] <= p['emision'] + pd.Timedelta(days=Config.AJUSTE_NEG_MAX_DIAS)) & \
                     (~neg['rpe'].isin(used))
            
            n1 = neg[n_mask].copy()
            if n1.empty:
                continue
            
            n1['adj_prima'] = p['suma_prima'] + n1['suma_prima']
            n1['absd_adj'] = (n1['adj_prima'] - target).abs()
            n1['rel_adj'] = n1['absd_adj'] / target
            n1['gap_days'] = n1['emision'].apply(lambda e: days_between(e, p['emision']))
            
            n1 = n1[(n1['rel_adj'] <= Config.TOL_PRIMA) | (n1['absd_adj'] <= Config.TOL_ABS)]
            if n1.empty:
                continue
            
            cand = n1.sort_values(['absd_adj', 'gap_days']).iloc[0]
            
            if (cand['absd_adj'] < best_abs_adj or
                (cand['absd_adj'] == best_abs_adj and cand['gap_days'] < best_gap)):
                best = {'pos': p, 'neg': cand}
                best_abs_adj = cand['absd_adj']
                best_gap = cand['gap_days']
        
        if best is None:
            return None
        
        p, n = best['pos'], best['neg']
        detalle = (f"{Config.EXC_01_CODE} | RPE+={p['rpe']} {fmt_num(p['suma_prima'])} | "
                   f"RPE-={n['rpe']} {fmt_num(n['suma_prima'])} | "
                   f"neto={fmt_num(n['adj_prima'])} | target={fmt_num(target)}")
        
        return {
            'RPE1': p['rpe'],
            'RPE2': n['rpe'],
            'RPE3': None,
            'emision': p['emision'],
            'fin_vig': p['fin_vig'],
            'prima1': float(p['suma_prima']),
            'prima2': float(n['suma_prima']),
            'prima3': np.nan,
            'comision': float(p['suma_comis']) + float(n['suma_comis']) if pd.notna(p['suma_comis']) and pd.notna(n['suma_comis']) else 0,
            'multi_n': 2,
            'ramo_ok': True,
            'razon_no_rpe': '',
            'exc_code': Config.EXC_01_CODE,
            'exc_detalle': detalle,
            'revision_manual': True,
            'revision_manual_txt': 'SI',
            'anomalia': False,
            'anomalia_detalle': '',
        }
    
    def _try_cancelado_posterior(self, target: float, ingreso: pd.Timestamp, ramo: Optional[int], used: set) -> Optional[Dict]:
        if pd.isna(target) or target <= 0 or pd.isna(ingreso):
            return None
        
        lim = add_months_cached(ingreso, Config.MAX_MESES)
        
        # Buscar positivos cancelados
        pos_mask = (~self._com_df['is_negative']) & (self._com_df['is_canceled']) & \
                   self._com_df['base'].notna() & self._com_df['suma_prima'].notna() & \
                   (self._com_df['suma_prima'] > 0) & self._com_df['emision'].notna() & \
                   (self._com_df['emision'] >= ingreso) & (self._com_df['emision'] <= lim) & \
                   (~self._com_df['rpe'].isin(used))
        
        if ramo is not None:
            pos_mask = pos_mask & (self._com_df['rpe_code'] == ramo)
        
        pos = self._com_df[pos_mask].copy()
        if pos.empty:
            return None
        
        pos['absd'] = (pos['suma_prima'] - target).abs()
        pos['rel'] = pos['absd'] / target
        pos = pos[(pos['rel'] <= Config.TOL_PRIMA) | (pos['absd'] <= Config.TOL_ABS)]
        
        if pos.empty:
            return None
        
        pos = pos.sort_values(['absd', 'emision'])
        
        # Buscar negativos
        neg_mask = self._com_df['is_negative'] & self._com_df['base'].notna() & \
                   self._com_df['suma_prima'].notna() & self._com_df['emision'].notna()
        neg = self._com_df[neg_mask]
        if neg.empty:
            return None
        
        best = None
        best_abs = float('inf')
        best_gap = float('inf')
        
        for _, p in pos.iterrows():
            n_mask = (neg['base'] == p['base']) & \
                     (neg['emision'] >= p['emision']) & \
                     (neg['emision'] <= p['emision'] + pd.Timedelta(days=Config.CANCELADO_POST_MAX_DIAS)) & \
                     (~neg['rpe'].isin(used))
            
            n1 = neg[n_mask].copy()
            if n1.empty:
                continue
            
            n1['abs_neg'] = (n1['suma_prima'].abs() - p['suma_prima']).abs()
            n1 = n1[n1['abs_neg'] <= np.maximum(Config.TOL_CANCEL_ABS, Config.TOL_CANCEL_REL * p['suma_prima'])]
            if n1.empty:
                continue
            
            n1['gap_days'] = n1['emision'].apply(lambda e: days_between(e, p['emision']))
            cand = n1.sort_values(['gap_days', 'abs_neg']).iloc[0]
            
            if p['absd'] < best_abs or (p['absd'] == best_abs and cand['gap_days'] < best_gap):
                best = {'pos': p, 'neg': cand}
                best_abs = p['absd']
                best_gap = cand['gap_days']
        
        if best is None:
            return None
        
        p, n = best['pos'], best['neg']
        detalle = (f"{Config.EXC_03_CODE} | RPE+={p['rpe']} {fmt_num(p['suma_prima'])} | "
                   f"RPE-={n['rpe']} {fmt_num(n['suma_prima'])} | cancelacion posterior")
        
        return {
            'RPE1': p['rpe'],
            'RPE2': n['rpe'],
            'RPE3': None,
            'emision': p['emision'],
            'fin_vig': p['fin_vig'],
            'prima1': float(p['suma_prima']),
            'prima2': float(n['suma_prima']),
            'prima3': np.nan,
            'comision': float(p['suma_comis']) + float(n['suma_comis']) if pd.notna(p['suma_comis']) and pd.notna(n['suma_comis']) else 0,
            'multi_n': 2,
            'ramo_ok': True,
            'razon_no_rpe': '',
            'exc_code': Config.EXC_03_CODE,
            'exc_detalle': detalle,
            'revision_manual': True,
            'revision_manual_txt': 'SI',
            'anomalia': False,
            'anomalia_detalle': '',
        }
    
    def _try_emision_antes_ingreso(self, target: float, ingreso: pd.Timestamp, ramo: Optional[int], used: set) -> Optional[Dict]:
        if pd.isna(target) or target <= 0 or pd.isna(ingreso) or ramo is None:
            return None
        
        lim = add_months_cached(ingreso, Config.MAX_MESES)
        
        # Buscar emitidos antes del ingreso (1 día)
        mask = (~self._com_df['is_negative']) & (~self._com_df['is_canceled']) & \
               self._com_df['suma_prima'].notna() & (self._com_df['suma_prima'] > 0) & \
               self._com_df['emision'].notna() & (self._com_df['emision'] < ingreso) & \
               (self._com_df['emision'] >= ingreso - pd.Timedelta(days=Config.EMISION_ANTES_INGRESO_MAX_DIAS)) & \
               (self._com_df['emision'] <= lim) & (self._com_df['rpe_code'] == ramo) & \
               (~self._com_df['rpe'].isin(used))
        
        candidatos = self._com_df[mask].copy()
        if candidatos.empty:
            return None
        
        candidatos['absd'] = (candidatos['suma_prima'] - target).abs()
        candidatos['rel'] = candidatos['absd'] / target
        candidatos = candidatos[(candidatos['rel'] <= Config.TOL_PRIMA) | (candidatos['absd'] <= Config.TOL_ABS)]
        
        if candidatos.empty:
            return None
        
        candidatos['dias_antes'] = candidatos['emision'].apply(lambda e: days_between(ingreso, e))
        candidatos = candidatos[candidatos['dias_antes'] == Config.EMISION_ANTES_INGRESO_MAX_DIAS]
        
        if candidatos.empty:
            return None
        
        b = candidatos.sort_values(['absd', 'emision']).iloc[0]
        detalle = (f"{Config.EXC_04_CODE} | RPE={b['rpe']} | "
                   f"emision={pd.Timestamp(b['emision']).strftime('%d/%m/%Y')} | "
                   f"ingreso={pd.Timestamp(ingreso).strftime('%d/%m/%Y')} | "
                   f"dias={days_between(ingreso, b['emision'])} | "
                   f"prima={fmt_num(b['suma_prima'])} | target={fmt_num(target)}")
        
        return {
            'RPE1': b['rpe'],
            'RPE2': None,
            'RPE3': None,
            'emision': b['emision'],
            'fin_vig': b['fin_vig'],
            'prima1': float(b['suma_prima']),
            'prima2': np.nan,
            'prima3': np.nan,
            'comision': float(b['suma_comis']) if pd.notna(b['suma_comis']) else 0,
            'multi_n': 1,
            'ramo_ok': True,
            'razon_no_rpe': '',
            'exc_code': Config.EXC_04_CODE,
            'exc_detalle': detalle,
            'revision_manual': True,
            'revision_manual_txt': 'SI',
            'anomalia': False,
            'anomalia_detalle': '',
        }
    
    def _match_multi(self, grupo: pd.DataFrame, com_sub: pd.DataFrame, cedula: str) -> List[Dict]:
        resultados = []
        
        # Construir edges optimizado
        if com_sub.empty:
            for _, row in grupo.iterrows():
                resultados.append(self._crear_vacio(row, "Sin candidatos"))
            return resultados
        
        # Preparar datos para merge rápido
        grupo_flat = grupo[['row_id', 'prima_ventas', 'ingreso', 'ramo_code']].copy()
        grupo_flat['_key'] = 1
        
        com_flat = com_sub[['rpe', 'suma_prima', 'emision', 'fin_vig', 'suma_comis', 
                            'rpe_code', 'is_negative', 'is_canceled']].copy()
        com_flat = com_flat[(~com_flat['is_negative']) & (~com_flat['is_canceled']) & 
                            com_flat['suma_prima'].notna() & (com_flat['suma_prima'] > 0) &
                            com_flat['emision'].notna()]
        
        if com_flat.empty:
            for _, row in grupo.iterrows():
                resultados.append(self._crear_vacio(row, "Sin candidatos válidos"))
            return resultados
        
        com_flat['_key'] = 1
        edges = grupo_flat.merge(com_flat, on='_key').drop('_key', axis=1)
        
        # Filtrar por fecha y ramo
        lim_series = edges['ingreso'].apply(lambda x: add_months_cached(x, Config.MAX_MESES))
        edges = edges[(edges['emision'] >= edges['ingreso']) & 
                      (edges['emision'] <= lim_series) &
                      (edges['rpe_code'] == edges['ramo_code'])]
        
        if edges.empty:
            for _, row in grupo.iterrows():
                resultados.append(self._crear_vacio(row, "Sin match por fecha/ramo"))
            return resultados
        
        # Calcular diferencias
        edges['abs_diff'] = (edges['suma_prima'] - edges['prima_ventas']).abs()
        edges['rel_diff'] = edges['abs_diff'] / edges['prima_ventas']
        edges = edges[(edges['rel_diff'] <= Config.TOL_PRIMA) | (edges['abs_diff'] <= Config.TOL_ABS)]
        
        if edges.empty:
            for _, row in grupo.iterrows():
                resultados.append(self._crear_vacio(row, "Sin match por prima"))
            return resultados
        
        # Costo
        edges['days_gap'] = (edges['emision'] - edges['ingreso']).dt.days.clip(lower=0)
        edges['cost'] = (edges['abs_diff'] * 1e8 + edges['days_gap'] * 1e2)
        
        # Resolver asignación
        row_ids = grupo['row_id'].tolist()
        asignaciones = self._resolver_asignacion(edges, row_ids)
        used_rpes = set(asignaciones.values())
        
        for _, row in grupo.iterrows():
            row_id = row['row_id']
            if row_id in asignaciones:
                rpe = asignaciones[row_id]
                match_row = edges[edges['rpe'] == rpe].iloc[0]
                res = self._crear_desde_match(row, match_row)
            else:
                exc_match = self._try_exceptions(row, used_rpes)
                if exc_match:
                    res = self._crear_vacio(row, "")
                    res.update(exc_match)
                    res['revision_manual_txt'] = "SI"
                    if exc_match.get('RPE1'):
                        used_rpes.add(exc_match['RPE1'])
                    if exc_match.get('RPE2'):
                        used_rpes.add(exc_match['RPE2'])
                else:
                    res = self._crear_vacio(row, "No asignado")
            res['sla_dias'] = self._get_sla(row)
            resultados.append(res)
        
        return resultados
    
    def _resolver_asignacion(self, edges: pd.DataFrame, row_ids: List[int]) -> Dict[int, str]:
        if edges.empty:
            return {}
        
        if (Config.USAR_LSAP and 
            len(row_ids) <= Config.LSAP_MAX_FILAS and 
            edges['rpe'].nunique() <= Config.LSAP_MAX_CANDIDATOS):
            return self._resolver_lsap(edges, row_ids)
        else:
            return self._resolver_greedy(edges, row_ids)
    
    def _resolver_lsap(self, edges: pd.DataFrame, row_ids: List[int]) -> Dict[int, str]:
        rows = sorted(row_ids)
        rpes = sorted(edges['rpe'].unique())
        nr, nc = len(rows), len(rpes)
        cost = np.full((nr, nc + nr), Config.IMPOSSIBLE_COST, dtype=float)
        cost[:, nc:nc + nr] = Config.DUMMY_UNMATCH_COST
        
        ridx = {r: i for i, r in enumerate(rows)}
        cidx = {c: j for j, c in enumerate(rpes)}
        
        for _, row in edges[['row_id', 'rpe', 'cost']].drop_duplicates().iterrows():
            if row['row_id'] in ridx and row['rpe'] in cidx:
                cost[ridx[row['row_id']], cidx[row['rpe']]] = row['cost']
        
        ri, ci = linear_sum_assignment(cost)
        return {rows[i]: rpes[j] for i, j in zip(ri, ci) if j < nc}
    
    def _resolver_greedy(self, edges: pd.DataFrame, row_ids: List[int]) -> Dict[int, str]:
        edges = edges.sort_values(['cost', 'row_id', 'rpe'])
        asignados = {}
        rpes_usados = set()
        for _, row in edges.iterrows():
            if row['row_id'] not in asignados and row['rpe'] not in rpes_usados:
                asignados[row['row_id']] = row['rpe']
                rpes_usados.add(row['rpe'])
        return asignados
    
    def _crear_vacio(self, row: pd.Series, razon: str) -> Dict:
        return {
            'row_id': row.get('row_id', 0),
            'row_excel': row.get('row_excel', 0),
            'cedula10': row.get('cedula10', ''),
            'cedula_raw': row.get('cedula_raw', ''),
            'prima_ventas': row.get('prima_ventas', np.nan),
            'prima_ventas_raw': row.get('prima_raw', ''),
            'ingreso': row.get('ingreso', pd.NaT),
            'ingreso_raw': row.get('ingreso_raw', ''),
            'ramo_raw': row.get('ramo_raw', ''),
            'ramo_code': row.get('ramo_code', None),
            'producto_norm': row.get('producto_norm', ''),
            'producto_raw': row.get('producto_raw', ''),
            'group_n': row.get('group_n', 1),
            'RPE1': None,
            'RPE2': None,
            'RPE3': None,
            'emision': pd.NaT,
            'fin_vig': pd.NaT,
            'prima1': np.nan,
            'prima2': np.nan,
            'prima3': np.nan,
            'comision': np.nan,
            'multi_n': 0,
            'ramo_ok': False,
            'razon_no_rpe': razon,
            'exc_code': '',
            'exc_detalle': '',
            'revision_manual': False,
            'revision_manual_txt': 'NO',
            'anomalia': False,
            'anomalia_detalle': '',
            'sla_dias': self._get_sla(row),
        }
    
    def _crear_desde_match(self, row: pd.Series, match: pd.Series) -> Dict:
        return {
            'row_id': row.get('row_id', 0),
            'row_excel': row.get('row_excel', 0),
            'cedula10': row.get('cedula10', ''),
            'cedula_raw': row.get('cedula_raw', ''),
            'prima_ventas': row.get('prima_ventas', np.nan),
            'prima_ventas_raw': row.get('prima_raw', ''),
            'ingreso': row.get('ingreso', pd.NaT),
            'ingreso_raw': row.get('ingreso_raw', ''),
            'ramo_raw': row.get('ramo_raw', ''),
            'ramo_code': row.get('ramo_code', None),
            'producto_norm': row.get('producto_norm', ''),
            'producto_raw': row.get('producto_raw', ''),
            'group_n': row.get('group_n', 1),
            'RPE1': match.get('rpe', None),
            'RPE2': None,
            'RPE3': None,
            'emision': match.get('emision', pd.NaT),
            'fin_vig': match.get('fin_vig', pd.NaT),
            'prima1': match.get('suma_prima', np.nan),
            'prima2': np.nan,
            'prima3': np.nan,
            'comision': match.get('suma_comis', 0) if pd.notna(match.get('suma_comis', np.nan)) else 0,
            'multi_n': 1,
            'ramo_ok': True,
            'razon_no_rpe': '',
            'exc_code': '',
            'exc_detalle': '',
            'revision_manual': False,
            'revision_manual_txt': 'NO',
            'anomalia': False,
            'anomalia_detalle': '',
            'sla_dias': self._get_sla(row),
        }

# ============================================================
# G) SALIDA CON COLUMNAS ADICIONALES
# ============================================================

EXC_TO_ESTADO = {
    "EXC_03_CANCELACION_POSTERIOR_TOTAL": "INACTIVO",
    "EXC_01_AJUSTE_POSTERIOR_PARCIAL": "ACTIVO",
    "EXC_04_EMISION_ANTES_INGRESO_1D": "EMISION ANTES DE INGRESO"
}

BLOQUE_COLS = [
    "mes_emision_rpe", "status_rpe_pago", "status_sva_reservado", "comentario_sva_reservado",
    "fecha_emision_rpe", "fecha_fin_vigencia_rpe", "mes_comision_rpe", "rpe_1_principal",
    "rpe_2_asistencia", "prima_neta_rpe_1", "prima_neta_rpe_2", "prima_neta_final_match",
    "comision_total_rpe", "validacion_prima_match", "diferencia_prima_rpe_vs_venta",
    "porcentaje_comision_sobre_prima", "rpe_3_reserva", "prima_neta_rpe_3_reserva",
    "cantidad_rpe_asignados", "excepcion_codigo", "detalle_excepcion", "revision_manual",
    "dictamen_color_powerbi", "Estado_Poliza", "Fecha_Estado", "SLA_DIAS"
]

def armar_bloque(res: pd.DataFrame, com_key: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    res = res.copy()
    
    # Prima final
    def prima_final(r):
        if not (isinstance(r.get('prima3', np.nan), float) and math.isnan(r.get('prima3', np.nan))):
            return r.get('prima1', 0) + r.get('prima2', 0) + r.get('prima3', 0)
        if not (isinstance(r.get('prima2', np.nan), float) and math.isnan(r.get('prima2', np.nan))):
            return r.get('prima1', 0) + r.get('prima2', 0)
        return r.get('prima1', 0)
    
    res['PRIMA_NETA_FINAL'] = res.apply(prima_final, axis=1)
    
    # Revision manual
    res['revision_manual_txt'] = res['exc_code'].apply(
        lambda x: "SI" if x and x != '' and not (isinstance(x, float) and math.isnan(x)) else "NO"
    )
    
    # Dictamen
    def get_dictamen(r):
        exc = r.get('exc_code', '')
        if exc and exc != '' and not (isinstance(exc, float) and math.isnan(exc)):
            return "NARANJA_EXCEPCION_REVISION_MANUAL"
        if r.get('group_n', 1) >= 2:
            return "AZUL_GRUPO_DUPLICADO"
        if r.get('RPE1') is None or r.get('RPE1') == '' or (isinstance(r.get('RPE1'), float) and math.isnan(r.get('RPE1'))):
            return "ROJO_SIN_RPE"
        return "SIN_COLOR_OK"
    
    res['dictamen_color_powerbi'] = res.apply(get_dictamen, axis=1)
    
    # Meses
    MESES_ES = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
    
    def mes_es(ts):
        if ts is None or pd.isna(ts):
            return None
        try:
            d = pd.Timestamp(ts)
            return f"{MESES_ES[d.month - 1]}-{d.strftime('%y')}"
        except:
            return None
    
    res['MES_DE_EMISION'] = res['emision'].apply(mes_es)
    res['STATUS'] = res['RPE1'].apply(lambda v: "" if (v is None or v == '' or (isinstance(v, float) and math.isnan(v))) else "PAGADA")
    res['MES_COMISION'] = res['MES_DE_EMISION']
    
    # Validaciones
    def validacion(r):
        pf = r.get('PRIMA_NETA_FINAL', np.nan)
        pv = r.get('prima_ventas', np.nan)
        if (isinstance(pf, float) and math.isnan(pf)) or (isinstance(pv, float) and math.isnan(pv)) or pv == 0:
            return None
        return abs((pf - pv) / pv) <= Config.TOL_PRIMA
    
    res['VALIDACION'] = res.apply(validacion, axis=1)
    res['DIF'] = res.apply(lambda r: (r.get('PRIMA_NETA_FINAL', np.nan) - r.get('prima_ventas', np.nan)) 
                           if not (math.isnan(r.get('PRIMA_NETA_FINAL', np.nan)) or math.isnan(r.get('prima_ventas', np.nan))) 
                           else None, axis=1)
    
    def porc(r):
        com = r.get('comision', np.nan)
        pf = r.get('PRIMA_NETA_FINAL', np.nan)
        if (isinstance(com, float) and math.isnan(com)) or (isinstance(pf, float) and math.isnan(pf)) or pf == 0:
            return None
        return com / pf
    
    res['PORC_COMISION'] = res.apply(porc, axis=1)
    
    # Estado_Poliza
    res['Estado_Poliza'] = res['exc_code'].apply(
        lambda x: EXC_TO_ESTADO.get(x, "ACTIVO") if x and x != '' and not (isinstance(x, float) and math.isnan(x)) else "ACTIVO"
    )
    
    # Fecha_Estado - usando rpe_2_asistencia
    fecha_map = {}
    for _, row in com_key.iterrows():
        rpe = row.get('rpe', '')
        emision = row.get('emision', pd.NaT)
        if rpe and pd.notna(emision):
            if rpe not in fecha_map or fecha_map[rpe] < emision:
                fecha_map[rpe] = emision
    
    res['Fecha_Estado'] = res['RPE2'].apply(lambda x: fecha_map.get(x, pd.NaT) if x and x != '' and not (isinstance(x, float) and math.isnan(x)) else pd.NaT)
    
    # Bloque final
    bloque = pd.DataFrame({
        "mes_emision_rpe": res["MES_DE_EMISION"],
        "status_rpe_pago": res["STATUS"],
        "status_sva_reservado": "",
        "comentario_sva_reservado": "",
        "fecha_emision_rpe": res["emision"],
        "fecha_fin_vigencia_rpe": res["fin_vig"],
        "mes_comision_rpe": res["MES_COMISION"],
        "rpe_1_principal": res["RPE1"],
        "rpe_2_asistencia": res["RPE2"],
        "prima_neta_rpe_1": res["prima1"],
        "prima_neta_rpe_2": res["prima2"],
        "prima_neta_final_match": res["PRIMA_NETA_FINAL"],
        "comision_total_rpe": res["comision"],
        "validacion_prima_match": res["VALIDACION"],
        "diferencia_prima_rpe_vs_venta": res["DIF"],
        "porcentaje_comision_sobre_prima": res["PORC_COMISION"],
        "rpe_3_reserva": res.get("RPE3", pd.Series([np.nan] * len(res))),
        "prima_neta_rpe_3_reserva": res.get("prima3", pd.Series([np.nan] * len(res))),
        "cantidad_rpe_asignados": res["multi_n"].astype("Int64"),
        "excepcion_codigo": res.get("exc_code", ""),
        "detalle_excepcion": res.get("exc_detalle", ""),
        "revision_manual": res["revision_manual_txt"],
        "dictamen_color_powerbi": res["dictamen_color_powerbi"],
        "Estado_Poliza": res["Estado_Poliza"],
        "Fecha_Estado": res["Fecha_Estado"],
        "SLA_DIAS": res.get("sla_dias", np.nan),
    })[BLOQUE_COLS]
    
    # Log
    mask_log = (
        res["RPE1"].isna() | (res["RPE1"] == "") |
        ((res.get("razon_no_rpe", "") != "") & res.get("razon_no_rpe", "").notna()) |
        (res.get("anomalia", False)) | (res.get("group_n", 1) >= 2) |
        ((res.get("exc_code", "") != "") & res.get("exc_code", "").notna())
    )
    
    log_cols = ["row_excel", "group_n", "cedula10", "cedula_raw", "ramo_raw", "ramo_code",
                "prima_ventas", "prima_ventas_raw", "ingreso", "ingreso_raw",
                "RPE1", "RPE2", "RPE3", "emision", "fin_vig", "prima1", "prima2", "prima3",
                "PRIMA_NETA_FINAL", "comision", "VALIDACION", "DIF", "PORC_COMISION",
                "multi_n", "ramo_ok", "razon_no_rpe",
                "anomalia", "anomalia_detalle", "exc_code", "exc_detalle", "revision_manual",
                "excepcion_codigo", "detalle_excepcion", "revision_manual_txt", "dictamen_color_powerbi",
                "Estado_Poliza", "sla_dias", "producto_norm", "producto_raw"]
    
    for col in log_cols:
        if col not in res.columns:
            if col in ['exc_code', 'exc_detalle', 'revision_manual', 'excepcion_codigo', 
                       'detalle_excepcion', 'revision_manual_txt', 'dictamen_color_powerbi', 'Estado_Poliza']:
                res[col] = ""
            elif col in ['anomalia', 'ramo_ok']:
                res[col] = False
            elif col in ['row_excel', 'group_n', 'cedula10', 'cedula_raw', 'ramo_raw', 'ramo_code',
                         'prima_ventas', 'prima_ventas_raw', 'ingreso', 'ingreso_raw']:
                res[col] = np.nan
            else:
                res[col] = np.nan
    
    log_df = res[mask_log][log_cols].copy()
    
    return bloque, log_df

def escribir_salida(res: pd.DataFrame, bloque: pd.DataFrame, log_df: pd.DataFrame) -> Tuple[str, str]:
    hrow = detectar_header_ventas()
    path_ven = os.path.join(Config.CARPETA, Config.NOMBRE_VENTAS)
    top = read_excel_fast(path_ven, Config.SHEET_VENTAS, n_rows=hrow + 5)
    
    if len(top) < hrow:
        raise RuntimeError(f"VENTAS no tiene header en fila {hrow}")
    
    orig_headers = ["" if v is None else str(v) for v in top[hrow - 1]]
    ncol = len(orig_headers)
    
    # Leer datos
    data = []
    if _HAS_CALAMINE:
        try:
            all_data = read_excel_fast(path_ven, Config.SHEET_VENTAS)
            for i in range(hrow, len(all_data)):
                row = list(all_data[i][:ncol])
                if len(row) < ncol:
                    row.extend([None] * (ncol - len(row)))
                data.append(row)
        except:
            data = None
    
    if data is None:
        wb = load_workbook(path_ven, read_only=True, data_only=True)
        ws = wb[Config.SHEET_VENTAS]
        for row in ws.iter_rows(min_row=hrow + 1, min_col=1, max_col=ncol, values_only=True):
            rr = list(row[:ncol])
            if len(rr) < ncol:
                rr.extend([None] * (ncol - len(rr)))
            data.append(rr)
        wb.close()
    
    headers = list(orig_headers)
    output_start = column_index_from_string(Config.COL_START_BN)
    output_end = output_start + len(bloque.columns) - 1
    
    while len(headers) < output_end:
        headers.append(f"__BLANK_{len(headers)+1}")
        for row in data:
            row.append("")
    
    # Mapear columnas
    for j in range(len(headers)):
        nh = normalize_name(headers[j])
        if nh == normalize_name("AÑO"):
            headers[j] = "excepcion_codigo"
        elif nh == normalize_name("MES AÑO"):
            headers[j] = "detalle_excepcion"
    
    # Agregar nombres de bloque
    for k, name in enumerate(bloque.columns):
        headers[output_start - 1 + k] = name
    
    # Llenar datos
    bloque_idx = bloque.reset_index(drop=True)
    res_idx = res.reset_index(drop=True)
    
    for i in range(len(res_idx)):
        try:
            row_excel = int(res_idx.iloc[i]["row_excel"])
        except:
            continue
        data_i = row_excel - (hrow + 1)
        if data_i < 0 or data_i >= len(data):
            continue
        for k in range(len(bloque.columns)):
            val = bloque_idx.iat[i, k]
            data[data_i][output_start - 1 + k] = val
    
    # Guardar
    wb = Workbook()
    ws = wb.active
    ws.title = Config.SHEET_VENTAS
    ws.append(headers)
    
    for row in data:
        ws.append([None if (v is None or v is pd.NaT or (isinstance(v, float) and math.isnan(v))) else v for v in row])
    
    # Estilos
    bold = Font(bold=True)
    alc = Alignment(horizontal="center", vertical="center")
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=j)
        c.font = bold
        c.alignment = alc
    
    # Formato de fechas
    def set_fmt(col, fmt, start, end):
        for r in range(start, end + 1):
            ws.cell(row=r, column=col).number_format = fmt
    
    hn = [normalize_name(h) for h in headers]
    for j, h in enumerate(hn):
        if "fecha" in h or "inicio" in h or "fin" in h:
            set_fmt(j + 1, "DD/MM/YYYY", 2, len(data) + 1)
    
    # Colores
    bcol = {name: output_start + k for k, name in enumerate(bloque.columns)}
    fill_bad = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_group = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_exc = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    
    rpe1_col = bcol.get("rpe_1_principal")
    dict_col = bcol.get("dictamen_color_powerbi")
    color_start = rpe1_col if rpe1_col else output_start
    color_end = output_end
    
    res_idx = res.reset_index(drop=True)
    
    def out_row(r):
        try:
            row_excel = int(r["row_excel"])
            data_i = row_excel - (hrow + 1)
            return data_i + 2 if 0 <= data_i < len(data) else None
        except:
            return None
    
    for _, r in res_idx.iterrows():
        out_r = out_row(r)
        if out_r is None:
            continue
        rpe1_blank = (r.get("RPE1") is None or r.get("RPE1") == "" or 
                     (isinstance(r.get("RPE1"), float) and math.isnan(r.get("RPE1"))))
        if rpe1_blank and rpe1_col:
            ws.cell(row=out_r, column=rpe1_col).fill = fill_bad
        if dict_col and rpe1_blank:
            ws.cell(row=out_r, column=dict_col).fill = fill_bad
        if r.get("group_n", 1) >= 2:
            for cc in range(color_start, color_end + 1):
                ws.cell(row=out_r, column=cc).fill = fill_group
        exc_code = r.get("exc_code", "")
        if exc_code not in (None, "", np.nan):
            for cc in range(color_start, color_end + 1):
                ws.cell(row=out_r, column=cc).fill = fill_exc
    
    # Ajustar columnas
    ws.freeze_panes = "A2"
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
    for j in range(output_start, output_end + 1):
        ws.column_dimensions[get_column_letter(j)].width = 22
    
    os.makedirs(Config.CARPETA_RESULTADOS, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = os.path.join(Config.CARPETA_RESULTADOS, f"VENTAS_con_RPE_{timestamp}.xlsx")
    
    max_attempts = 5
    attempt = 0
    saved = False
    while attempt < max_attempts and not saved:
        try:
            if os.path.exists(out_file):
                try:
                    os.remove(out_file)
                except PermissionError:
                    base, ext = os.path.splitext(out_file)
                    out_file = f"{base}_{attempt+1}{ext}"
            wb.save(out_file)
            saved = True
            msg(f"Archivo guardado: {out_file}")
        except PermissionError:
            attempt += 1
            if attempt < max_attempts:
                msg(f"Bloqueado, intento {attempt}/{max_attempts}...")
                time.sleep(2)
                base, ext = os.path.splitext(out_file)
                out_file = f"{base}_{attempt}{ext}"
            else:
                raise RuntimeError(f"No se pudo guardar después de {max_attempts} intentos")
    
    # Guardar log
    out_log = os.path.join(Config.CARPETA_RESULTADOS, f"LOG_RPE_{timestamp}.csv")
    df_log = log_df.copy()
    for c in df_log.columns:
        df_log[c] = df_log[c].map(lambda v: "" if (v is None or v is pd.NaT or (isinstance(v, float) and math.isnan(v))) else
                                  (pd.Timestamp(v).strftime("%Y-%m-%d") if isinstance(v, (pd.Timestamp, datetime, date)) else v))
    df_log.to_csv(out_log, index=False, encoding="utf-8-sig")
    
    return out_file, out_log

# ============================================================
# H) MAIN
# ============================================================

def main():
    start = time.time()
    msg("=" * 70)
    msg("RPE MATCH - VERSIÓN OPTIMIZADA CON SLA (Hoja2)")
    msg("=" * 70)
    msg()
    
    if not os.path.isdir(Config.CARPETA):
        raise RuntimeError(f"No existe: {Config.CARPETA}")
    
    os.makedirs(Config.CARPETA_RESULTADOS, exist_ok=True)
    
    # Cargar SLA maestro (usando Hoja2)
    msg("0. Cargando tabla maestra de SLA (Hoja2)...")
    sla_df = cargar_sla_maestro()
    if not sla_df.empty:
        msg(f"   ✅ {len(sla_df)} registros de SLA cargados")
        # Mostrar algunos ejemplos
        msg(f"   Ejemplos:")
        for _, row in sla_df.head(3).iterrows():
            msg(f"     - {row['producto_original']} | {row['ramo_original']} → {row['sla_dias']} días")
    else:
        msg("   ⚠️ No se cargaron registros de SLA")
    
    msg("1. Cargando COMISIONES...")
    com = cargar_comisiones()
    msg(f"   {len(com)} filas")
    
    msg("2. Construyendo COM_KEY...")
    com_key = build_com_key(com)
    msg(f"   {len(com_key)} registros únicos")
    del com
    
    msg("3. Cargando VENTAS...")
    ven = leer_ventas()
    msg(f"   {len(ven)} filas")
    
    msg("4. Ejecutando match...")
    engine = MatchEngine(com_key, sla_df)
    res = engine.ejecutar_match(ven)
    msg(f"   {len(res)} procesados")
    
    msg("5. Armando salida...")
    bloque, log_df = armar_bloque(res, com_key)
    out_file, out_log = escribir_salida(res, bloque, log_df)
    
    elapsed = time.time() - start
    msg()
    msg("=" * 70)
    msg("PROCESO COMPLETADO")
    msg("=" * 70)
    msg(f"Archivo: {out_file}")
    msg(f"Log: {out_log}")
    msg(f"Tiempo: {elapsed:.2f}s")
    
    total = len(res)
    con_rpe = res[res['RPE1'].notna() & (res['RPE1'] != '')].shape[0]
    msg(f"Total: {total} | Con RPE: {con_rpe} ({con_rpe/total*100:.1f}%) | Sin RPE: {total-con_rpe} ({(total-con_rpe)/total*100:.1f}%)")
    
    # Mostrar estadísticas de SLA
    if 'sla_dias' in res.columns:
        sla_count = res['sla_dias'].notna().sum()
        msg(f"SLA asignados: {sla_count} ({sla_count/total*100:.1f}%)")
        if sla_count > 0:
            sla_mean = res['sla_dias'].mean()
            sla_min = res['sla_dias'].min()
            sla_max = res['sla_dias'].max()
            msg(f"SLA - Promedio: {sla_mean:.1f} días | Mín: {sla_min} | Máx: {sla_max}")
    
    for code in [Config.EXC_01_CODE, Config.EXC_02_CODE, Config.EXC_03_CODE, Config.EXC_04_CODE]:
        count = (res['exc_code'] == code).sum() if 'exc_code' in res.columns else 0
        msg(f"{code}: {count}")
    
    msg("=" * 70)

if __name__ == "__main__":
    main()