# -*- coding: utf-8 -*-
# ============================================================
# BG - Automatizacion RPE + Actualizacion Maestros (PYTHON / Pandas)
# COMISIONES + VENTAS + CRUCE RPE
#
# Port fiel del proceso original en R. Misma logica de negocio,
# mismas tolerancias, mismas excepciones (EXC_01, EXC_03, EXC_04),
# mismas salidas: VENTAS_con_RPE_YYYYMMDD.xlsx y LOG_RPE_YYYYMMDD.csv.
#
# Robusto a la estructura real de los archivos:
#  - Detecta la hoja correcta por CONTENIDO, no por nombre
#    (PRELIMINAR VENTAS viene con nombre tipo GUID).
#  - Detecta la fila real de encabezados (no asume fila 1).
#  - Lee solo las columnas que existen (no pide columnas fuera de rango).
#  - Tolera el crecimiento semanal de los archivos.
#
# IMPORTANTE (validacion antes de produccion):
#  Esta version reproduce la logica de R. Antes de calendarizar en el
#  servidor, conviene correr una vez el proceso en R y una vez en Python
#  con los mismos archivos y comparar VENTAS_con_RPE y LOG_RPE. R y Python
#  pueden diferir en empates del algoritmo de asignacion; el costo esta
#  disenado para que los empates sean raros, pero la comparacion es la
#  garantia.
#
# Dependencias:
#   pip install pandas numpy openpyxl scipy python-dateutil
#
# Ejecucion:
#   python bg_rpe_automatizacion.py
# ============================================================

import os
import re
import math
import time
import shutil
import tempfile
import warnings
import unicodedata
from datetime import date, datetime, timedelta

import numpy as np
import pandas as pd
from dateutil.relativedelta import relativedelta

import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment, PatternFill

from scipy.optimize import linear_sum_assignment

# Lector rapido opcional (Rust). Si esta instalado, acelera la lectura de los
# Excel grandes ~7x. Si no, el script usa openpyxl automaticamente.
# Para instalarlo:  pip install python-calamine
try:
    from python_calamine import CalamineWorkbook
    _HAS_CALAMINE = True
except Exception:
    _HAS_CALAMINE = False

warnings.filterwarnings("ignore")  # silencia avisos de openpyxl sobre celdas con formato de fecha invalido


# ============================================================
# A) CONFIGURACION
# ============================================================

# En tu PC usa estos valores por defecto. En el servidor, define las
# variables de entorno RPE_LOOK y RPE_RESULTADOS apuntando a la carpeta
# compartida y NO toques el codigo.
carpeta = os.environ.get(
    "RPE_LOOK",
    r"C:/Users/fplaza/Desktop/F/Rodri/r_comisiones_venta/LOOK",
)
carpeta_resultados = os.environ.get(
    "RPE_RESULTADOS",
    r"C:/Users/fplaza/Desktop/F/Rodri/r_comisiones_venta/RESULTADOS",
)

# Nombres de archivo (acepta "PRELIMINAR COMISIONES.xlsx" o "PRELIMINAR_COMISIONES.xlsx")
NOMBRE_VENTAS = "VENTAS.xlsx"
NOMBRE_COMISIONES = "COMISIONES.xlsx"
NOMBRES_PRE_COMISIONES = ["PRELIMINAR COMISIONES.xlsx", "PRELIMINAR_COMISIONES.xlsx"]
NOMBRES_PRE_VENTAS = ["PRELIMINAR VENTAS.xlsx", "PRELIMINAR_VENTAS.xlsx"]

sheet_comisiones = "Base"
sheet_ventas = "BASE"

ACTUALIZAR_COMISIONES_CON_PRELIMINAR = True
ACTUALIZAR_VENTAS_CON_PRELIMINAR = True

# Muestra avance del cruce cada 5000 cedulas (util para procesos largos).
MOSTRAR_PROGRESO = True

HACER_BACKUP_COMISIONES = False
HACER_BACKUP_VENTAS = False
BORRAR_BACKUPS_VENTAS_EXISTENTES = True

# Fila preferida de encabezados en los preliminares (si la deteccion por
# contenido no encuentra algo mejor, usa estas).
HEADER_ROW_PRE_COMISIONES = 4
HEADER_ROW_PRE_VENTAS = 9

# La hoja BASE de VENTAS tiene los encabezados en la fila 2 (la fila 1
# son titulos de grupo). Los datos del cruce se leen desde la fila 2,
# igual que en R (la fila de encabezado se descarta sola por traer
# cedula/prima/fecha invalidas).
ventas_row_ini = 2
ventas_row_fin = None  # se detecta solo

# Formato de fechas en TEXTO. En Ecuador es DIA primero (DD/MM/AAAA), que es
# el valor por defecto. Las fechas reales de Excel (no texto) se leen sin
# ambiguedad. Si tu CRM exportara fechas en texto formato MES primero
# (MM/DD/AAAA, estilo EE.UU.) y vieras dias/meses invertidos, cambia a False.
FECHA_DIA_PRIMERO = True

# Columnas del cruce en VENTAS (letras de Excel), igual que en R.
col_ventas_cedula = "B"    # 10 DIGITOS (cedula ya a 10 digitos)
col_ventas_prima = "AC"    # Prima Neta
col_ventas_ingreso = "M"   # Fecha de Ingreso
col_ventas_ramo = "U"      # Ramo

col_start_BN = "BN"        # inicio del bloque RPE en la salida

tol_prima = 0.03
tol_abs = 10
max_meses = 11

tol_cancel_rel = 0.005
tol_cancel_abs = 2

HABILITAR_PAIR_ASISTENCIA = True
# Dias maximos entre la emision del RPE principal y la del RPE de asistencia.
# La poliza principal y su rider de asistencia se emiten en el mismo ciclo. Se
# observo que R acepta pares con ~8 dias de diferencia y rechaza los de ~12, por
# lo que el limite efectivo es ~10. El par SOLO se forma si las dos primas suman
# el total de la venta dentro de tolerancia (esa es la prueba real del cruce);
# el limite de dias es un filtro secundario.
MAX_DIAS_ASISTENCIA = 10
ASISTENCIA_MAX_FRAC = 0.35
MAIN_TOP = 40
OTHER_TOP = 200

USAR_LSAP_POR_CEDULA = True
LSAP_MAX_FILAS_CEDULA = 220
LSAP_MAX_CANDIDATOS_CEDULA = 1200
DUMMY_UNMATCH_COST = 1e12
IMPOSSIBLE_COST = 1e15

HABILITAR_PRIORIDAD_ANTIGUEDAD_MATCH = True
PRIORIDAD_ANTIGUEDAD_COST = 1e4

HABILITAR_AJUSTE_POSTERIOR = True
AJUSTE_POS_MAX_REL = 0.10
AJUSTE_POS_MAX_ABS = 150
AJUSTE_NEG_MAX_DIAS = 45
AJUSTE_POS_TOP = 60
EXC_01_CODE = "EXC_01_AJUSTE_POSTERIOR_PARCIAL"

# EXC_02 existe como configuracion en el R original pero NUNCA se implemento
# (no hay funcion ni se invoca). Se conserva identico: su contador siempre es 0.
EXC_02_CODE = "EXC_02_GRUPO_REPARTO_UNICO"

HABILITAR_CANCELADO_POSTERIOR = True
CANCELADO_POST_MAX_DIAS = 365
EXC_03_CODE = "EXC_03_CANCELACION_POSTERIOR_TOTAL"

HABILITAR_EMISION_ANTES_INGRESO_1D = True
EMISION_ANTES_INGRESO_MAX_DIAS = 1
EXC_04_CODE = "EXC_04_EMISION_ANTES_INGRESO_1D"

EPS_ZERO_CONJUNTO = 1e-9
EPOCH = date(1970, 1, 1)


def msg(*args):
    print("".join(str(a) for a in args), flush=True)


def first_existing(folder, names):
    for n in names:
        p = os.path.join(folder, n)
        if os.path.exists(p):
            return p
    return os.path.join(folder, names[0])


file_ven = os.path.join(carpeta, NOMBRE_VENTAS)
file_com = os.path.join(carpeta, NOMBRE_COMISIONES)
file_pre_com = first_existing(carpeta, NOMBRES_PRE_COMISIONES)
file_pre_ven = first_existing(carpeta, NOMBRES_PRE_VENTAS)


# ============================================================
# C) HELPERS GENERALES (parseo, normalizacion)
# ============================================================

def col2num(col):
    return column_index_from_string(str(col).upper())


def normalize_name(x):
    if x is None:
        s = ""
    elif isinstance(x, float) and math.isnan(x):
        s = ""
    else:
        s = str(x)
    s = s.strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def is_blank_scalar(x):
    if x is None or x is pd.NaT:
        return True
    try:
        if isinstance(x, float) and math.isnan(x):
            return True
    except Exception:
        pass
    return str(x).strip() == ""


def excel_serial_to_ts(n):
    try:
        days = int(math.floor(float(n)))
    except Exception:
        return pd.NaT
    return pd.Timestamp(date(1899, 12, 30) + timedelta(days=days))


def parse_general_date(s):
    """Parsea fechas en TEXTO. Respeta FECHA_DIA_PRIMERO (Ecuador = dia primero).
    Las fechas ISO (ANIO-MES-DIA, con o sin 'T' y microsegundos) son inequivocas
    y NUNCA se interpretan dia-primero (evita que 2026-05-08 se vuelva 2026-08-05)."""
    s2 = s.strip()
    # ISO inequivoco (empieza con anio de 4 digitos): parsear sin dayfirst.
    if re.match(r"^\d{4}-\d{1,2}-\d{1,2}", s2):
        for f in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S.%f",
                  "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S",
                  "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
            try:
                return pd.Timestamp(datetime.strptime(s2, f))
            except ValueError:
                continue
        try:
            return pd.Timestamp(pd.to_datetime(s2, errors="raise")).normalize()
        except Exception:
            return pd.NaT
    fmts = []
    if FECHA_DIA_PRIMERO:
        for sep in ["-", "/", "."]:
            fmts += [f"%d{sep}%m{sep}%Y", f"%d{sep}%m{sep}%y"]
    else:
        for sep in ["-", "/", "."]:
            fmts += [f"%m{sep}%d{sep}%Y", f"%m{sep}%d{sep}%y"]
    timed = []
    for f in fmts:
        timed += [f + " %H:%M:%S", f + " %H:%M"]
    for f in fmts + timed:
        try:
            return pd.Timestamp(datetime.strptime(s2, f))
        except ValueError:
            continue
    try:
        return pd.Timestamp(pd.to_datetime(s2, dayfirst=FECHA_DIA_PRIMERO, errors="raise")).normalize()
    except Exception:
        return pd.NaT


def parse_fecha_scalar(x):
    if x is None or x is pd.NaT:
        return pd.NaT
    if isinstance(x, datetime):
        return pd.Timestamp(x).normalize()
    if isinstance(x, date):
        return pd.Timestamp(x)
    if isinstance(x, pd.Timestamp):
        return x.normalize()
    if isinstance(x, bool):
        return pd.NaT
    if isinstance(x, (int, float, np.integer, np.floating)):
        v = float(x)
        if math.isnan(v):
            return pd.NaT
        if v >= 1e7:
            s = "%08.0f" % v
            try:
                return pd.Timestamp(datetime.strptime(s, "%Y%m%d"))
            except Exception:
                return pd.NaT
        if 20000 < v < 60000:
            return excel_serial_to_ts(v)
        return pd.NaT
    s = str(x).strip()
    if s == "":
        return pd.NaT
    try:
        xn = float(s)
    except ValueError:
        xn = None
    if xn is not None and 20000 < xn < 60000:
        return excel_serial_to_ts(xn)
    if re.fullmatch(r"\d{8}", s):
        try:
            return pd.Timestamp(datetime.strptime(s, "%Y%m%d"))
        except Exception:
            return pd.NaT
    return parse_general_date(s)


def parse_fecha(series):
    return pd.Series([parse_fecha_scalar(v) for v in series], index=getattr(series, "index", None))


def to_num_scalar(x):
    if x is None or x is pd.NaT:
        return float("nan")
    if isinstance(x, bool):
        return float("nan")
    if isinstance(x, (int, float, np.integer, np.floating)):
        v = float(x)
        return float("nan") if math.isnan(v) else v
    s = str(x).strip()
    if s == "":
        return float("nan")
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^0-9,\.\-]", "", s)
    s = re.sub(r"[.,]+$", "", s)
    if s in ("", "-"):
        return float("nan")
    has_comma = "," in s
    has_dot = "." in s
    try:
        if has_comma and has_dot:
            if s.rfind(",") > s.rfind("."):
                t = s.replace(".", "").replace(",", ".", 1)
            else:
                t = s.replace(",", "")
            return float(t)
        elif has_comma and not has_dot:
            if s.count(",") > 1:
                return float(s.replace(",", ""))
            right = s.split(",")[-1]
            return float(s.replace(",", "")) if len(right) == 3 else float(s.replace(",", ".", 1))
        elif has_dot and not has_comma:
            if s.count(".") > 1:
                return float(s.replace(".", ""))
            right = s.split(".")[-1]
            return float(s.replace(".", "")) if len(right) == 3 else float(s)
        return float(s)
    except ValueError:
        return float("nan")


def to_num_commission_scalar(x):
    if x is None or x is pd.NaT:
        return float("nan")
    if isinstance(x, bool):
        return float("nan")
    if isinstance(x, (int, float, np.integer, np.floating)):
        v = float(x)
        return float("nan") if math.isnan(v) else v
    s = str(x).strip()
    if s == "":
        return float("nan")
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^0-9,\.\-]", "", s)
    s = re.sub(r"[.,]+$", "", s)
    if s in ("", "-"):
        return float("nan")
    has_comma = "," in s
    has_dot = "." in s
    try:
        if has_comma and has_dot:
            if s.rfind(",") > s.rfind("."):
                t = s.replace(".", "").replace(",", ".", 1)
            else:
                t = s.replace(",", "")
            return float(t)
        elif has_comma and not has_dot:
            if s.count(",") > 1:
                idx = s.rfind(",")
                return float(s[:idx].replace(",", "") + "." + s[idx + 1:])
            return float(s.replace(",", ".", 1))
        elif has_dot and not has_comma:
            if s.count(".") > 1:
                parts = s.split(".")
                if len(parts) > 2 and all(len(p) == 3 for p in parts[1:]):
                    return float("".join(parts))
                return float("".join(parts[:-1]) + "." + parts[-1])
            return float(s)
        return float(s)
    except ValueError:
        return float("nan")


def to_num(series):
    return pd.Series([to_num_scalar(v) for v in series], index=getattr(series, "index", None), dtype="float64")


def to_num_commission(series):
    return pd.Series([to_num_commission_scalar(v) for v in series], index=getattr(series, "index", None), dtype="float64")


def clean_cedula10_scalar(x):
    if x is None or x is pd.NaT:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float, np.integer, np.floating)):
        v = float(x)
        if math.isnan(v):
            return None
        x = "%.0f" % v
    s = str(x).strip()
    if s == "":
        return None
    if re.search(r"[eE]", s):
        try:
            s = "%.0f" % float(s)
        except ValueError:
            pass
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^0-9]", "", s)
    if s == "":
        return None
    out = s
    if len(out) == 13 and out.endswith("001"):
        out = out[:10]
    if len(out) > 10 and out.endswith("001"):
        out = out[: len(out) - 3]
    if len(out) > 10:
        out = out[:10]
    if len(out) < 10:
        out = out.rjust(10, "0")
    if len(out) != 10:
        return None
    return out


def clean_cedula10(series):
    return pd.Series([clean_cedula10_scalar(v) for v in series], index=getattr(series, "index", None), dtype="object")


MESES_ES = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]


def mes_es_scalar(ts):
    if ts is None or ts is pd.NaT:
        return None
    try:
        d = pd.Timestamp(ts)
    except Exception:
        return None
    if pd.isna(d):
        return None
    return f"{MESES_ES[d.month - 1]}-{d.strftime('%y')}"


def mes_es(series):
    return pd.Series([mes_es_scalar(v) for v in series], index=getattr(series, "index", None))


def fmt_num2(x):
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return None
    return "%.2f" % round(float(x), 2)


def date_to_epoch_days(ts):
    return (pd.Timestamp(ts).date() - EPOCH).days


def add_months(ts, n):
    if ts is None or pd.isna(ts):
        return pd.NaT
    return pd.Timestamp(pd.Timestamp(ts).date() + relativedelta(months=n))


def days_between(a, b):
    return (pd.Timestamp(a).normalize() - pd.Timestamp(b).normalize()).days


def _id_part(v):
    """Formatea una parte de un identificador (Ramo/Poliza/Endoso/etc) como
    texto limpio, SIN decimales. calamine lee numeros como float (1.0), asi que
    1.0 -> '1', 437140.0 -> '437140', 0.0 -> '0'. Strings se devuelven tal cual."""
    if v is None or v is pd.NaT:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (np.integer, int)):
        return str(int(v))
    if isinstance(v, (np.floating, float)):
        fv = float(v)
        if math.isnan(fv):
            return ""
        if fv == int(fv) and abs(fv) < 1e15:
            return str(int(fv))
        return ("%f" % fv).rstrip("0").rstrip(".")
    s = str(v).strip()
    # texto numerico tipo "1.0" / "437140.0" -> sin decimales
    if re.fullmatch(r"-?\d+\.0+", s):
        return s.split(".")[0]
    return s


def to_char_cell(v):
    if v is None or v is pd.NaT:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    # numpy.datetime64 NO es subclase de datetime: convertir a Timestamp para
    # no producir cadenas tipo "2026-05-08T05:00:00.000000" que luego se
    # mal-parsean (invirtiendo dia/mes).
    if isinstance(v, np.datetime64):
        ts = pd.Timestamp(v)
        if pd.isna(ts):
            return ""
        return ts.strftime("%Y-%m-%d %H:%M:%S") if (ts.hour or ts.minute or ts.second) else ts.strftime("%Y-%m-%d")
    if isinstance(v, pd.Timestamp):
        if pd.isna(v):
            return ""
        return v.strftime("%Y-%m-%d %H:%M:%S") if (v.hour or v.minute or v.second) else v.strftime("%Y-%m-%d")
    if isinstance(v, (np.integer, int)):
        return str(int(v))
    if isinstance(v, (np.floating, float)):
        fv = float(v)
        if math.isnan(fv):
            return ""
        if fv == int(fv) and abs(fv) < 1e15:
            return str(int(fv))
        return repr(fv)
    if isinstance(v, datetime):
        if v.hour or v.minute or v.second:
            return v.strftime("%Y-%m-%d %H:%M:%S")
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    return str(v)


def coerce_cell_for_excel(v):
    if v is None or v is pd.NaT:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    if isinstance(v, np.floating):
        fv = float(v)
        return None if math.isnan(fv) else fv
    if isinstance(v, np.integer):
        return int(v)
    if isinstance(v, np.bool_):
        return bool(v)
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    return v


def resolve_conjunto_value(conj_vals, raw_vals, source_conj, source_raw, eps_zero=EPS_ZERO_CONJUNTO):
    conj = [v for v in conj_vals if v is not None and not (isinstance(v, float) and math.isnan(v))]
    raw_nonan = [v for v in raw_vals if v is not None and not (isinstance(v, float) and math.isnan(v))]
    raw_sum = sum(raw_nonan) if len(raw_nonan) > 0 else float("nan")
    if isinstance(raw_sum, float) and not math.isfinite(raw_sum):
        raw_sum = float("nan")
    if len(conj) == 0:
        return {"value": raw_sum, "source": source_raw}
    conj_pick = conj[0]
    best = abs(conj[0])
    for v in conj[1:]:
        if abs(v) > best:
            best = abs(v)
            conj_pick = v
    if conj_pick is None or (isinstance(conj_pick, float) and math.isnan(conj_pick)):
        return {"value": raw_sum, "source": source_raw}
    if abs(conj_pick) <= eps_zero and not (isinstance(raw_sum, float) and math.isnan(raw_sum)) and abs(raw_sum) > eps_zero:
        return {"value": raw_sum, "source": source_raw + "_FALLBACK_" + source_conj + "_CERO"}
    return {"value": conj_pick, "source": source_conj}


# ============================================================
# LECTURA DE EXCEL (robusta a hojas/encabezados/columnas variables)
# ============================================================

# --- Cache de lecturas calamine (evita releer el mismo archivo varias veces) ---
_CALAMINE_CACHE = {}


def _calamine_sheet_names(path):
    wb = CalamineWorkbook.from_path(path)
    return list(wb.sheet_names)


def _calamine_read_all(path, sheet):
    """Lee una hoja completa con calamine como lista de listas. Cacheado."""
    key = (path, sheet, os.path.getmtime(path))
    if key in _CALAMINE_CACHE:
        return _CALAMINE_CACHE[key]
    wb = CalamineWorkbook.from_path(path)
    ws = wb.get_sheet_by_name(sheet)
    data = ws.to_python(skip_empty_area=False)
    data = [list(r) for r in data]
    # mantener el cache chico: solo el ultimo archivo grande
    if len(_CALAMINE_CACHE) > 2:
        _CALAMINE_CACHE.clear()
    _CALAMINE_CACHE[key] = data
    return data


def excel_sheet_names(path):
    """Nombres de hoja, via calamine si esta disponible, si no openpyxl."""
    if _HAS_CALAMINE:
        try:
            return _calamine_sheet_names(path)
        except Exception:
            pass
    wb = load_workbook(path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def read_top_rows(path, sheet, n_rows, max_cols=None, fill_merged=True):
    """Lee solo las primeras n_rows de una hoja (para detectar hoja/encabezado).
    Usa calamine si esta disponible; si no, openpyxl read_only (barato en
    memoria). El merge de celdas solo se resuelve en archivos pequenos."""
    if _HAS_CALAMINE:
        try:
            data = _calamine_read_all(path, sheet)
            rows = [list(r[:max_cols]) if max_cols else list(r) for r in data[:n_rows]]
            return rows
        except Exception:
            pass
    return _read_top_rows_openpyxl(path, sheet, n_rows, max_cols=max_cols, fill_merged=fill_merged)


def _read_top_rows_openpyxl(path, sheet, n_rows, max_cols=None, fill_merged=True):
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"No existe la hoja '{sheet}' en {os.path.basename(path)}")
    ws = wb[sheet]
    max_col = ws.max_column or 0
    if max_cols is not None and max_col > 0:
        max_col = min(max_col, max_cols)
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=n_rows, min_col=1,
                            max_col=(max_col if max_col > 0 else None), values_only=True):
        rows.append(list(row))
    wb.close()

    if fill_merged:
        try:
            size_mb = os.path.getsize(path) / (1024 * 1024)
        except Exception:
            size_mb = 999
        if size_mb < 25:
            wb2 = load_workbook(path, read_only=False, data_only=True)
            ws2 = wb2[sheet]
            mc = ws2.max_column or 0
            if max_cols is not None and mc > 0:
                mc = min(mc, max_cols)
            for rng in list(ws2.merged_cells.ranges):
                if rng.min_row <= n_rows and (rng.min_col - 1) < mc:
                    if (rng.min_row - 1) < len(rows) and (rng.min_col - 1) < len(rows[rng.min_row - 1]):
                        tl = rows[rng.min_row - 1][rng.min_col - 1]
                    else:
                        tl = None
                    for rr in range(rng.min_row, min(rng.max_row, n_rows) + 1):
                        if (rr - 1) >= len(rows):
                            continue
                        for cc in range(rng.min_col, min(rng.max_col, mc) + 1):
                            while (cc - 1) >= len(rows[rr - 1]):
                                rows[rr - 1].append(None)
                            rows[rr - 1][cc - 1] = tl
            wb2.close()
    return rows


def read_sheet_to_tbl(path, sheet, header_row, max_cols=None):
    """Lee una hoja completa y devuelve un Tbl. header_row 1-indexado.
    Usa calamine si esta disponible (rapido); si no, openpyxl read_only.
    Recorta columnas finales vacias y filas totalmente en blanco."""
    if _HAS_CALAMINE:
        try:
            return _read_sheet_to_tbl_calamine(path, sheet, header_row, max_cols=max_cols)
        except Exception:
            pass

    # encabezados desde las primeras filas
    top = read_top_rows(path, sheet, header_row, max_cols=max_cols, fill_merged=True)
    if len(top) < header_row:
        return Tbl([], pd.DataFrame())
    header_vals = top[header_row - 1]
    headers = [("" if v is None else str(v)) for v in header_vals]
    ncol = len(headers)

    # datos (lectura por filas read_only)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    data = []
    for row in ws.iter_rows(min_row=header_row + 1, min_col=1, max_col=ncol, values_only=True):
        r = list(row[:ncol])
        if len(r) < ncol:
            r = r + [None] * (ncol - len(r))
        if any(not is_blank_scalar(v) for v in r):
            data.append(r)
    wb.close()

    df = pd.DataFrame(data, columns=list(range(ncol)))
    return _trim_tbl(headers, df, ncol)


def _read_sheet_to_tbl_calamine(path, sheet, header_row, max_cols=None):
    data_all = _calamine_read_all(path, sheet)
    if len(data_all) < header_row:
        return Tbl([], pd.DataFrame())
    header_vals = data_all[header_row - 1]
    if max_cols is not None:
        header_vals = header_vals[:max_cols]
    headers = [("" if v is None else str(v)) for v in header_vals]
    ncol = len(headers)
    data = []
    for r in data_all[header_row:]:
        rr = list(r[:ncol])
        if len(rr) < ncol:
            rr = rr + [None] * (ncol - len(rr))
        if any(not is_blank_scalar(v) for v in rr):
            data.append(rr)
    df = pd.DataFrame(data, columns=list(range(ncol)))
    return _trim_tbl(headers, df, ncol)


def _trim_tbl(headers, df, ncol):
    # recortar columnas finales sin encabezado y sin datos
    keep_until = -1
    for j in range(ncol):
        header_nonblank = not is_blank_scalar(headers[j])
        col_has_data = df[j].map(lambda v: not is_blank_scalar(v)).any() if len(df) > 0 else False
        if header_nonblank or col_has_data:
            keep_until = j
    if keep_until < ncol - 1 and keep_until >= 0:
        headers = headers[:keep_until + 1]
        df = df[list(range(keep_until + 1))]
    elif keep_until < 0:
        headers = []
        df = pd.DataFrame()
    return Tbl(headers, df)


def sheet_names(path):
    return excel_sheet_names(path)


class Tbl:
    """Tabla con encabezados crudos (headers) y un DataFrame con columnas
    enteras 0..n-1 alineadas a headers."""
    def __init__(self, headers, df):
        self.headers = list(headers)
        self.df = df


def matrix_to_tbl(matrix, header_row_idx, max_cols=None):
    """Construye un Tbl a partir de una lista de filas ya leidas (para
    archivos pequenos: preliminares). matrix es lista de listas."""
    if len(matrix) <= header_row_idx:
        return Tbl([], pd.DataFrame())
    header_vals = matrix[header_row_idx]
    if max_cols is not None:
        header_vals = header_vals[:max_cols]
    headers = [("" if v is None else str(v)) for v in header_vals]
    ncol = len(headers)
    norm_data = []
    for r in matrix[header_row_idx + 1:]:
        rr = list(r[:ncol])
        if len(rr) < ncol:
            rr = rr + [None] * (ncol - len(rr))
        if any(not is_blank_scalar(v) for v in rr):
            norm_data.append(rr)
    df = pd.DataFrame(norm_data, columns=list(range(ncol)))
    keep_until = -1
    for j in range(ncol):
        header_nonblank = not is_blank_scalar(headers[j])
        col_has_data = df[j].map(lambda v: not is_blank_scalar(v)).any() if len(df) > 0 else False
        if header_nonblank or col_has_data:
            keep_until = j
    if keep_until < ncol - 1 and keep_until >= 0:
        headers = headers[:keep_until + 1]
        df = df[list(range(keep_until + 1))]
    elif keep_until < 0:
        headers = []
        df = pd.DataFrame()
    return Tbl(headers, df)


def read_excel_table_raw(path, sheet, header_row=1):
    """Equivalente a read_excel_table_raw de R. header_row 1-indexado.
    Usa el lector eficiente en memoria (read_only)."""
    return read_sheet_to_tbl(path, sheet, header_row)


def get_col_idx(tbl, label, occurrence=1, contains=False):
    norm_headers = [normalize_name(h) for h in tbl.headers]
    nl = normalize_name(label)
    if contains:
        idx = [i for i, h in enumerate(norm_headers) if nl in h]
    else:
        idx = [i for i, h in enumerate(norm_headers) if h == nl]
    if len(idx) < occurrence:
        return None
    return idx[occurrence - 1]


def get_col_any(tbl, labels, occurrence=1, contains=False):
    for lab in labels:
        i = get_col_idx(tbl, lab, occurrence=occurrence, contains=contains)
        if i is not None:
            return i
    return None


def find_source_col_alias(tbl, aliases):
    i = get_col_any(tbl, aliases, contains=False)
    if i is not None:
        return i
    return get_col_any(tbl, aliases, contains=True)


# ----- deteccion de hoja y fila de encabezado por contenido -----

def detect_header_row(path, sheet, preferred_row=1, keywords=(), scan_to=40):
    matrix = read_top_rows(path, sheet, scan_to, fill_merged=True)
    keys = [normalize_name(k) for k in keywords]
    keys = [k for k in keys if k]
    if not keys:
        return preferred_row
    scan = min(scan_to, len(matrix))

    def score_row(vals):
        vn = [normalize_name(v) for v in vals]
        vn = [v for v in vn if v]
        if not vn:
            return 0
        s = 0
        for k in keys:
            if any(v == k for v in vn) or any(k in v for v in vn):
                s += 1
        return s

    scores = [score_row(matrix[i]) for i in range(scan)]
    if not scores:
        return preferred_row
    pref_score = scores[preferred_row - 1] if preferred_row - 1 < len(scores) else 0
    best_row = int(np.argmax(scores)) + 1
    best_score = max(scores)
    if pref_score >= 2:
        return preferred_row
    if best_score >= 2:
        return best_row
    return preferred_row


def select_sheet(path, preferred_sheet=None, keywords=()):
    shs = sheet_names(path)
    if preferred_sheet is not None and preferred_sheet in shs:
        return preferred_sheet
    if len(shs) == 1:
        return shs[0]
    keys = [normalize_name(k) for k in keywords]
    best_sheet, best_score = shs[0], -1
    for sh in shs:
        try:
            m = read_top_rows(path, sh, 30, fill_merged=True)
        except Exception:
            continue
        top = m[:30]
        vals = set()
        for row in top:
            for v in row:
                vals.add(normalize_name(v))
        score = sum(1 for k in keys if any(k in v for v in vals if v))
        if score > best_score:
            best_score, best_sheet = score, sh
    return best_sheet


# ============================================================
# D) MAPEO RAMO
# ============================================================

RAMO_MAP = {
    normalize_name("ACCIDENTES PERSONALES"): 8,
    normalize_name("CALDERAS Y MAQUINARIAS"): 11,
    normalize_name("EQUIPO DE CONTRATISTAS"): 5,
    normalize_name("INCENDIO"): 1,
    normalize_name("RESPONSABILIDAD CIVIL"): 12,
    normalize_name("TRANSPORTE"): 16,
    normalize_name("VEHICULOS"): 4,
    normalize_name("VIDA TERMINO"): 30,
}


def ramo_code_from_text_scalar(x):
    xn = normalize_name(x)
    if xn == "":
        return None
    if xn in RAMO_MAP:
        return RAMO_MAP[xn]
    for k, code in RAMO_MAP.items():
        if k in xn:
            return code
    return None


def _rpe_vacio(v):
    """True si un RPE esta vacio: None, NaN, NaT o cadena vacia. Robusto para
    no confundir un NaN con un RPE valido (causaba que filas sin RPE quedaran
    como SIN_COLOR_OK / PAGADA en vez de ROJO_SIN_RPE)."""
    if v is None or v is pd.NaT:
        return True
    try:
        if isinstance(v, float) and math.isnan(v):
            return True
    except Exception:
        pass
    return str(v).strip() == "" or str(v).strip().lower() in ("nan", "none", "nat")


def _clean_rpe(v):
    """Normaliza un RPE leido a 'ramo-poliza-endoso' sin decimales.
    Cubre el caso en que las partes vienen como float (1.0-437140.0-0.0)
    o el RPE viene como un solo numero. Si no tiene el patron, lo deja igual."""
    if v is None or v is pd.NaT:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    s = str(v).strip()
    if s == "":
        return ""
    if "-" in s:
        partes = s.split("-")
        return "-".join(_id_part(p) for p in partes)
    return _id_part(v)


def rpe_prefix_code(rpe):
    m = re.match(r"^\d+", str(rpe))
    return int(m.group(0)) if m else None


def rpe_base(rpe):
    m = re.match(r"^\d+-\d+", str(rpe))
    return m.group(0) if m else None


# ============================================================
# E) ACTUALIZACION MAESTRO COMISIONES
# ============================================================

ALIAS_COMISIONES = {
    "Cédula": ["Cédula", "Cedula", "Identificación", "Identificacion", "RUC", "Documento"],
    "Fecha Emision": ["Fecha Emision", "Fecha Emisión", "F Emision", "F EMISION"],
    "F EMISION": ["F EMISION", "Fecha Emision", "Fecha Emisión"],
    "Fin Vig": ["Fin Vig", "Fecha Fin Vigencia", "Fin Vigencia", "FECHA FIN VIG"],
    "FECHA FIN VIG": ["FECHA FIN VIG", "Fin Vig", "Fecha Fin Vigencia", "Fin Vigencia"],
    "Prima": ["Prima", "Prima Neta", "Prima Total"],
    "Sub Total Comisión": ["Sub Total Comisión", "Sub Total Comision", "Subtotal Comisión", "Subtotal Comision"],
    "Ramo": ["Ramo"],
    "Póliza": ["Póliza", "Poliza"],
    "Endoso": ["Endoso"],
    "Clave R-P-E": ["Clave R-P-E", "Clave RPE", "Clave 3", "RPE"],
}


def copy_common_columns_with_aliases(src_tbl, target_headers, alias_map):
    """Copia por nombre normalizado (con ocurrencia) y rellena con alias."""
    src_norm = [normalize_name(h) for h in src_tbl.headers]
    target_norm = [normalize_name(h) for h in target_headers]
    src_map = {}
    for i, n in enumerate(src_norm):
        src_map.setdefault(n, []).append(i)

    n_rows = len(src_tbl.df)
    out = pd.DataFrame({j: [None] * n_rows for j in range(len(target_headers))})

    for j in range(len(target_headers)):
        tn = target_norm[j]
        if tn == "":
            continue
        occ = sum(1 for k in range(j + 1) if target_norm[k] == tn)
        if tn in src_map and len(src_map[tn]) >= occ:
            src_pos = src_map[tn][occ - 1]
            out[j] = src_tbl.df[src_pos].values

    for target_label, aliases in alias_map.items():
        tnl = normalize_name(target_label)
        positions = [j for j in range(len(target_headers)) if target_norm[j] == tnl]
        if not positions:
            continue
        src_pos = find_source_col_alias(src_tbl, aliases)
        if src_pos is None:
            continue
        for p in positions:
            if out[p].map(lambda v: is_blank_scalar(v)).all():
                out[p] = src_tbl.df[src_pos].values

    return Tbl(list(target_headers), out)


def recalcular_columnas_comisiones(tbl):
    h = tbl.headers
    df = tbl.df
    def gi(label):
        return get_col_idx(tbl, label)
    col_cedula = gi("Cédula"); col_ced_real = gi("Cédula Real"); col_cc = gi("cc")
    col_clave = gi("CLAVE"); col_rpe = gi("Clave R-P-E"); col_mes_com = gi("Mes comisión")
    col_f_emision = gi("F EMISION"); col_ramo = gi("Ramo"); col_poliza = gi("Póliza")
    col_endoso = gi("Endoso"); col_fecha_emis = gi("Fecha Emision"); col_prima = gi("Prima")
    col_subtotal = gi("Sub Total Comisión"); col_prima_conj = gi("PRIMA CONJUNTO")
    col_sub_conj = gi("SUBTOTAL CONJUNTO"); col_fin_vig = gi("Fin Vig"); col_fecha_fin = gi("FECHA FIN VIG")

    if col_cedula is not None and col_ced_real is not None:
        df[col_ced_real] = df[col_cedula].map(lambda v: str(v).strip() if not is_blank_scalar(v) else "")
    if col_cc is not None and col_ced_real is not None:
        def cc_calc(v):
            cr = re.sub(r"[^0-9]", "", str(v).strip()) if not is_blank_scalar(v) else ""
            return cr[:10] if len(cr) == 13 else cr
        df[col_cc] = df[col_ced_real].map(cc_calc)
    if col_rpe is not None and col_ramo is not None and col_poliza is not None and col_endoso is not None:
        df[col_rpe] = [
            f"{_id_part(a)}-{_id_part(b)}-{_id_part(c)}"
            for a, b, c in zip(df[col_ramo], df[col_poliza], df[col_endoso])
        ]
    if col_f_emision is not None and col_fecha_emis is not None:
        df[col_f_emision] = list(parse_fecha(df[col_fecha_emis]))
    if col_mes_com is not None and col_f_emision is not None:
        df[col_mes_com] = list(parse_fecha(df[col_f_emision]))
    if col_fecha_fin is not None and col_fin_vig is not None:
        df[col_fecha_fin] = list(parse_fecha(df[col_fin_vig]))
    if col_prima is not None:
        df[col_prima] = list(to_num(df[col_prima]))
    if col_subtotal is not None:
        df[col_subtotal] = list(to_num_commission(df[col_subtotal]))
    if col_clave is not None and col_ced_real is not None and col_prima is not None:
        df[col_clave] = [
            (str(cr).strip() + ("" if is_blank_scalar(p) else (str(int(p)) if (isinstance(p, float) and p == int(p)) else str(p))))
            for cr, p in zip(df[col_ced_real], df[col_prima])
        ]
    if col_prima_conj is not None and col_rpe is not None and col_prima is not None:
        tmp = pd.DataFrame({"rpe": df[col_rpe].values, "prima": to_num(df[col_prima]).values})
        df[col_prima_conj] = tmp.groupby("rpe")["prima"].transform(lambda s: s.sum(skipna=True)).values
    if col_sub_conj is not None and col_rpe is not None and col_subtotal is not None:
        tmp = pd.DataFrame({"rpe": df[col_rpe].values, "sub": to_num_commission(df[col_subtotal]).values})
        df[col_sub_conj] = tmp.groupby("rpe")["sub"].transform(lambda s: s.sum(skipna=True)).values
    tbl.df = df
    return tbl


# ----- escritura de masters con formatos -----

def _norm_in(label_norm, label_list):
    return label_norm in [normalize_name(x) for x in label_list]


def escribir_master_xlsx(tbl, path, sheet, kind):
    headers = tbl.headers
    df = tbl.df.copy()
    n = len(df)
    hn = [normalize_name(h) for h in headers]

    # coercion de tipos (equivale a preparar_tipos_*_para_excel)
    if kind == "comisiones":
        date_cols, money_cols, pct_cols, text_cols = [], [], [], []
        for j, h in enumerate(hn):
            if ("fecha" in h) or ("inicio vig" in h) or ("fin vig" in h) or _norm_in(h, ["F EMISION", "Mes comisión", "FECHA FIN VIG"]):
                date_cols.append(j)
            if _norm_in(h, ["Prima", "Sub Total Comisión", "Total Adicional", "Total", "PRIMA CONJUNTO", "SUBTOTAL CONJUNTO"]):
                money_cols.append(j)
            if _norm_in(h, ["% Comisión", "% Adicional", "Tasas", "% Comision"]):
                pct_cols.append(j)
        for j in date_cols:
            df[j] = list(parse_fecha(df[j]))
        for j in money_cols:
            df[j] = list(to_num_commission(df[j]))
        for j in pct_cols:
            df[j] = list(to_num(df[j]))
    else:  # ventas
        date_cols, money_cols, pct_cols = [], [], []
        text_cols = []
        for j, h in enumerate(hn):
            if ("fecha" in h) or ("inicio" in h) or ("fin" in h) or ("vigencia" in h):
                date_cols.append(j)
            if ("prima" in h) or ("valor asegurado" in h) or ("monto" in h) or ("total" in h):
                money_cols.append(j)
            if _norm_in(h, ["CLAVE", "10 DIGITOS", "Código Venta", "Código Cliente", "Cédula", "Teléfono",
                            "Comprobante", "Autorización", "ID Facturación", "Identificación", "RUC", "Número Crédito"]):
                text_cols.append(j)
        # limpiar NA->"" en columnas tipo texto
        for j in range(len(headers)):
            df[j] = df[j].map(lambda v: "" if (is_blank_scalar(v) or str(v) in ("NA", "NaN", "<NA>")) else v)
        for j in date_cols:
            df[j] = list(parse_fecha(df[j]))
        for j in money_cols:
            df[j] = list(to_num(df[j]))

    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    for i in range(n):
        ws.append([coerce_cell_for_excel(df.iat[i, j]) for j in range(len(headers))])

    bold_center = Font(bold=True)
    align_c = Alignment(horizontal="center", vertical="center")
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=j)
        c.font = bold_center
        c.alignment = align_c

    def apply_fmt(cols, fmt):
        for j in cols:
            for r in range(2, n + 2):
                ws.cell(row=r, column=j + 1).number_format = fmt

    if kind == "comisiones":
        apply_fmt(date_cols, "DD/MM/YYYY")
        apply_fmt(money_cols, '"$"#,##0.00')
        apply_fmt(pct_cols, "0.00%")
    else:
        apply_fmt(date_cols, "DD/MM/YYYY")
        apply_fmt(money_cols, '"$"#,##0.00')
        apply_fmt(text_cols, "@")

    ws.freeze_panes = "A2"
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 15

    tmp = tempfile.NamedTemporaryFile(prefix="TMP_MASTER_", suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    try:
        load_workbook(tmp.name).close()
    except Exception:
        os.unlink(tmp.name)
        raise RuntimeError(f"El archivo temporal de {kind} no se pudo validar. No se reemplazo el maestro.")
    shutil.copy(tmp.name, path)
    os.unlink(tmp.name)
    msg(f"{os.path.basename(path)} actualizado y reemplazado correctamente.")


def row_key_from_columns(tbl, col_idxs):
    if not col_idxs:
        return pd.Series([""] * len(tbl.df))
    partes = []
    for cc in col_idxs:
        header_label = tbl.headers[cc] if cc < len(tbl.headers) else ""
        h = normalize_name(header_label)
        vals = tbl.df[cc]
        if re.search(r"fecha|inicio|fin", h):
            d = parse_fecha(vals)
            col = [(dd.strftime("%Y-%m-%d") if not pd.isna(dd) else ("" if is_blank_scalar(v) else str(v).strip()))
                   for dd, v in zip(d, vals)]
        elif re.search(r"prima|valor|total|comision|monto", h):
            nums = to_num(vals)
            col = [("%.2f" % round(nn, 2) if not (isinstance(nn, float) and math.isnan(nn)) else ("" if is_blank_scalar(v) else str(v).strip()))
                   for nn, v in zip(nums, vals)]
        else:
            # texto: normalizar numeros con decimal cero ("369749.0" -> "369749")
            # para que la llave de deduplicacion coincida sin importar si el
            # numero quedo guardado como float (calamine) o como texto.
            col = []
            for v in vals:
                if is_blank_scalar(v):
                    col.append("")
                else:
                    s = str(v).strip()
                    if re.fullmatch(r"-?\d+\.0+", s):
                        s = s.split(".")[0]
                    col.append(s)
        col = [unicodedata.normalize("NFKD", str(c)).encode("ascii", "ignore").decode("ascii").lower() for c in col]
        col = [re.sub(r"\s+", " ", c).strip() for c in col]
        partes.append(col)
    return pd.Series(["||".join(t) for t in zip(*partes)])


def actualizar_archivo_comisiones():
    if not ACTUALIZAR_COMISIONES_CON_PRELIMINAR:
        msg("INFO: Actualizacion de COMISIONES desactivada por configuracion.")
        return None
    if not os.path.exists(file_pre_com):
        msg("INFO: No existe preliminar de comisiones. No se actualizo COMISIONES: ", file_pre_com)
        return None
    if not os.path.exists(file_com):
        raise RuntimeError("No existe archivo maestro de comisiones: " + file_com)

    kw = ["Línea", "Ramo", "Póliza", "Clave 3", "Sub Total Comisión", "Fecha Emision"]
    sheet_pre = select_sheet(file_pre_com, preferred_sheet="Preliminar Comisiones", keywords=kw)
    header_pre = detect_header_row(file_pre_com, sheet_pre, preferred_row=HEADER_ROW_PRE_COMISIONES, keywords=kw)
    msg("Hoja preliminar comisiones usada: ", sheet_pre)
    msg("Fila encabezado preliminar comisiones usada: ", header_pre)

    master = read_excel_table_raw(file_com, sheet_comisiones, header_row=1)
    pre = read_excel_table_raw(file_pre_com, sheet_pre, header_row=header_pre)

    pre_as_master = copy_common_columns_with_aliases(pre, master.headers, ALIAS_COMISIONES)
    pre_as_master = recalcular_columnas_comisiones(pre_as_master)
    master = recalcular_columnas_comisiones(master)

    cfp = get_col_idx(pre_as_master, "F EMISION")
    if cfp is None:
        raise RuntimeError("No encontre F EMISION calculada en preliminar de comisiones.")
    pre_dates = parse_fecha(pre_as_master.df[cfp])
    if pre_dates.isna().all():
        raise RuntimeError("No pude interpretar F EMISION del preliminar de comisiones.")
    fecha_min = pre_dates.min()
    fecha_max = pre_dates.max()

    cfm = get_col_idx(master, "F EMISION")
    if cfm is None:
        raise RuntimeError("No encontre F EMISION en COMISIONES.xlsx.")
    master_dates = parse_fecha(master.df[cfm]).reset_index(drop=True)

    keep_mask = master_dates.isna() | (master_dates < fecha_min) | (master_dates > fecha_max)
    master.df = master.df.reset_index(drop=True)
    master_keep = Tbl(master.headers, master.df[keep_mask.values].reset_index(drop=True))
    master_rango = Tbl(master.headers, master.df[(~keep_mask).values].reset_index(drop=True))

    cols_cmp = [j for j in range(len(pre_as_master.headers))
                if pre_as_master.df[j].map(lambda v: not is_blank_scalar(v)).any()]
    old_keys = sorted(row_key_from_columns(master_rango, cols_cmp).tolist())
    new_keys = sorted(row_key_from_columns(pre_as_master, cols_cmp).tolist())
    if old_keys == new_keys:
        msg(f"COMISIONES no se actualizo: el rango {fecha_min} a {fecha_max} ya esta igual al preliminar.")
        return None

    actualizado_df = pd.concat([master_keep.df, pre_as_master.df], ignore_index=True)
    actualizado = Tbl(master.headers, actualizado_df)
    actualizado = recalcular_columnas_comisiones(actualizado)

    cf = get_col_idx(actualizado, "F EMISION")
    if cf is not None:
        ord_d = parse_fecha(actualizado.df[cf])
        order = ord_d.sort_values(kind="stable", na_position="first").index
        actualizado.df = actualizado.df.loc[order].reset_index(drop=True)

    escribir_master_xlsx(actualizado, file_com, sheet_comisiones, kind="comisiones")
    msg("COMISIONES actualizado. Rango: ", fecha_min, " a ", fecha_max,
        " | nuevas: ", len(pre_as_master.df), " | finales: ", len(actualizado.df))
    return actualizado


# ============================================================
# F) ACTUALIZACION MAESTRO VENTAS
# ============================================================

PRE_VENTAS_HEADERS_OFICIAL = [
    "Código Venta", "Tipo ID", "Código Cliente", "Cédula", "Asegurado", "Segmento  Estratégico",
    "Oficial Venta", "Agencia", "Opid", "Ejecutiva Seguros", "Fecha de Ingreso", "Región",
    "Observaciones", "Email", "Teléfono", "Producto", "Tipo Producto", "Plan", "Ramo",
    "Valor Asegurado", "Fecha Pago", "Forma Pago", "Tipo Pago", "Tipo Comprobante", "Comprobante",
    "Autorización", "Prima Neta", "Prima Total", "ID Facturación", "Identificación", "Nombre",
    "Identificación", "Nombre", "Inicio Póliza", "Fin Póliza", "Envio Documentos", "Cotización",
    "Ficha Cliente", "Carta Débito", "Facturación Electrónica", "Votación", "Planilla",
    "Declaración         Salud", "Declaración Imp. Renta", "RUC", "Representante           Legal",
    "Matrícula", "Inspeción", "Exámenes Médicos", "Otros Documentos", "Observaciones",
    "Motivo Devolución", "Usuario de Ingreso", "Fecha de Modificación", "Usuario  Modificó",
    "Tipo Periodo Tabla", "Modelo Seguro", "Fecha Dividendo", "Número Dividendo", "Tipo Crédito",
    "Número Crédito", "Inicio Crédito", "Fin Crédito",
]

ALIAS_VENTAS = {
    "Código Venta": ["Código Venta", "Codigo Venta", "Cod Venta", "Cod. Venta", "Código", "Codigo"],
    "Tipo ID": ["Tipo ID", "Tipo Identificación", "Tipo Identificacion"],
    "Código Cliente": ["Código Cliente", "Codigo Cliente", "Cod Cliente"],
    "Cédula": ["Cédula", "Cedula"],
    "Asegurado": ["Asegurado", "Cliente", "Nombre Asegurado"],
    "Segmento  Estratégico": ["Segmento  Estratégico", "Segmento Estratégico", "Segmento Estrategico"],
    "Oficial Venta": ["Oficial Venta"], "Agencia": ["Agencia"], "Opid": ["Opid", "OPID"],
    "Ejecutiva Seguros": ["Ejecutiva Seguros", "Ejecutivo Seguros"],
    "Fecha de Ingreso": ["Fecha de Ingreso", "Fecha Ingreso", "F Ingreso", "Fecha Venta"],
    "Región": ["Región", "Region"], "Observaciones": ["Observaciones", "Observación", "Observacion"],
    "Email": ["Email", "Correo", "Correo Electrónico", "Correo Electronico"],
    "Teléfono": ["Teléfono", "Telefono", "Celular"], "Producto": ["Producto"],
    "Tipo Producto": ["Tipo Producto"], "Plan": ["Plan"], "Ramo": ["Ramo", "Tipo Ramo"],
    "Valor Asegurado": ["Valor Asegurado", "Suma Asegurada"], "Fecha Pago": ["Fecha Pago", "Fecha de Pago"],
    "Forma Pago": ["Forma Pago", "Forma de Pago"], "Tipo Pago": ["Tipo Pago", "Tipo de Pago"],
    "Tipo Comprobante": ["Tipo Comprobante"], "Comprobante": ["Comprobante"],
    "Autorización": ["Autorización", "Autorizacion"], "Prima Neta": ["Prima Neta"], "Prima Total": ["Prima Total"],
    "ID Facturación": ["ID Facturación", "ID Facturacion"], "Identificación": ["Identificación", "Identificacion"],
    "Nombre": ["Nombre"], "Inicio Póliza": ["Inicio Póliza", "Inicio Poliza"],
    "Fin Póliza": ["Fin Póliza", "Fin Poliza"], "Envio Documentos": ["Envio Documentos", "Envío Documentos"],
    "Cotización": ["Cotización", "Cotizacion"], "Ficha Cliente": ["Ficha Cliente"],
    "Carta Débito": ["Carta Débito", "Carta Debito"],
    "Facturación Electrónica": ["Facturación Electrónica", "Facturacion Electronica"],
    "Votación": ["Votación", "Votacion"], "Planilla": ["Planilla"],
    "Declaración         Salud": ["Declaración         Salud", "Declaración Salud", "Declaracion Salud"],
    "Declaración Imp. Renta": ["Declaración Imp. Renta", "Declaracion Imp. Renta"], "RUC": ["RUC"],
    "Representante           Legal": ["Representante           Legal", "Representante Legal"],
    "Matrícula": ["Matrícula", "Matricula"], "Inspeción": ["Inspeción", "Inspección", "Inspeccion"],
    "Exámenes Médicos": ["Exámenes Médicos", "Examenes Medicos"], "Otros Documentos": ["Otros Documentos"],
    "Motivo Devolución": ["Motivo Devolución", "Motivo Devolucion"],
    "Usuario de Ingreso": ["Usuario de Ingreso", "Usuario Ingreso"],
    "Fecha de Modificación": ["Fecha de Modificación", "Fecha de Modificacion"],
    "Usuario  Modificó": ["Usuario  Modificó", "Usuario Modificó", "Usuario Modifico"],
    "Tipo Periodo Tabla": ["Tipo Periodo Tabla"], "Modelo Seguro": ["Modelo Seguro"],
    "Fecha Dividendo": ["Fecha Dividendo"], "Número Dividendo": ["Número Dividendo", "Numero Dividendo"],
    "Tipo Crédito": ["Tipo Crédito", "Tipo Credito"],
    "Número Crédito": ["Número Crédito", "Numero Credito", "No Crédito", "No Credito", "Credito", "Crédito"],
    "Inicio Crédito": ["Inicio Crédito", "Inicio Credito"], "Fin Crédito": ["Fin Crédito", "Fin Credito"],
}

VENTAS_ESENCIALES = [
    "Código Venta", "Tipo ID", "Código Cliente", "Cédula", "Asegurado", "Fecha de Ingreso",
    "Producto", "Tipo Producto", "Plan", "Ramo", "Valor Asegurado", "Fecha Pago",
    "Prima Neta", "Prima Total", "Número Crédito", "Inicio Crédito", "Fin Crédito",
]


def score_header_ventas(vals):
    vn = set(normalize_name(v) for v in vals if not is_blank_scalar(v))
    return sum(1 for k in VENTAS_ESENCIALES if normalize_name(k) in vn)


def detectar_fila_encabezado_ventas_master(path, sheet, preferida=2, scan_to=15):
    """Detecta la fila de encabezados del maestro VENTAS por contenido.
    Busca la fila con mas columnas esenciales (Cédula, Fecha de Ingreso,
    Producto, Prima Neta, etc.). Usa 'preferida' si empata o si nada supera
    el umbral. Robusto si el archivo cambia de fila de encabezados."""
    top = read_top_rows(path, sheet, scan_to, fill_merged=True)
    if not top:
        return preferida
    scores = [score_header_ventas(r) for r in top]
    best_row = int(np.argmax(scores)) + 1
    best_score = max(scores) if scores else 0
    pref_score = scores[preferida - 1] if (preferida - 1) < len(scores) else 0
    # si la fila preferida ya es buena (>=8 esenciales), usarla
    if pref_score >= 8:
        return preferida
    # si hay una fila claramente mejor, usarla
    if best_score >= 8:
        return best_row
    # ultimo recurso: la preferida
    return preferida


def select_sheet_header_ventas(path, preferred_sheet="Preliminar Ventas", preferred_row=9, scan_to=40):
    """Detecta hoja + fila de encabezados de PRELIMINAR VENTAS por contenido.
    No depende del nombre de la hoja (puede venir como GUID)."""
    shs = sheet_names(path)
    candidatos = []  # (score, sheet, row_excel)
    for sh in shs:
        try:
            m = read_top_rows(path, sh, scan_to, fill_merged=True)
        except Exception:
            continue
        top = min(scan_to, len(m))
        for i in range(top):
            sc = score_header_ventas(m[i])
            candidatos.append((sc, sh, i + 1))
    if not candidatos:
        raise RuntimeError("No pude escanear ninguna hoja valida en PRELIMINAR VENTAS.xlsx.")
    # preferida si cumple umbral
    for sc, sh, rr in candidatos:
        if sh == preferred_sheet and rr == preferred_row and sc >= 12:
            return sh, rr, sc
    candidatos.sort(key=lambda t: (-t[0], t[1], t[2]))
    best = candidatos[0]
    if best[0] < 12:
        raise RuntimeError(
            "No encontre una fila de encabezado confiable en PRELIMINAR VENTAS.xlsx. "
            f"Mejor score={best[0]} | Hoja={best[1]} | Fila={best[2]}. No actualizo VENTAS."
        )
    return best[1], best[2], best[0]


def read_preliminar_ventas_limpio(path, sheet, header_row, max_col=120):
    return read_sheet_to_tbl(path, sheet, header_row, max_cols=max_col)


def source_col_ventas(tbl, target_label, occurrence=1):
    src_norm = [normalize_name(h) for h in tbl.headers]
    aliases = ALIAS_VENTAS.get(target_label, [target_label])
    idx = []
    for a in aliases:
        na = normalize_name(a)
        idx += [i for i, h in enumerate(src_norm) if h == na]
    idx = sorted(set(idx))
    return idx[occurrence - 1] if len(idx) >= occurrence else None


def copy_ventas_preliminar_a_master(pre, master_headers):
    n = len(pre.df)
    out = pd.DataFrame({j: [""] * n for j in range(len(master_headers))})
    target_norm = [normalize_name(h) for h in master_headers]
    common_norm = set(normalize_name(h) for h in PRE_VENTAS_HEADERS_OFICIAL)
    # mapa de nombre-oficial-normalizado -> etiqueta oficial (para usar alias)
    oficial_por_norm = {normalize_name(h): h for h in PRE_VENTAS_HEADERS_OFICIAL}
    for j in range(len(master_headers)):
        tln = target_norm[j]
        if tln in (normalize_name("CLAVE"), normalize_name("10 DIGITOS")):
            continue
        if tln == "":
            continue
        occ = sum(1 for k in range(j + 1) if target_norm[k] == tln)
        # 1) si el nombre del maestro coincide con un encabezado oficial, usar su alias
        target_label = oficial_por_norm.get(tln, master_headers[j])
        src_pos = source_col_ventas(pre, target_label, occurrence=occ)
        # 2) intento por la etiqueta literal del maestro
        if src_pos is None:
            src_pos = source_col_ventas(pre, master_headers[j], occurrence=occ)
        # 3) ultimo intento: match directo por nombre normalizado en el preliminar
        if src_pos is None:
            pre_norm = [normalize_name(h) for h in pre.headers]
            cand = [i for i, h in enumerate(pre_norm) if h == tln]
            if len(cand) >= occ:
                src_pos = cand[occ - 1]
        if src_pos is not None:
            out[j] = [to_char_cell(v) for v in pre.df[src_pos].values]
    return Tbl(list(master_headers), out)


def recalcular_columnas_ventas(tbl):
    col_clave = get_col_idx(tbl, "CLAVE")
    col_10 = get_col_idx(tbl, "10 DIGITOS")
    col_ced = get_col_idx(tbl, "Cédula")
    col_prima = get_col_idx(tbl, "Prima Neta")
    df = tbl.df
    if col_10 is not None and col_ced is not None:
        def left10(v):
            c = re.sub(r"[^0-9]", "", re.sub(r"\s+", "", str(v).strip())) if not is_blank_scalar(v) else ""
            return c[:10] if c != "" else ""
        df[col_10] = df[col_ced].map(left10)
    if col_clave is not None and col_10 is not None and col_prima is not None:
        primas = to_num(df[col_prima])
        ced = df[col_10].map(lambda v: str(v).strip())
        df[col_clave] = [
            (cc + str(int(math.trunc(p)))) if (cc != "" and not (isinstance(p, float) and math.isnan(p))) else ""
            for cc, p in zip(ced, primas)
        ]
    # limpiar NA->""
    for j in range(len(tbl.headers)):
        df[j] = df[j].map(lambda v: "" if (is_blank_scalar(v) or str(v) in ("NA", "NaN", "<NA>")) else v)
    tbl.df = df
    return tbl


def detectar_filas_mal_cargadas_ventas(tbl):
    cp = get_col_idx(tbl, "Producto"); cc = get_col_idx(tbl, "Cédula")
    cf = get_col_idx(tbl, "Fecha de Ingreso"); cpr = get_col_idx(tbl, "Prima Neta")
    cl = get_col_idx(tbl, "CLAVE")
    if None in (cp, cc, cf, cpr):
        return pd.Series([False] * len(tbl.df))
    prod = tbl.df[cp].map(lambda v: not is_blank_scalar(v))
    ced = tbl.df[cc].map(is_blank_scalar)
    fec = tbl.df[cf].map(is_blank_scalar)
    pri = tbl.df[cpr].map(is_blank_scalar)
    clv = tbl.df[cl].map(is_blank_scalar) if cl is not None else pd.Series([True] * len(tbl.df))
    return prod & ced & fec & pri & clv


def validar_mapeo_ventas(tbl):
    cc = get_col_idx(tbl, "Cédula"); cf = get_col_idx(tbl, "Fecha de Ingreso")
    cpr = get_col_idx(tbl, "Prima Neta"); cprod = get_col_idx(tbl, "Producto")
    ccod = get_col_idx(tbl, "Código Venta")
    if None in (cc, cf, cpr, cprod):
        faltan = []
        if cc is None: faltan.append("Cédula")
        if cf is None: faltan.append("Fecha de Ingreso")
        if cpr is None: faltan.append("Prima Neta")
        if cprod is None: faltan.append("Producto")
        muestra = ", ".join(repr(h) for h in tbl.headers[:40])
        raise RuntimeError(
            "El mapeo de VENTAS no encontro columnas esenciales. Faltan: "
            + ", ".join(faltan)
            + ". Esto suele significar que el maestro VENTAS.xlsx tiene los encabezados "
            "en una fila distinta o con nombres diferentes. Encabezados detectados: "
            + muestra
        )
    n = len(tbl.df)
    if n == 0:
        raise RuntimeError("PRELIMINAR VENTAS no tiene filas de datos.")
    pct_cod = tbl.df[ccod].map(lambda v: not is_blank_scalar(v)).mean() if ccod is not None else 0
    pct_ced = tbl.df[cc].map(lambda v: not is_blank_scalar(v)).mean()
    pct_fec = parse_fecha(tbl.df[cf]).notna().mean()
    pct_pri = to_num(tbl.df[cpr]).notna().mean()
    pct_pro = tbl.df[cprod].map(lambda v: not is_blank_scalar(v)).mean()
    msg(f"Validacion mapeo VENTAS | Código Venta: {round(100*pct_cod,1)}% | Cédula: {round(100*pct_ced,1)}% | "
        f"Fecha: {round(100*pct_fec,1)}% | Prima: {round(100*pct_pri,1)}% | Producto: {round(100*pct_pro,1)}%")
    if pct_pro >= 0.5 and (pct_ced < 0.5 or pct_fec < 0.5 or pct_pri < 0.5):
        raise RuntimeError("Mapeo incorrecto: Producto con datos pero Cédula/Fecha/Prima vacias. No se reemplazo VENTAS.")
    return True


def build_ventas_key(tbl):
    labels = ["Cédula", "Fecha de Ingreso", "Prima Neta", "Prima Total", "Producto", "Tipo Producto",
              "Plan", "Ramo", "Número Crédito", "Asegurado", "Código Cliente"]
    idxs = [get_col_idx(tbl, l) for l in labels]
    idxs = [i for i in idxs if i is not None]
    key = row_key_from_columns(tbl, idxs)
    out = []
    for k in key:
        if k is None or k.strip() == "" or re.fullmatch(r"\|*", k or ""):
            out.append(None)
        else:
            out.append("VENTAS_COMMON|" + k)
    return pd.Series(out)


def actualizar_archivo_ventas():
    if not ACTUALIZAR_VENTAS_CON_PRELIMINAR:
        msg("INFO: Actualizacion de VENTAS desactivada por configuracion.")
        return None
    if not os.path.exists(file_pre_ven):
        msg("INFO: No existe PRELIMINAR VENTAS.xlsx. No se actualizo VENTAS: ", file_pre_ven)
        return None
    if not os.path.exists(file_ven):
        raise RuntimeError("No existe VENTAS.xlsx: " + file_ven)

    sheet_pre, header_pre, score = select_sheet_header_ventas(
        file_pre_ven, preferred_sheet="Preliminar Ventas", preferred_row=HEADER_ROW_PRE_VENTAS)
    msg("Hoja preliminar ventas usada: ", sheet_pre)
    msg("Fila encabezado preliminar ventas usada: ", header_pre, " (score ", score, ")")

    # El maestro VENTAS suele tener los encabezados en la fila 2 (fila 1 =
    # titulos de grupo), pero detectamos la fila real por contenido para ser
    # robustos si el archivo cambia de estructura.
    header_master = detectar_fila_encabezado_ventas_master(file_ven, sheet_ventas, preferida=2)
    msg("Fila encabezado maestro VENTAS usada: ", header_master)
    master = read_excel_table_raw(file_ven, sheet_ventas, header_row=header_master)

    # diagnostico: confirmar que el maestro tiene las columnas clave
    _diag = {lab: get_col_idx(master, lab) for lab in ["Cédula", "Fecha de Ingreso", "Prima Neta", "Producto", "Ramo", "Código Venta"]}
    if any(v is None for v in _diag.values()):
        msg("AVISO: el maestro VENTAS no tiene todas las columnas clave en la fila ", header_master,
            ". Mapeo detectado: ", _diag)

    filas_malas = detectar_filas_mal_cargadas_ventas(master)
    if filas_malas.any():
        msg("Filas mal cargadas eliminadas de VENTAS: ", int(filas_malas.sum()))
        master.df = master.df[~filas_malas.values].reset_index(drop=True)

    pre = read_preliminar_ventas_limpio(file_pre_ven, sheet_pre, header_pre, max_col=120)
    master_headers = master.headers

    pre_as_master = copy_ventas_preliminar_a_master(pre, master_headers)
    pre_as_master = recalcular_columnas_ventas(pre_as_master)
    master = recalcular_columnas_ventas(master)

    validar_mapeo_ventas(pre_as_master)

    cfp = get_col_idx(pre_as_master, "Fecha de Ingreso")
    pre_dates = parse_fecha(pre_as_master.df[cfp])
    fecha_min = pre_dates.min() if pre_dates.notna().any() else None
    fecha_max = pre_dates.max() if pre_dates.notna().any() else None

    key_master = build_ventas_key(master).reset_index(drop=True)
    key_pre = build_ventas_key(pre_as_master).reset_index(drop=True)
    master_keys_set = set(k for k in key_master.tolist() if k is not None)
    nuevos_idx = [i for i, k in enumerate(key_pre.tolist()) if (k is not None and k not in master_keys_set)]

    if len(nuevos_idx) == 0:
        msg("VENTAS no se actualizo: el preliminar ya esta cargado o no tiene registros nuevos.")
        msg("Rango preliminar ventas: ", fecha_min, " a ", fecha_max, " | filas revisadas: ", len(pre_as_master.df))
        if filas_malas.any():
            escribir_master_xlsx(master, file_ven, sheet_ventas, kind="ventas")
            msg("VENTAS reconstruido limpio tras eliminar filas mal cargadas.")
        return None

    nuevos = Tbl(master_headers, pre_as_master.df.iloc[nuevos_idx].reset_index(drop=True))
    actualizado_df = pd.concat([master.df, nuevos.df], ignore_index=True)
    actualizado = Tbl(master_headers, actualizado_df)
    actualizado = recalcular_columnas_ventas(actualizado)

    cf = get_col_idx(actualizado, "Fecha de Ingreso")
    if cf is not None:
        ord_d = parse_fecha(actualizado.df[cf])
        order = ord_d.sort_values(kind="stable", na_position="first").index
        actualizado.df = actualizado.df.loc[order].reset_index(drop=True)

    escribir_master_xlsx(actualizado, file_ven, sheet_ventas, kind="ventas")
    msg("VENTAS actualizado. Rango: ", fecha_min, " a ", fecha_max,
        " | nuevas: ", len(nuevos.df), " | finales: ", len(actualizado.df))
    return actualizado


# ============================================================
# H) LEER COMISIONES PARA MATCH
# ============================================================

def build_com_for_match_light(path, sheet, header_row=1):
    """Lee de COMISIONES SOLO las columnas necesarias para el cruce, en modo
    read_only fila por fila. Evita cargar las 61 columnas como objetos
    (clave para no agotar la RAM con archivos grandes). Devuelve el mismo
    DataFrame 'com' que build_com_from_master."""
    top = read_top_rows(path, sheet, header_row, fill_merged=True)
    if len(top) < header_row:
        raise RuntimeError("COMISIONES no tiene la fila de encabezados esperada.")
    headers = [("" if v is None else str(v)) for v in top[header_row - 1]]
    norm = [normalize_name(h) for h in headers]

    def find(labels):
        for lab in labels:
            nl = normalize_name(lab)
            for i, h in enumerate(norm):
                if h == nl:
                    return i
        for lab in labels:  # fallback: contiene
            nl = normalize_name(lab)
            for i, h in enumerate(norm):
                if nl in h and h != "":
                    return i
        return None

    i_cc = find(["cc", "Cédula Real", "Cédula"])
    i_rpe = find(["Clave R-P-E", "Clave RPE", "Clave 3"])
    i_femis = find(["F EMISION", "Fecha Emision"])
    i_finv = find(["FECHA FIN VIG", "Fin Vig", "Fin Vigencia"])
    i_prima = find(["Prima"])
    i_subcom = find(["Sub Total Comisión", "Sub Total Comisi"])
    i_pconj = find(["PRIMA CONJUNTO"])
    i_sconj = find(["SUBTOTAL CONJUNTO", "SUB TOTAL CONJUNTO"])
    if None in (i_cc, i_rpe, i_finv, i_prima):
        raise RuntimeError("No pude mapear columnas esenciales en COMISIONES para match.")

    wanted = {k: v for k, v in {
        "cc": i_cc, "rpe": i_rpe, "femis": i_femis, "finv": i_finv,
        "prima": i_prima, "subcom": i_subcom, "pconj": i_pconj, "sconj": i_sconj
    }.items() if v is not None}
    maxc = max(wanted.values()) + 1

    cols = {k: [] for k in wanted}
    if _HAS_CALAMINE:
        try:
            data_all = _calamine_read_all(path, sheet)
            for r in data_all[header_row:]:
                if all(is_blank_scalar(r[ci]) if ci < len(r) else True for ci in wanted.values()):
                    continue
                for k, ci in wanted.items():
                    cols[k].append(r[ci] if ci < len(r) else None)
        except Exception:
            cols = None
    else:
        cols = None

    if cols is None:
        cols = {k: [] for k in wanted}
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        r_idx = 0
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=maxc, values_only=True):
            r_idx += 1
            if r_idx <= header_row:
                continue
            if all(is_blank_scalar(row[wanted[k]]) for k in wanted):
                continue
            for k, ci in wanted.items():
                cols[k].append(row[ci] if ci < len(row) else None)
        wb.close()

    def col(k):
        return pd.Series(cols[k]) if k in cols else None

    n = len(cols["cc"])
    com = pd.DataFrame({
        "cedula10": clean_cedula10(col("cc")).values,
        "rpe": [_clean_rpe(v) for v in cols["rpe"]],
        "fecha_emis": parse_fecha(col("femis")).values if "femis" in cols else [pd.NaT] * n,
        "fin_vig": parse_fecha(col("finv")).values,
        "prima": to_num(col("prima")).values,
        "subcom": to_num_commission(col("subcom")).values if "subcom" in cols else [np.nan] * n,
        "prima_conj": to_num(col("pconj")).values if "pconj" in cols else [np.nan] * n,
        "subcom_conj": to_num_commission(col("sconj")).values if "sconj" in cols else [np.nan] * n,
    })
    return com


def build_com_from_master(tbl):
    col_cc = get_col_any(tbl, ["cc", "Cédula Real", "Cédula"])
    col_rpe = get_col_any(tbl, ["Clave R-P-E", "Clave RPE", "Clave 3"])
    col_femis = get_col_any(tbl, ["F EMISION", "Fecha Emision"])
    col_finv = get_col_any(tbl, ["FECHA FIN VIG", "Fin Vig", "Fin Vigencia"])
    col_prima = get_col_any(tbl, ["Prima"])
    col_subcom = get_col_any(tbl, ["Sub Total Comisión", "Sub Total Comisi"])
    col_prima_conj = get_col_any(tbl, ["PRIMA CONJUNTO"])
    col_sub_conj = get_col_any(tbl, ["SUBTOTAL CONJUNTO", "SUB TOTAL CONJUNTO"])
    if None in (col_cc, col_rpe, col_finv, col_prima):
        raise RuntimeError("No pude mapear columnas esenciales en COMISIONES para match.")
    df = tbl.df
    com = pd.DataFrame({
        "cedula10": clean_cedula10(df[col_cc]).values,
        "rpe": [_clean_rpe(v) for v in df[col_rpe].values],
        "fecha_emis": parse_fecha(df[col_femis]).values if col_femis is not None else pd.NaT,
        "fin_vig": parse_fecha(df[col_finv]).values,
        "prima": to_num(df[col_prima]).values,
        "subcom": to_num_commission(df[col_subcom]).values if col_subcom is not None else np.nan,
        "prima_conj": to_num(df[col_prima_conj]).values if col_prima_conj is not None else np.nan,
        "subcom_conj": to_num_commission(df[col_sub_conj]).values if col_sub_conj is not None else np.nan,
    })
    return com


# ============================================================
# I) COM_KEY + FLAGS
# ============================================================

def build_com_key(com):
    com = com.copy()
    com["fecha_emis"] = pd.to_datetime(com["fecha_emis"], errors="coerce")
    com["fin_vig"] = pd.to_datetime(com["fin_vig"], errors="coerce")
    mask = (com["cedula10"].notna() & (com["cedula10"] != "") &
            com["rpe"].notna() & (com["rpe"] != "") &
            com["fecha_emis"].notna() &
            (com["prima"].notna() | com["prima_conj"].notna()))
    base = com[mask].copy()

    rows = []
    for (ced, rpe), g in base.groupby(["cedula10", "rpe"], sort=False):
        fe = g["fecha_emis"].dropna()
        fv = g["fin_vig"].dropna()
        pp = resolve_conjunto_value(g["prima_conj"].tolist(), g["prima"].tolist(), "PRIMA_CONJUNTO", "SUM_PRIMA")
        pc = resolve_conjunto_value(g["subcom_conj"].tolist(), g["subcom"].tolist(), "SUBTOTAL_CONJUNTO", "SUM_SUBTOTAL")
        rows.append({
            "cedula10": ced, "rpe": rpe,
            "emision": fe.min() if len(fe) > 0 else pd.NaT,
            "fin_vig": fv.max() if len(fv) > 0 else pd.NaT,
            "suma_prima": pp["value"], "suma_comis": pc["value"],
        })
    ck = pd.DataFrame(rows)
    if ck.empty:
        ck = pd.DataFrame(columns=["cedula10", "rpe", "emision", "fin_vig", "suma_prima", "suma_comis"])
    ck["emision"] = pd.to_datetime(ck["emision"], errors="coerce")
    ck["fin_vig"] = pd.to_datetime(ck["fin_vig"], errors="coerce")
    ck["rpe_code"] = ck["rpe"].map(rpe_prefix_code)
    ck["base"] = ck["rpe"].map(rpe_base)
    ck["is_negative"] = ck["suma_prima"].notna() & (ck["suma_prima"] < 0)
    ck["is_canceled"] = False

    # deteccion de cancelacion (negativo posterior de igual magnitud, mismo base)
    neg = ck[ck["is_negative"] & ck["base"].notna()][["cedula10", "base", "emision", "suma_prima"]].copy()
    neg = neg.rename(columns={"emision": "em_neg"})
    neg["neg_abs"] = neg["suma_prima"].abs()
    pos = ck[(~ck["is_negative"]) & ck["base"].notna() & ck["suma_prima"].notna()][["cedula10", "base", "rpe", "emision", "suma_prima"]].copy()
    pos = pos.rename(columns={"emision": "em_pos"})
    pos["pos_abs"] = pos["suma_prima"].abs()
    if len(neg) > 0 and len(pos) > 0:
        pos = pos.reset_index(drop=True)
        pos["_pid"] = pos.index
        m = pos.merge(neg[["cedula10", "base", "em_neg", "neg_abs"]], on=["cedula10", "base"], how="inner")
        m = m[m["em_neg"] >= m["em_pos"]]
        if len(m) > 0:
            m = m.sort_values(["_pid", "em_neg"], kind="stable").drop_duplicates("_pid", keep="first")
            m["cancel_ok"] = (m["neg_abs"] - m["pos_abs"]).abs() <= np.maximum(tol_cancel_abs, tol_cancel_rel * m["pos_abs"])
            canc = m[m["cancel_ok"]][["cedula10", "rpe"]].drop_duplicates()
            if len(canc) > 0:
                canc_set = set(zip(canc["cedula10"], canc["rpe"]))
                ck["is_canceled"] = [ (c, r) in canc_set for c, r in zip(ck["cedula10"], ck["rpe"]) ]
    return ck


# ============================================================
# J) LEER VENTAS PARA MATCH (lectura unica de la hoja)
# ============================================================

def _detectar_cols_ventas_match(headers):
    """Devuelve posiciones 1-indexadas de cedula/prima/fecha/ramo en el maestro
    VENTAS, buscando por NOMBRE. Si no encuentra, cae a las posiciones fijas
    B/AC/M/U (configurables arriba). Para cedula prioriza '10 DIGITOS'."""
    norm = [normalize_name(h) for h in headers]

    def find(names):
        for nm in names:
            nn = normalize_name(nm)
            for i, h in enumerate(norm):
                if h == nn:
                    return i + 1  # 1-indexado
        return None

    pos_ced = find(["10 DIGITOS", "Cédula", "Cedula"])
    pos_pri = find(["Prima Neta"])
    pos_ing = find(["Fecha de Ingreso", "Fecha Ingreso", "Fecha Venta"])
    pos_ram = find(["Ramo"])
    # fallback a posiciones fijas configuradas
    if pos_ced is None: pos_ced = col2num(col_ventas_cedula)
    if pos_pri is None: pos_pri = col2num(col_ventas_prima)
    if pos_ing is None: pos_ing = col2num(col_ventas_ingreso)
    if pos_ram is None: pos_ram = col2num(col_ventas_ramo)
    return pos_ced, pos_pri, pos_ing, pos_ram


def leer_ventas_para_match():
    global ventas_row_fin
    # Detectar la fila real del header del maestro VENTAS (robusto: fila 1 o 2).
    hrow = detectar_fila_encabezado_ventas_master(file_ven, sheet_ventas, preferida=2)
    top = read_top_rows(file_ven, sheet_ventas, hrow, fill_merged=True)
    headers_master = [("" if v is None else str(v)) for v in (top[hrow - 1] if len(top) >= hrow else [])]
    colB, colAC, colM, colU = _detectar_cols_ventas_match(headers_master)
    maxc = max(colB, colAC, colM, colU)
    # Los datos del cruce empiezan en la fila SIGUIENTE al header (igual que R,
    # que arranca en ventas_row_ini=2 con el header en la fila 1). Asi no se
    # cuela la fila de encabezados como una venta fantasma.
    row_ini = hrow + 1
    msg("VENTAS match | fila header: ", hrow, " | col cedula: ", colB,
        " prima: ", colAC, " fecha: ", colM, " ramo: ", colU)

    rows_data = []  # (row_excel, ced, pri, ing, ram)
    if _HAS_CALAMINE:
        try:
            data_all = _calamine_read_all(file_ven, sheet_ventas)
            if data_all and len(data_all[0]) < colAC:
                raise RuntimeError(f"VENTAS tiene menos columnas de las esperadas (se requiere hasta col {colAC}).")
            for i in range(row_ini - 1, len(data_all)):
                row = data_all[i]
                ced = row[colB - 1] if colB - 1 < len(row) else None
                pri = row[colAC - 1] if colAC - 1 < len(row) else None
                ing = row[colM - 1] if colM - 1 < len(row) else None
                ram = row[colU - 1] if colU - 1 < len(row) else None
                rows_data.append((i + 1, ced, pri, ing, ram))
        except RuntimeError:
            raise
        except Exception:
            rows_data = None
    else:
        rows_data = None

    if rows_data is None:
        rows_data = []
        wb = load_workbook(file_ven, read_only=True, data_only=True)
        if sheet_ventas not in wb.sheetnames:
            wb.close()
            raise RuntimeError(f"No existe la hoja '{sheet_ventas}' en VENTAS.xlsx")
        ws = wb[sheet_ventas]
        if (ws.max_column or 0) < colAC:
            wb.close()
            raise RuntimeError(f"VENTAS tiene menos columnas de las esperadas (se requiere hasta col {colAC}).")
        r_idx = 0
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=maxc, values_only=True):
            r_idx += 1
            if r_idx < row_ini:
                continue
            ced = row[colB - 1] if colB - 1 < len(row) else None
            pri = row[colAC - 1] if colAC - 1 < len(row) else None
            ing = row[colM - 1] if colM - 1 < len(row) else None
            ram = row[colU - 1] if colU - 1 < len(row) else None
            rows_data.append((r_idx, ced, pri, ing, ram))
        wb.close()

    # ultima fila con dato en columna de cedula
    last = None
    for (rx, ced, _, _, _) in rows_data:
        if not is_blank_scalar(ced):
            last = rx
    if ventas_row_fin is None:
        if last is None:
            raise RuntimeError("No pude detectar el ultimo row en VENTAS.")
        rf = last
    else:
        rf = ventas_row_fin
    msg("VENTAS row_ini: ", row_ini, " | row_fin: ", rf)

    rows_data = [t for t in rows_data if t[0] <= rf]
    idxs = [t[0] for t in rows_data]
    ced_raw = [t[1] for t in rows_data]
    pri_raw = [t[2] for t in rows_data]
    ing_raw = [t[3] for t in rows_data]
    ram_raw = [t[4] for t in rows_data]

    ven = pd.DataFrame({
        "row_id": list(range(1, len(idxs) + 1)),
        "row_excel": idxs,
        "cedula_raw": [to_char_cell(v) for v in ced_raw],
        "prima_ventas_raw": [to_char_cell(v) for v in pri_raw],
        "ingreso_raw": [to_char_cell(v) for v in ing_raw],
        "ramo_raw": [("" if v is None else str(v)) for v in ram_raw],
    })
    ven["cedula10"] = clean_cedula10(pd.Series(ced_raw)).values
    ven["prima_ventas"] = to_num(pd.Series(pri_raw)).values
    ven["ingreso"] = pd.to_datetime(parse_fecha(pd.Series(ing_raw)).values, errors="coerce")
    ven["ramo_code"] = [ramo_code_from_text_scalar(v) for v in ram_raw]

    def gk(r):
        rc = r["ramo_code"]
        rc_ok = (rc is not None) and not (isinstance(rc, float) and math.isnan(rc))
        if (not is_blank_scalar(r["cedula10"])) and (not pd.isna(r["ingreso"])) and rc_ok:
            return f"{r['cedula10']}|{pd.Timestamp(r['ingreso']).strftime('%Y-%m-%d')}|{int(rc)}"
        return f"ROW|{r['row_excel']}"
    ven["group_key"] = ven.apply(gk, axis=1)
    ven["group_n"] = ven.groupby("group_key")["group_key"].transform("size")
    return ven


# ============================================================
# K) MOTOR DE MATCH
# ============================================================

_ADDM_CACHE = {}


def _add_months_vec(ts_series, n):
    """Suma n meses a una serie de timestamps, vectorizado, con cache por
    fecha unica (acelera cuando hay fechas repetidas). Equivale a
    date + relativedelta(months=n)."""
    ts = pd.to_datetime(ts_series, errors="coerce")
    uniq = pd.unique(ts.dropna())
    m = {}
    for u in uniq:
        key = (u, n)
        v = _ADDM_CACHE.get(key)
        if v is None:
            v = pd.Timestamp(u) + pd.DateOffset(months=n)
            if len(_ADDM_CACHE) < 50000:
                _ADDM_CACHE[key] = v
        m[u] = v
    return ts.map(lambda x: m.get(x, pd.NaT) if not pd.isna(x) else pd.NaT)


def build_single_edges_one_cedula(rows_sd, com_sub, used_rpes):
    rows0 = rows_sd[
        rows_sd["cedula10"].notna() & (rows_sd["cedula10"] != "") &
        rows_sd["prima_ventas"].notna() & (rows_sd["prima_ventas"] > 0) &
        rows_sd["ingreso"].notna() & rows_sd["ramo_code"].notna()
    ][["row_id", "row_excel", "cedula10", "prima_ventas", "ingreso", "ramo_code"]].copy()
    if rows0.empty:
        return None
    cand0 = com_sub[
        (~com_sub["is_negative"]) & (~com_sub["is_canceled"]) &
        com_sub["rpe"].notna() & (com_sub["rpe"] != "") &
        (~com_sub["rpe"].isin(used_rpes)) &
        com_sub["suma_prima"].notna() & (com_sub["suma_prima"] > 0) &
        com_sub["emision"].notna()
    ][["rpe", "rpe_code", "emision", "fin_vig", "suma_prima", "suma_comis"]].copy()
    if cand0.empty:
        return None

    rows0["_k"] = 1
    cand0["_k"] = 1
    ed = rows0.merge(cand0, on="_k").drop(columns="_k")
    # limite vectorizado (ingreso + 11 meses)
    ed["lim"] = _add_months_vec(ed["ingreso"], max_meses)
    ed = ed[(ed["emision"] >= ed["ingreso"]) & (ed["emision"] <= ed["lim"]) & (ed["rpe_code"] == ed["ramo_code"])]
    if ed.empty:
        return None
    ed["absd"] = (ed["suma_prima"] - ed["prima_ventas"]).abs()
    ed["rel"] = ed["absd"] / ed["prima_ventas"]
    ed = ed[(ed["rel"] <= tol_prima) | (ed["absd"] <= tol_abs)]
    if ed.empty:
        return None
    # days_gap vectorizado: max(0, (emision - ingreso) en dias)
    gap = (pd.to_datetime(ed["emision"]).dt.normalize() - pd.to_datetime(ed["ingreso"]).dt.normalize()).dt.days
    ed["days_gap"] = gap.clip(lower=0)
    ed["absd_cent"] = ed["absd"].round(2)
    ed["row_deg"] = ed.groupby("row_id")["rpe"].transform("nunique")
    ed["cand_deg"] = ed.groupby("rpe")["row_id"].transform("nunique")

    if HABILITAR_PRIORIDAD_ANTIGUEDAD_MATCH:
        claim = ed[["rpe", "row_id", "ingreso", "row_excel"]].drop_duplicates()
        claim = claim.sort_values(["rpe", "ingreso", "row_excel"], kind="stable")
        claim["older_claims_for_rpe"] = claim.groupby("rpe").cumcount()
        ed = ed.merge(claim[["rpe", "row_id", "older_claims_for_rpe"]], on=["rpe", "row_id"], how="left")
        ed["older_claims_for_rpe"] = ed["older_claims_for_rpe"].fillna(0).astype(int)
    else:
        ed["older_claims_for_rpe"] = 0

    # emis_mod vectorizado: dias desde epoch % 10000
    emis_days = (pd.to_datetime(ed["emision"]).dt.normalize() - pd.Timestamp(EPOCH)).dt.days
    ed["emis_mod"] = emis_days % 10000
    ed["cost"] = (ed["absd_cent"] * 1e8 + ed["absd"] * 1e5 +
                  ed["older_claims_for_rpe"] * PRIORIDAD_ANTIGUEDAD_COST +
                  ed["days_gap"] * 1e2 + ed["cand_deg"] * 1e-1 + ed["emis_mod"] * 1e-6)
    ed = ed.sort_values(["row_id", "cost", "rpe"], kind="stable").reset_index(drop=True)
    return ed


def solve_lsap(ed, all_row_ids):
    if ed is None or ed.empty:
        return pd.DataFrame({"row_id": [], "rpe": []})
    rows = sorted(set(all_row_ids))
    cands = sorted(ed["rpe"].unique())
    nr, nc = len(rows), len(cands)
    ed_min = ed.sort_values("cost", kind="stable").drop_duplicates(["row_id", "rpe"], keep="first")
    cost = np.full((nr, nc + nr), IMPOSSIBLE_COST, dtype=float)
    cost[:, nc:nc + nr] = DUMMY_UNMATCH_COST
    ridx = {r: i for i, r in enumerate(rows)}
    cidx = {c: j for j, c in enumerate(cands)}
    for _, rr in ed_min.iterrows():
        cost[ridx[rr["row_id"]], cidx[rr["rpe"]]] = rr["cost"]
    ri, ci = linear_sum_assignment(cost)
    out = [(rows[i], cands[j]) for i, j in zip(ri, ci) if j < nc]
    out = pd.DataFrame(out, columns=["row_id", "rpe"])
    if out.empty:
        return out
    valid = ed_min[["row_id", "rpe"]].drop_duplicates()
    return out.merge(valid, on=["row_id", "rpe"], how="inner").drop_duplicates()


def solve_greedy(ed):
    if ed is None or ed.empty:
        return pd.DataFrame({"row_id": [], "rpe": []})
    e = ed.copy()
    e["row_deg"] = e.groupby("row_id")["rpe"].transform("nunique")
    e["cand_deg"] = e.groupby("rpe")["row_id"].transform("nunique")
    e = e.sort_values(["row_deg", "cand_deg", "cost", "row_id", "rpe"], kind="stable")
    ur, up, keep = set(), set(), []
    for _, r in e.iterrows():
        if r["row_id"] not in ur and r["rpe"] not in up:
            keep.append((r["row_id"], r["rpe"]))
            ur.add(r["row_id"]); up.add(r["rpe"])
    return pd.DataFrame(keep, columns=["row_id", "rpe"]).drop_duplicates()


def solve_one_cedula_fast_1row(row, com_sub):
    """Fast-path EQUIVALENTE para una cedula con UNA sola fila de venta.
    Replica exactamente el filtrado y el costo de build_single_edges_one_cedula
    y la eleccion del LSAP (que para 1 fila = candidato de menor costo, con
    desempate estable por rpe). Devuelve el mismo dict que solve_one_cedula."""
    empty_assign = pd.DataFrame(columns=["row_id", "rpe", "emision", "fin_vig", "suma_prima", "suma_comis", "cost"])
    empty_counts = pd.DataFrame({"row_id": [], "single_cand_n": []})

    ced = row["cedula10"]
    pv = row["prima_ventas"]
    ing = row["ingreso"]
    rc = row["ramo_code"]
    rid = row["row_id"]
    rc_ok = (rc is not None) and not (isinstance(rc, float) and math.isnan(rc))
    if is_blank_scalar(ced) or pd.isna(pv) or pv <= 0 or pd.isna(ing) or not rc_ok:
        return None  # cae al camino normal

    if com_sub is None or len(com_sub) == 0:
        return {"assign": empty_assign, "counts": empty_counts}

    c = com_sub
    mask = ((~c["is_negative"].values) & (~c["is_canceled"].values) &
            c["rpe"].notna().values & (c["rpe"].values != "") &
            c["suma_prima"].notna().values & (c["suma_prima"].values > 0) &
            c["emision"].notna().values &
            (c["rpe_code"].values == rc))
    if not mask.any():
        return {"assign": empty_assign, "counts": empty_counts}
    sub = c[mask]
    emis = pd.to_datetime(sub["emision"].values)
    ingt = pd.Timestamp(ing)
    lim = ingt + pd.DateOffset(months=max_meses)
    in_window = np.asarray((emis >= ingt) & (emis <= lim))
    if not in_window.any():
        return {"assign": empty_assign, "counts": empty_counts}
    sub = sub[in_window]
    emis = emis[in_window]
    sp = sub["suma_prima"].values.astype(float)
    absd = np.abs(sp - pv)
    rel = absd / pv
    tol_ok = (rel <= tol_prima) | (absd <= tol_abs)
    if not tol_ok.any():
        return {"assign": empty_assign, "counts": empty_counts}
    sub = sub[tol_ok]
    emis = emis[tol_ok]
    sp = sp[tol_ok]
    absd = absd[tol_ok]
    rpes = sub["rpe"].values
    single_cand_n = len(pd.unique(rpes))

    days_gap = (emis.normalize() - ingt.normalize()).days
    days_gap = np.clip(np.asarray(days_gap, dtype=float), 0, None)
    absd_cent = np.round(absd, 2)
    emis_days = (emis.normalize() - pd.Timestamp(EPOCH)).days
    emis_mod = np.asarray(emis_days, dtype=float) % 10000
    cand_deg = 1.0
    older = 0.0
    cost = (absd_cent * 1e8 + absd * 1e5 + older * PRIORIDAD_ANTIGUEDAD_COST +
            days_gap * 1e2 + cand_deg * 1e-1 + emis_mod * 1e-6)

    order = np.lexsort((rpes, cost))  # por cost, desempate por rpe
    best = order[0]
    fv_vals = sub["fin_vig"].values
    sc_vals = sub["suma_comis"].values
    assign = pd.DataFrame([{
        "row_id": rid,
        "rpe": rpes[best],
        "emision": pd.Timestamp(emis[best]),
        "fin_vig": (pd.Timestamp(fv_vals[best]) if not pd.isna(fv_vals[best]) else pd.NaT),
        "suma_prima": float(sp[best]),
        "suma_comis": (float(sc_vals[best]) if not pd.isna(sc_vals[best]) else float("nan")),
        "cost": float(cost[best]),
    }])
    counts = pd.DataFrame({"row_id": [rid], "single_cand_n": [single_cand_n]})
    return {"assign": assign, "counts": counts}


def solve_one_cedula(rows_sd, com_sub):
    ed = build_single_edges_one_cedula(rows_sd, com_sub, set())
    if ed is None or ed.empty:
        return {"assign": pd.DataFrame(columns=["row_id", "rpe", "emision", "fin_vig", "suma_prima", "suma_comis", "cost"]),
                "counts": pd.DataFrame({"row_id": [], "single_cand_n": []})}
    counts = ed.groupby("row_id")["rpe"].nunique().reset_index().rename(columns={"rpe": "single_cand_n"})
    nr = rows_sd["row_id"].nunique()
    nc = ed["rpe"].nunique()
    if USAR_LSAP_POR_CEDULA and nr <= LSAP_MAX_FILAS_CEDULA and nc <= LSAP_MAX_CANDIDATOS_CEDULA:
        assign = solve_lsap(ed, rows_sd["row_id"].tolist())
    else:
        assign = solve_greedy(ed)
    det = assign.merge(ed[["row_id", "rpe", "emision", "fin_vig", "suma_prima", "suma_comis", "cost"]],
                       on=["row_id", "rpe"], how="left")
    if not det.empty:
        det = det.sort_values("cost", kind="stable").drop_duplicates("row_id", keep="first")
    return {"assign": det, "counts": counts}


def _window(com_sub, ing, lim, used_rpes):
    return com_sub[
        (~com_sub["is_negative"]) & (~com_sub["is_canceled"]) &
        com_sub["suma_prima"].notna() & (com_sub["suma_prima"] > 0) &
        com_sub["emision"].notna() & (com_sub["emision"] >= ing) & (com_sub["emision"] <= lim) &
        (~com_sub["rpe"].isin(used_rpes))
    ]


def try_pair_asistencia(c0, target, ing, ramo_c, used_rpes):
    if not HABILITAR_PAIR_ASISTENCIA:
        return None
    if pd.isna(target) or target <= 0 or pd.isna(ing) or ramo_c is None:
        return None
    lim = add_months(ing, max_meses)
    c_all = _window(c0, ing, lim, used_rpes)
    if c_all.empty:
        return None
    main = c_all[c_all["rpe_code"] == ramo_c].copy()
    other = c_all[c_all["rpe_code"] != ramo_c].copy()
    if main.empty or other.empty:
        return None
    main["absd"] = (main["suma_prima"] - target).abs()
    main = main.sort_values(["absd", "emision"], kind="stable")
    if len(main) > MAIN_TOP:
        main = main.iloc[:MAIN_TOP]
    other = other[other["suma_prima"] <= ASISTENCIA_MAX_FRAC * target].copy()
    if other.empty:
        return None
    other["abs0"] = (other["suma_prima"] - (target / 10)).abs()
    other = other.sort_values(["abs0", "emision"], kind="stable")
    if len(other) > OTHER_TOP:
        other = other.iloc[:OTHER_TOP]
    best = None; best_abs = float("inf"); best_days = float("inf")
    tol_sum = max(tol_abs, tol_prima * target)
    for _, m1 in main.iterrows():
        rem = target - m1["suma_prima"]
        if rem <= 0:
            continue
        oth = other[(other["rpe"] != m1["rpe"]) & ((other["suma_prima"] - rem).abs() <= tol_sum)].copy()
        if oth.empty:
            continue
        oth["daygap"] = oth["emision"].map(lambda e: abs(days_between(e, m1["emision"])))
        oth = oth[oth["daygap"] <= MAX_DIAS_ASISTENCIA]
        if oth.empty:
            continue
        oth["sumdiff"] = ((m1["suma_prima"] + oth["suma_prima"]) - target).abs()
        oth = oth.sort_values(["sumdiff", "daygap", "emision"], kind="stable")
        cand = oth.iloc[0]
        if cand["sumdiff"] < best_abs or (cand["sumdiff"] == best_abs and cand["daygap"] < best_days):
            best_abs = cand["sumdiff"]; best_days = cand["daygap"]
            best = {"main": m1, "assist": cand}
            if best_abs == 0:
                break
    return best


def try_ajuste_posterior(c0, target, ing, ramo_c, used_rpes):
    if not HABILITAR_AJUSTE_POSTERIOR:
        return None
    if pd.isna(target) or target <= 0 or pd.isna(ing) or ramo_c is None:
        return None
    lim = add_months(ing, max_meses)
    pos = c0[
        (~c0["is_negative"]) & (~c0["is_canceled"]) & c0["base"].notna() &
        c0["suma_prima"].notna() & (c0["suma_prima"] > 0) &
        c0["emision"].notna() & (c0["emision"] >= ing) & (c0["emision"] <= lim) &
        (c0["rpe_code"] == ramo_c) & (~c0["rpe"].isin(used_rpes))
    ].copy()
    if pos.empty:
        return None
    pos["absd_target"] = (pos["suma_prima"] - target).abs()
    pos["rel_target"] = pos["absd_target"] / target
    near_cap = max(tol_abs + 0.01, min(AJUSTE_POS_MAX_ABS, AJUSTE_POS_MAX_REL * target))
    pos = pos[(pos["suma_prima"] > target) &
              ((pos["rel_target"] > tol_prima) | (pos["absd_target"] > tol_abs)) &
              (pos["absd_target"] <= near_cap)]
    if pos.empty:
        return None
    pos = pos.sort_values(["absd_target", "emision"], kind="stable")
    if len(pos) > AJUSTE_POS_TOP:
        pos = pos.iloc[:AJUSTE_POS_TOP]
    neg = c0[c0["is_negative"] & c0["base"].notna() & c0["suma_prima"].notna() & c0["emision"].notna()]
    if neg.empty:
        return None
    best = None; best_abs_adj = float("inf"); best_gap = float("inf"); best_posmiss = float("inf")
    for _, p in pos.iterrows():
        n1 = neg[(neg["base"] == p["base"]) & (neg["emision"] >= p["emision"]) &
                 (neg["emision"] <= (p["emision"] + pd.Timedelta(days=AJUSTE_NEG_MAX_DIAS)))].copy()
        if n1.empty:
            continue
        n1["adj_prima"] = p["suma_prima"] + n1["suma_prima"]
        n1["absd_adj"] = (n1["adj_prima"] - target).abs()
        n1["rel_adj"] = n1["absd_adj"] / target
        n1["gap_days"] = n1["emision"].map(lambda e: days_between(e, p["emision"]))
        n1 = n1[(n1["rel_adj"] <= tol_prima) | (n1["absd_adj"] <= tol_abs)]
        if n1.empty:
            continue
        n1 = n1.sort_values(["absd_adj", "gap_days"], kind="stable")
        cand = n1.iloc[0]
        if (cand["absd_adj"] < best_abs_adj or
                (cand["absd_adj"] == best_abs_adj and cand["gap_days"] < best_gap) or
                (cand["absd_adj"] == best_abs_adj and cand["gap_days"] == best_gap and p["absd_target"] < best_posmiss)):
            best = {"pos": p, "neg": cand}
            best_abs_adj = cand["absd_adj"]; best_gap = cand["gap_days"]; best_posmiss = p["absd_target"]
    if best is None:
        return None
    p, n = best["pos"], best["neg"]
    detalle = (f"{EXC_01_CODE} | RPE positivo={p['rpe']} prima={fmt_num2(p['suma_prima'])} | "
               f"ajuste negativo={n['rpe']} prima={fmt_num2(n['suma_prima'])} | "
               f"neto={fmt_num2(n['adj_prima'])} | target={fmt_num2(target)}")
    return {"ok": True, "rpe1": str(p["rpe"]), "rpe2": None, "em": p["emision"], "fv": p["fin_vig"],
            "p1": float(p["suma_prima"]), "p2": float("nan"), "com": float(p["suma_comis"]), "multi": 1,
            "exc_code": EXC_01_CODE, "exc_detalle": detalle, "revision_manual": True}


def try_cancelado_posterior(c0, target, ing, ramo_c, used_rpes):
    if not HABILITAR_CANCELADO_POSTERIOR:
        return None
    if pd.isna(target) or target <= 0 or pd.isna(ing):
        return None
    lim = add_months(ing, max_meses)
    pos = c0[
        (~c0["is_negative"]) & (c0["is_canceled"]) & c0["base"].notna() &
        c0["suma_prima"].notna() & (c0["suma_prima"] > 0) &
        c0["emision"].notna() & (c0["emision"] >= ing) & (c0["emision"] <= lim) &
        (c0["rpe_code"] == ramo_c) & (~c0["rpe"].isin(used_rpes))
    ].copy()
    if pos.empty:
        return None
    pos["absd"] = (pos["suma_prima"] - target).abs()
    pos["rel"] = pos["absd"] / target
    pos = pos[(pos["rel"] <= tol_prima) | (pos["absd"] <= tol_abs)]
    if pos.empty:
        return None
    pos = pos.sort_values(["absd", "emision"], kind="stable")
    best = None; best_abs = float("inf"); best_gap = float("inf")
    for _, p in pos.iterrows():
        n1 = c0[c0["is_negative"] & c0["base"].notna() & (c0["base"] == p["base"]) &
                c0["emision"].notna() & (c0["emision"] >= p["emision"]) &
                (c0["emision"] <= (p["emision"] + pd.Timedelta(days=CANCELADO_POST_MAX_DIAS)))].copy()
        if n1.empty:
            continue
        n1["abs_neg"] = (n1["suma_prima"].abs() - p["suma_prima"]).abs()
        n1 = n1[n1["abs_neg"] <= np.maximum(tol_cancel_abs, tol_cancel_rel * p["suma_prima"])]
        if n1.empty:
            continue
        n1["gap_days"] = n1["emision"].map(lambda e: days_between(e, p["emision"]))
        n1 = n1.sort_values(["gap_days", "abs_neg"], kind="stable")
        cand = n1.iloc[0]
        if p["absd"] < best_abs or (p["absd"] == best_abs and cand["gap_days"] < best_gap):
            best = {"pos": p, "neg": cand}; best_abs = p["absd"]; best_gap = cand["gap_days"]
    if best is None:
        return None
    p, n = best["pos"], best["neg"]
    detalle = (f"{EXC_03_CODE} | RPE positivo={p['rpe']} prima={fmt_num2(p['suma_prima'])} | "
               f"devolucion posterior={n['rpe']} prima={fmt_num2(n['suma_prima'])} | "
               f"match exacto, pero con cancelacion posterior del mismo base")
    return {"ok": True, "rpe1": str(p["rpe"]), "rpe2": None, "em": p["emision"], "fv": p["fin_vig"],
            "p1": float(p["suma_prima"]), "p2": float("nan"), "com": float(p["suma_comis"]), "multi": 1,
            "exc_code": EXC_03_CODE, "exc_detalle": detalle, "revision_manual": True}


def try_emision_antes_ingreso_1d(c0, target, ing, ramo_c, used_rpes):
    if not HABILITAR_EMISION_ANTES_INGRESO_1D:
        return None
    if pd.isna(target) or target <= 0 or pd.isna(ing) or ramo_c is None:
        return None
    lim = add_months(ing, max_meses)
    cand = c0[
        (~c0["is_negative"]) & (~c0["is_canceled"]) &
        c0["suma_prima"].notna() & (c0["suma_prima"] > 0) &
        c0["emision"].notna() & (c0["emision"] < ing) &
        (c0["emision"] >= (ing - pd.Timedelta(days=EMISION_ANTES_INGRESO_MAX_DIAS))) &
        (c0["emision"] <= lim) & (c0["rpe_code"] == ramo_c) & (~c0["rpe"].isin(used_rpes))
    ].copy()
    if cand.empty:
        return None
    cand["absd"] = (cand["suma_prima"] - target).abs()
    cand["rel"] = cand["absd"] / target
    cand = cand[(cand["rel"] <= tol_prima) | (cand["absd"] <= tol_abs)]
    if cand.empty:
        return None
    cand["dias_antes"] = cand["emision"].map(lambda e: days_between(ing, e))
    cand = cand.sort_values(["absd", "dias_antes", "emision"], kind="stable")
    b = cand.iloc[0]
    detalle = (f"{EXC_04_CODE} | RPE={b['rpe']} | emision={pd.Timestamp(b['emision']).strftime('%d/%m/%Y')} | "
               f"ingreso={pd.Timestamp(ing).strftime('%d/%m/%Y')} | dias_diff={days_between(ing, b['emision'])} | "
               f"prima_rpe={fmt_num2(b['suma_prima'])} | target={fmt_num2(target)} | "
               f"motivo=Emision registrada 1 dia antes del ingreso")
    return {"ok": True, "rpe1": str(b["rpe"]), "rpe2": None, "em": b["emision"], "fv": b["fin_vig"],
            "p1": float(b["suma_prima"]), "p2": float("nan"), "com": float(b["suma_comis"]), "multi": 1,
            "exc_code": EXC_04_CODE, "exc_detalle": detalle, "revision_manual": True}


def match_one_row(c0, target, ing, ramo_c, used_rpes):
    if c0 is None or len(c0) == 0:
        return {"ok": False, "reason": "1) no se encontro cedula"}
    if pd.isna(ing) or pd.isna(target) or target == 0:
        return {"ok": False, "reason": "(ventas invalido) fecha/prima neta invalida"}
    # ramo_code llega como NaN (float) cuando el ramo no esta mapeado, porque
    # pandas convierte None->NaN en columnas numericas. Normalizar a None para
    # que 'is None' funcione igual que is.na() en R (si no, las ventas con ramo
    # no mapeado caen erroneamente en "RAMO no coincide" en vez de la rama de
    # match sin validar prefijo).
    if ramo_c is not None and isinstance(ramo_c, float) and math.isnan(ramo_c):
        ramo_c = None
    lim = add_months(ing, max_meses)
    c_date_all = _window(c0, ing, lim, used_rpes)

    if ramo_c is None:
        if c_date_all.empty:
            ex3 = try_cancelado_posterior(c0, target, ing, 999999, used_rpes)
            if ex3 is not None:
                return ex3
            return {"ok": False, "reason": "3) no hay fecha (<= 11 meses y EMISION >= INGRESO)"}
        c = c_date_all.copy()
        c["absd"] = (c["suma_prima"] - target).abs()
        c["rel"] = c["absd"] / target
        c = c[(c["rel"] <= tol_prima) | (c["absd"] <= tol_abs)]
        if not c.empty:
            c = c.sort_values(["absd", "emision"], kind="stable")
            b = c.iloc[0]
            return {"ok": True, "rpe1": b["rpe"], "rpe2": None, "em": b["emision"], "fv": b["fin_vig"],
                    "p1": b["suma_prima"], "p2": float("nan"), "com": b["suma_comis"], "multi": 1,
                    "anom": True, "anom_det": "RAMO en VENTAS no mapeado; match sin validar prefijo RPE.",
                    "exc_code": "", "exc_detalle": "", "revision_manual": False}
        return {"ok": False, "reason": "2) no hay prima (+-3%)", "anom": True, "anom_det": "RAMO no mapeado"}

    c_main = c_date_all[c_date_all["rpe_code"] == ramo_c] if not c_date_all.empty else c_date_all
    if not c_main.empty:
        cm = c_main.copy()
        cm["absd"] = (cm["suma_prima"] - target).abs()
        cm["rel"] = cm["absd"] / target
        cpri = cm[(cm["rel"] <= tol_prima) | (cm["absd"] <= tol_abs)]
        if not cpri.empty:
            cpri = cpri.sort_values(["absd", "emision"], kind="stable")
            b = cpri.iloc[0]
            return {"ok": True, "rpe1": b["rpe"], "rpe2": None, "em": b["emision"], "fv": b["fin_vig"],
                    "p1": b["suma_prima"], "p2": float("nan"), "com": b["suma_comis"], "multi": 1,
                    "exc_code": "", "exc_detalle": "", "revision_manual": False}

    if (not c_main.empty) and HABILITAR_PAIR_ASISTENCIA:
        best = try_pair_asistencia(c0, target, ing, ramo_c, used_rpes)
        if best is not None:
            m1, a2 = best["main"], best["assist"]
            return {"ok": True, "rpe1": m1["rpe"], "rpe2": a2["rpe"], "em": m1["emision"], "fv": m1["fin_vig"],
                    "p1": m1["suma_prima"], "p2": a2["suma_prima"], "com": m1["suma_comis"] + a2["suma_comis"],
                    "multi": 2, "exc_code": "", "exc_detalle": "", "revision_manual": False}

    ex1 = try_ajuste_posterior(c0, target, ing, ramo_c, used_rpes)
    if ex1 is not None:
        return ex1
    ex3 = try_cancelado_posterior(c0, target, ing, ramo_c, used_rpes)
    if ex3 is not None:
        return ex3
    ex4 = try_emision_antes_ingreso_1d(c0, target, ing, ramo_c, used_rpes)
    if ex4 is not None:
        return ex4

    if c_date_all.empty:
        return {"ok": False, "reason": "3) no hay fecha (<= 11 meses y EMISION >= INGRESO)"}
    if c_main.empty:
        c_alt = c_date_all.copy()
        c_alt["absd"] = (c_alt["suma_prima"] - target).abs()
        c_alt["rel"] = c_alt["absd"] / target
        c_alt = c_alt[(c_alt["rel"] <= tol_prima) | (c_alt["absd"] <= tol_abs)]
        cand_alt = (c_alt.sort_values(["absd", "emision"], kind="stable").iloc[0]["rpe"]) if not c_alt.empty else None
        return {"ok": False, "reason": "4) RAMO no coincide (prefijo RPE vs RAMO VENTAS)", "candidato_sin_ramo": cand_alt}
    return {"ok": False, "reason": "2) no hay prima (+-3%) que corresponda a la cedula"}


# ============================================================
# L) EJECUTAR MATCH
# ============================================================

def ejecutar_match(ven, com_key):
    res = {}
    for _, r in ven.iterrows():
        rid = r["row_id"]
        res[rid] = dict(
            row_id=rid, row_excel=r["row_excel"], group_n=r["group_n"],
            cedula10=r["cedula10"], cedula_raw=r["cedula_raw"], ramo_raw=r["ramo_raw"], ramo_code=r["ramo_code"],
            prima_ventas=r["prima_ventas"], prima_ventas_raw=r["prima_ventas_raw"],
            ingreso=r["ingreso"], ingreso_raw=r["ingreso_raw"],
            RPE1=None, RPE2=None, RPE3=None, emision=pd.NaT, fin_vig=pd.NaT,
            prima1=float("nan"), prima2=float("nan"), prima3=float("nan"),
            comision=float("nan"), multi_n=0, ramo_ok=None,
            razon_no_rpe="", candidato_sin_ramo=None, anomalia=False, anomalia_detalle="",
            exc_code="", exc_detalle="", revision_manual=False, single_cand_n=None,
            excepcion_codigo="", detalle_excepcion="", revision_manual_txt="",
        )

    com_by_ced = {ced: g for ced, g in com_key.groupby("cedula10", sort=False)} if len(com_key) else {}
    ven_valid = ven[ven["cedula10"].notna() & (ven["cedula10"] != "")]
    ven_by_ced = {ced: g for ced, g in ven_valid.groupby("cedula10", sort=False)}
    valid_ceds = list(dict.fromkeys(ven_valid["cedula10"].tolist()))

    empty_com = com_key.iloc[0:0]

    _total = len(valid_ceds)
    for _i, ced in enumerate(valid_ceds):
        if MOSTRAR_PROGRESO and _i > 0 and (_i % 5000 == 0):
            msg(f"  ... cruce en progreso: {_i}/{_total} cedulas")
        rows_sd = ven_by_ced[ced]
        com_sub = com_by_ced.get(ced, empty_com)
        # Fast-path para cedulas con UNA sola fila de venta (la mayoria).
        # Validado como equivalente al camino LSAP; cae al normal si no aplica.
        solved = None
        if len(rows_sd) == 1:
            solved = solve_one_cedula_fast_1row(rows_sd.iloc[0], com_sub)
        if solved is None:
            solved = solve_one_cedula(rows_sd, com_sub)

        for _, cr in solved["counts"].iterrows():
            res[cr["row_id"]]["single_cand_n"] = int(cr["single_cand_n"])

        assigned_rows = set()
        if not solved["assign"].empty:
            for _, a in solved["assign"].iterrows():
                rid = a["row_id"]
                res[rid].update(RPE1=a["rpe"], emision=a["emision"], fin_vig=a["fin_vig"],
                                prima1=a["suma_prima"], comision=a["suma_comis"], multi_n=1, ramo_ok=True)
                assigned_rows.add(rid)

        used_rpes_ced = set(solved["assign"]["rpe"].tolist()) if not solved["assign"].empty else set()
        rem = rows_sd[~rows_sd["row_id"].isin(assigned_rows)].copy()
        if rem.empty:
            continue
        rem["single_cand_n"] = rem["row_id"].map(lambda rid: res[rid]["single_cand_n"])
        rem["single_cand_n"] = rem["single_cand_n"].fillna(999999).astype(int)
        rem["ingreso_ord"] = rem["ingreso"].fillna(pd.Timestamp("2999-12-31"))
        rem = rem.sort_values(["single_cand_n", "ingreso_ord", "row_excel"], kind="stable")

        rows_sd_by_id = {row["row_id"]: row for _, row in rows_sd.iterrows()}
        for _, rr0 in rem.iterrows():
            rid = rr0["row_id"]
            if res[rid]["RPE1"] not in (None, ""):
                continue
            rowk = rows_sd_by_id[rid]
            rr = match_one_row(com_sub, rowk["prima_ventas"], rowk["ingreso"], rowk["ramo_code"], used_rpes_ced)
            if not rr.get("ok"):
                res[rid]["razon_no_rpe"] = rr.get("reason", "")
                if "candidato_sin_ramo" in rr and rr["candidato_sin_ramo"] is not None:
                    res[rid]["candidato_sin_ramo"] = rr["candidato_sin_ramo"]
                if rr.get("anom"):
                    res[rid]["anomalia"] = True
                    res[rid]["anomalia_detalle"] = rr.get("anom_det", "")
            else:
                res[rid].update(
                    RPE1=(None if rr["rpe1"] is None else str(rr["rpe1"])),
                    RPE2=(None if rr["rpe2"] is None else str(rr["rpe2"])),
                    emision=pd.Timestamp(rr["em"]) if not pd.isna(rr["em"]) else pd.NaT,
                    fin_vig=pd.Timestamp(rr["fv"]) if not pd.isna(rr["fv"]) else pd.NaT,
                    prima1=rr["p1"], prima2=rr["p2"], comision=rr["com"], multi_n=rr["multi"], ramo_ok=True)
                if rr.get("exc_code"):
                    res[rid].update(exc_code=rr["exc_code"], exc_detalle=rr["exc_detalle"],
                                    revision_manual=bool(rr.get("revision_manual")),
                                    excepcion_codigo=rr["exc_code"], detalle_excepcion=rr["exc_detalle"],
                                    revision_manual_txt=("SI" if rr.get("revision_manual") else "NO"))
                nuevos = [x for x in (rr["rpe1"], rr["rpe2"]) if x is not None and x != ""]
                used_rpes_ced |= set(nuevos)

    # filas con cedula invalida: solo razon
    invalid = ven[ven["cedula10"].isna() | (ven["cedula10"] == "")]
    for _, r in invalid.iterrows():
        rr = match_one_row(empty_com, r["prima_ventas"], r["ingreso"], r["ramo_code"], set())
        if not rr.get("ok"):
            res[r["row_id"]]["razon_no_rpe"] = rr.get("reason", "")

    res_df = pd.DataFrame(list(res.values())).sort_values("row_excel", kind="stable").reset_index(drop=True)
    return res_df


# ============================================================
# M) ARMAR SALIDA (bloque + log)
# ============================================================

BLOQUE_COLS = [
    "mes_emision_rpe", "status_rpe_pago", "status_sva_reservado", "comentario_sva_reservado",
    "fecha_emision_rpe", "fecha_fin_vigencia_rpe", "mes_comision_rpe", "rpe_1_principal",
    "rpe_2_asistencia", "prima_neta_rpe_1", "prima_neta_rpe_2", "prima_neta_final_match",
    "comision_total_rpe", "validacion_prima_match", "diferencia_prima_rpe_vs_venta",
    "porcentaje_comision_sobre_prima", "rpe_3_reserva", "prima_neta_rpe_3_reserva",
    "cantidad_rpe_asignados", "excepcion_codigo", "detalle_excepcion", "revision_manual",
    "dictamen_color_powerbi",
]


def armar_bloque_y_log(res):
    res = res.copy()
    def prima_final(r):
        if not (isinstance(r["prima3"], float) and math.isnan(r["prima3"])):
            return r["prima1"] + r["prima2"] + r["prima3"]
        if not (isinstance(r["prima2"], float) and math.isnan(r["prima2"])):
            return r["prima1"] + r["prima2"]
        return r["prima1"]
    res["PRIMA_NETA_FINAL"] = res.apply(prima_final, axis=1)

    def dictamen(r):
        if not _rpe_vacio(r["exc_code"]):
            return "NARANJA_EXCEPCION_REVISION_MANUAL"
        if r["group_n"] >= 2:
            return "AZUL_GRUPO_DUPLICADO"
        if _rpe_vacio(r["RPE1"]):
            return "ROJO_SIN_RPE"
        return "SIN_COLOR_OK"
    res["dictamen_color_powerbi"] = res.apply(dictamen, axis=1)

    res["MES_DE_EMISION"] = mes_es(res["emision"]).values
    res["STATUS"] = res["RPE1"].map(lambda v: "" if _rpe_vacio(v) else "PAGADA")
    res["MES_COMISION"] = res["MES_DE_EMISION"]

    def validacion(r):
        pf, pv = r["PRIMA_NETA_FINAL"], r["prima_ventas"]
        if (isinstance(pf, float) and math.isnan(pf)) or (isinstance(pv, float) and math.isnan(pv)) or pv == 0:
            return None
        return abs((pf - pv) / pv) <= tol_prima
    res["VALIDACION"] = res.apply(validacion, axis=1)

    def dif(r):
        pf, pv = r["PRIMA_NETA_FINAL"], r["prima_ventas"]
        if (isinstance(pf, float) and math.isnan(pf)) or (isinstance(pv, float) and math.isnan(pv)):
            return None
        return pf - pv
    res["DIF"] = res.apply(dif, axis=1)

    def porc(r):
        com, pf = r["comision"], r["PRIMA_NETA_FINAL"]
        if (isinstance(com, float) and math.isnan(com)) or (isinstance(pf, float) and math.isnan(pf)) or pf == 0:
            return None
        return com / pf
    res["PORC_COMISION"] = res.apply(porc, axis=1)

    bloque = pd.DataFrame({
        "mes_emision_rpe": res["MES_DE_EMISION"],
        "status_rpe_pago": res["STATUS"],
        "status_sva_reservado": "",
        "comentario_sva_reservado": "",
        "fecha_emision_rpe": res["emision"],
        "fecha_fin_vigencia_rpe": res["fin_vig"],
        "mes_comision_rpe": res["MES_COMISION"],
        "rpe_1_principal": res["RPE1"],
        "rpe_2_asistencia": res["RPE2"],
        "prima_neta_rpe_1": res["prima1"],
        "prima_neta_rpe_2": res["prima2"],
        "prima_neta_final_match": res["PRIMA_NETA_FINAL"],
        "comision_total_rpe": res["comision"],
        "validacion_prima_match": res["VALIDACION"],
        "diferencia_prima_rpe_vs_venta": res["DIF"],
        "porcentaje_comision_sobre_prima": res["PORC_COMISION"],
        "rpe_3_reserva": res["RPE3"],
        "prima_neta_rpe_3_reserva": res["prima3"],
        "cantidad_rpe_asignados": res["multi_n"].astype("Int64"),
        "excepcion_codigo": res["excepcion_codigo"],
        "detalle_excepcion": res["detalle_excepcion"],
        "revision_manual": res["revision_manual_txt"],
        "dictamen_color_powerbi": res["dictamen_color_powerbi"],
    })[BLOQUE_COLS]

    mask_log = (
        res["RPE1"].isna() | (res["RPE1"] == "") |
        ((res["razon_no_rpe"] != "") & res["razon_no_rpe"].notna()) |
        (res["anomalia"]) | (res["group_n"] >= 2) |
        ((res["exc_code"] != "") & res["exc_code"].notna())
    )
    log_cols = ["row_excel", "group_n", "cedula10", "cedula_raw", "ramo_raw", "ramo_code",
                "prima_ventas", "prima_ventas_raw", "ingreso", "ingreso_raw",
                "RPE1", "RPE2", "RPE3", "emision", "fin_vig", "prima1", "prima2", "prima3",
                "PRIMA_NETA_FINAL", "comision", "VALIDACION", "DIF", "PORC_COMISION",
                "multi_n", "ramo_ok", "single_cand_n", "razon_no_rpe", "candidato_sin_ramo",
                "anomalia", "anomalia_detalle", "exc_code", "exc_detalle", "revision_manual",
                "excepcion_codigo", "detalle_excepcion", "revision_manual_txt", "dictamen_color_powerbi"]
    log_df = res[mask_log][log_cols].copy()
    return res, bloque, log_df


# ============================================================
# N) ESCRIBIR SALIDA DEL CRUCE
# ============================================================

COLUMNAS_LEGACY = [
    "%Comisión Venta", "$Comisión Venta", "Valida %", "Valida $", "Observacion Alice", "Ingreso",
    "Caso", "ESTADO TIPIFICADO", "ESTADO CHUBB", "MotivoCancelado", "Nombre UA", "Ramo",
    "Póliza", "Endoso", "CLAVE R-P-E", "cc", "Cédula Real", "# casos UA",
]


def escribir_salida(res, bloque):
    # Detectar la fila real de encabezados del maestro VENTAS (robusto).
    hrow = detectar_fila_encabezado_ventas_master(file_ven, sheet_ventas, preferida=2)
    top2 = read_top_rows(file_ven, sheet_ventas, hrow, fill_merged=True)
    if len(top2) < hrow:
        raise RuntimeError(f"VENTAS BASE no tiene la fila de encabezados esperada (fila {hrow}).")
    orig_headers = [("" if v is None else str(v)) for v in top2[hrow - 1]]
    ncol = len(orig_headers)

    norm_data = []
    if _HAS_CALAMINE:
        try:
            data_all = _calamine_read_all(file_ven, sheet_ventas)
            for i in range(hrow, len(data_all)):  # datos empiezan tras la fila de encabezados
                row = data_all[i]
                rr = list(row[:ncol])
                if len(rr) < ncol:
                    rr = rr + [None] * (ncol - len(rr))
                norm_data.append(rr)
        except Exception:
            norm_data = None
    else:
        norm_data = None

    if norm_data is None:
        norm_data = []
        wb_in = load_workbook(file_ven, read_only=True, data_only=True)
        ws_in = wb_in[sheet_ventas]
        r_idx = 0
        for row in ws_in.iter_rows(min_row=1, min_col=1, max_col=ncol, values_only=True):
            r_idx += 1
            if r_idx <= hrow:
                continue
            rr = list(row[:ncol])
            if len(rr) < ncol:
                rr = rr + [None] * (ncol - len(rr))
            norm_data.append(rr)
        wb_in.close()

    headers = list(orig_headers)
    data = [list(row) for row in norm_data]

    output_start = col2num(col_start_BN)  # 66
    legacy_norm = set(normalize_name(x) for x in COLUMNAS_LEGACY)

    # eliminar columnas legacy en posiciones >= output_start (1-indexado)
    keep_idx = []
    for j in range(len(headers)):
        pos1 = j + 1
        if pos1 >= output_start and normalize_name(headers[j]) in legacy_norm:
            continue
        keep_idx.append(j)
    headers = [headers[j] for j in keep_idx]
    data = [[row[j] for j in keep_idx] for row in data]

    # renombrar legacy AÑO / MES.AÑO si aparecen
    for j in range(len(headers)):
        nh = normalize_name(headers[j])
        if nh == normalize_name("AÑO"):
            headers[j] = "excepcion_codigo"
        elif nh == normalize_name("MES AÑO"):
            headers[j] = "detalle_excepcion"

    output_end = output_start + len(bloque.columns) - 1  # 88
    # asegurar columnas hasta output_end
    while len(headers) < output_end:
        headers.append(f"__BLANK_OUT_{len(headers)+1}")
        for row in data:
            row.append("")

    # set headers del bloque
    for k, name in enumerate(bloque.columns):
        headers[output_start - 1 + k] = name

    # parsear columnas de fecha originales (las del bloque se sobreescriben)
    hn = [normalize_name(h) for h in headers]
    date_cols_out = [j for j in range(len(headers))
                     if ("fecha" in hn[j]) or ("inicio" in hn[j]) or ("fin" in hn[j])]
    nrow = len(data)
    for j in date_cols_out:
        for i in range(nrow):
            data[i][j] = parse_fecha_scalar(data[i][j])

    # escribir bloque alineando por row_excel REAL (no por posicion), para que
    # nunca se corra si el header esta en fila 1 o en fila 2.
    # data[i] corresponde a la fila Excel (hrow + 1 + i) del maestro.
    bloque_idx = bloque.reset_index(drop=True)
    res_idx = res.reset_index(drop=True)
    for i in range(len(res_idx)):
        try:
            row_excel = int(res_idx.iloc[i]["row_excel"])
        except Exception:
            continue
        if row_excel <= hrow:   # fila de encabezado: no es dato
            continue
        data_i = row_excel - (hrow + 1)
        if data_i < 0 or data_i >= nrow:
            continue
        for k in range(len(bloque.columns)):
            data[data_i][output_start - 1 + k] = bloque_idx.iat[i, k]

    # construir workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_ventas
    ws.append(headers)
    for i in range(nrow):
        ws.append([coerce_cell_for_excel(v) for v in data[i]])

    bold = Font(bold=True)
    alc = Alignment(horizontal="center", vertical="center")
    for j in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=j)
        c.font = bold
        c.alignment = alc

    def set_fmt(col1, fmt, r0, r1):
        for r in range(r0, r1 + 1):
            ws.cell(row=r, column=col1).number_format = fmt

    # formato fecha original (filas 2..nrow+1)
    for j in date_cols_out:
        set_fmt(j + 1, "DD/MM/YYYY", 2, nrow + 1)

    bcol = {name: output_start + k for k, name in enumerate(bloque.columns)}  # 1-indexado
    # fechas del bloque (mismo formato que el resto: DD/MM/YYYY)
    for name in ("fecha_emision_rpe", "fecha_fin_vigencia_rpe"):
        set_fmt(bcol[name], "DD/MM/YYYY", 2, nrow + 1)
    # numeros del bloque
    for name in ("prima_neta_rpe_1", "prima_neta_rpe_2", "prima_neta_final_match",
                 "comision_total_rpe", "diferencia_prima_rpe_vs_venta", "prima_neta_rpe_3_reserva"):
        set_fmt(bcol[name], "#,##0.00", 2, nrow + 1)
    set_fmt(bcol["cantidad_rpe_asignados"], "0", 2, nrow + 1)
    set_fmt(bcol["porcentaje_comision_sobre_prima"], "0.00%", 2, nrow + 1)

    # colores (orden: rojo, azul, naranja -> el ultimo gana en celda solapada)
    fill_bad = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_group = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_exc = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")

    rpe1_col = bcol["rpe_1_principal"]
    dict_col = bcol["dictamen_color_powerbi"]
    color_start = rpe1_col
    color_end = output_end

    res_idx = res.reset_index(drop=True)
    # mapear row_excel REAL -> fila Excel de salida (header en fila 1 de salida).
    # data[i] (Excel maestro hrow+1+i) -> fila salida i+2.
    def out_row_for_res_row(r):
        try:
            row_excel = int(r["row_excel"])
        except Exception:
            return None
        if row_excel <= hrow:
            return None
        data_i = row_excel - (hrow + 1)
        if data_i < 0 or data_i >= nrow:
            return None
        return data_i + 2

    for _, r in res_idx.iterrows():
        out_r = out_row_for_res_row(r)
        if out_r is None:
            continue
        rpe1_blank = _rpe_vacio(r["RPE1"])
        if rpe1_blank:
            ws.cell(row=out_r, column=rpe1_col).fill = fill_bad
            ws.cell(row=out_r, column=dict_col).fill = fill_bad
    for _, r in res_idx.iterrows():
        out_r = out_row_for_res_row(r)
        if out_r is None:
            continue
        if r["group_n"] >= 2:
            for cc in range(color_start, color_end + 1):
                ws.cell(row=out_r, column=cc).fill = fill_group
    for _, r in res_idx.iterrows():
        out_r = out_row_for_res_row(r)
        if out_r is None:
            continue
        if r["exc_code"] not in (None, ""):
            for cc in range(color_start, color_end + 1):
                ws.cell(row=out_r, column=cc).fill = fill_exc

    ws.freeze_panes = "A2"
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 14
    for j in range(output_start, output_end + 1):
        ws.column_dimensions[get_column_letter(j)].width = 22

    if not os.path.isdir(carpeta_resultados):
        os.makedirs(carpeta_resultados, exist_ok=True)
    out_ventas = os.path.join(carpeta_resultados, "VENTAS_con_RPE_" + datetime.now().strftime("%Y%m%d") + ".xlsx")
    wb.save(out_ventas)
    return out_ventas


def escribir_log(log_df):
    out_log = os.path.join(carpeta_resultados, "LOG_RPE_" + datetime.now().strftime("%Y%m%d") + ".csv")
    df = log_df.copy()
    for c in df.columns:
        df[c] = df[c].map(lambda v: "" if (v is None or v is pd.NaT or (isinstance(v, float) and math.isnan(v))) else
                          (pd.Timestamp(v).strftime("%Y-%m-%d") if isinstance(v, (pd.Timestamp, datetime, date)) else v))
    df.to_csv(out_log, index=False, encoding="utf-8-sig")
    return out_log


# ============================================================
# G + MAIN) ORQUESTACION
# ============================================================

def main():
    if not os.path.isdir(carpeta):
        raise RuntimeError("No existe la carpeta LOOK: " + carpeta)
    if not os.path.isdir(carpeta_resultados):
        os.makedirs(carpeta_resultados, exist_ok=True)
        msg("Carpeta RESULTADOS creada: ", carpeta_resultados)

    if BORRAR_BACKUPS_VENTAS_EXISTENTES:
        for f in os.listdir(carpeta_resultados):
            if re.match(r"^BACKUP_VENTAS_.*\.xlsx$", f):
                try:
                    os.remove(os.path.join(carpeta_resultados, f))
                except Exception:
                    pass

    if not os.path.exists(file_ven):
        raise RuntimeError("No existe VENTAS.xlsx en LOOK: " + file_ven)
    if not os.path.exists(file_com):
        raise RuntimeError("No existe COMISIONES.xlsx en LOOK: " + file_com)

    msg("VENTAS maestro     : ", os.path.basename(file_ven))
    msg("COMISIONES maestro : ", os.path.basename(file_com))
    msg("PRE COMISIONES     : ", os.path.basename(file_pre_com) if os.path.exists(file_pre_com) else "(NO ENCONTRADO)")
    msg("PRE VENTAS         : ", os.path.basename(file_pre_ven) if os.path.exists(file_pre_ven) else "(NO ENCONTRADO)")

    comisiones_actualizadas = None
    if ACTUALIZAR_COMISIONES_CON_PRELIMINAR:
        comisiones_actualizadas = actualizar_archivo_comisiones()
    if ACTUALIZAR_VENTAS_CON_PRELIMINAR:
        actualizar_archivo_ventas()

    # H) comisiones para match
    if comisiones_actualizadas is not None:
        com = build_com_from_master(comisiones_actualizadas)
        del comisiones_actualizadas
    else:
        # Lector ligero: solo las columnas necesarias para el cruce.
        # Evita cargar las 61 columnas en memoria (clave con archivos grandes).
        com = build_com_for_match_light(file_com, sheet_comisiones, header_row=1)

    msg("Construyendo COM_KEY...")
    com_key = build_com_key(com)
    msg("COM_KEY filas: ", len(com_key))

    msg("Leyendo VENTAS para match...")
    ven = leer_ventas_para_match()
    msg("VENTAS filas: ", len(ven))

    msg("Ejecutando cruce RPE...")
    res = ejecutar_match(ven, com_key)

    res, bloque, log_df = armar_bloque_y_log(res)

    out_ventas = escribir_salida(res, bloque)
    out_log = escribir_log(log_df)

    msg("LISTO")
    msg("VENTAS salida: ", out_ventas)
    msg("LOG salida   : ", out_log)
    msg("Matches RPE_1: ", int(((res["RPE1"].notna()) & (res["RPE1"] != "")).sum()))
    msg("Pairs (RPE1+RPE2): ", int(((res["RPE2"].notna()) & (res["RPE2"] != "")).sum()))
    msg("EXC_01: ", int((res["exc_code"] == EXC_01_CODE).sum()))
    msg("EXC_02: ", int((res["exc_code"] == EXC_02_CODE).sum()))
    msg("EXC_03: ", int((res["exc_code"] == EXC_03_CODE).sum()))
    msg("EXC_04: ", int((res["exc_code"] == EXC_04_CODE).sum()))
    msg("Dictamen rojo sin RPE: ", int((res["dictamen_color_powerbi"] == "ROJO_SIN_RPE").sum()))
    msg("Dictamen azul grupo  : ", int((res["dictamen_color_powerbi"] == "AZUL_GRUPO_DUPLICADO").sum()))
    msg("Dictamen naranja exc : ", int((res["dictamen_color_powerbi"] == "NARANJA_EXCEPCION_REVISION_MANUAL").sum()))
    msg("Dictamen sin color OK: ", int((res["dictamen_color_powerbi"] == "SIN_COLOR_OK").sum()))
    msg("LOG filas: ", len(log_df))


if __name__ == "__main__":
    main()
